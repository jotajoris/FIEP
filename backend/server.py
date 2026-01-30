from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Body
from fastapi.security import HTTPBearer
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import asyncio
from pathlib import Path
import base64
import shutil
import io

# Configure logging FIRST - before any logger usage
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

from typing import List, Optional, Dict
import uuid
from datetime import datetime, timezone, timedelta
import openpyxl
import resend
import fitz  # PyMuPDF
import re
from pydantic import BaseModel

# OCR para PDFs escaneados
try:
    import pytesseract
    from PIL import Image
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False
    logger.warning("pytesseract ou PIL não instalados - OCR não disponível")

# Importar modelos do módulo models
from models import (
    UserRole, ItemStatus, User, UserCreate, LoginRequest, LoginResponse,
    ChangePasswordRequest, ResetPasswordRequest, ConfirmResetPasswordRequest,
    ReferenceItem, FonteCompra, Notificacao, NotaFiscalDoc,
    POItem, PurchaseOrder, PurchaseOrderCreate, ItemStatusUpdate, ItemFullUpdate,
    DashboardStats, AdminSummary, CommissionPayment, CommissionPaymentCreate,
    CommissionPaymentUpdate
)

from auth import (
    verify_password, get_password_hash, create_access_token,
    get_current_user, require_admin
)

import requests  # Para buscar CEP

# Importar routers modulares
from routes.auth_routes import router as auth_router
from routes.rastreio_routes import router as rastreio_router
from routes.notificacao_routes import router as notificacao_router
from routes.admin_routes import router as admin_router, init_admin_routes
from routes.backup_routes import router as backup_router, init_backup_routes
from routes.fornecedores_routes import router as fornecedores_router, init_fornecedores_routes
from routes.dashboard_routes import router as dashboard_router, init_dashboard_routes
from routes.estoque_routes import router as estoque_router, init_estoque_routes
from routes.limites_routes import router as limites_router, init_limites_routes

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Resend configuration
resend.api_key = os.environ.get('RESEND_API_KEY', '')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'onboarding@resend.dev')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Background task control
rastreio_task = None

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# Mapeamento de lotes para responsáveis
LOT_ASSIGNMENTS = {
    'Maria': list(range(1, 13)) + list(range(43, 54)),
    'Mateus': list(range(13, 21)) + list(range(54, 67)),
    'João': list(range(21, 32)) + list(range(67, 80)),
    'Mylena': list(range(80, 98)),
    'Fabio': list(range(32, 43))
}

# Criar reverse mapping (lote -> responsável)
LOT_TO_OWNER = {}
for owner, lots in LOT_ASSIGNMENTS.items():
    for lot in lots:
        LOT_TO_OWNER[lot] = owner

# Email to owner name mapping
EMAIL_TO_OWNER = {
    'maria.onsolucoes@gmail.com': 'Maria',
    'mylena.onsolucoes@gmail.com': 'Mylena',
    'fabioonsolucoes@gmail.com': 'Fabio'
}

# OCs excluídas do cálculo de comissão
EXCLUDED_OCS_FROM_COMMISSION = ['OC-2.118938', 'OC-2.118941']

# Health check endpoint (for Kubernetes)
@app.get("/health")
async def health_check():
    """Health check endpoint for Kubernetes deployment"""
    return {"status": "healthy", "service": "fiep-oc-backend"}

# Helper function
def get_responsible_by_lot(lot_number: int) -> str:
    return LOT_TO_OWNER.get(lot_number, "Não atribuído")


async def buscar_cep_por_endereco(endereco: str) -> Optional[str]:
    """
    Busca CEP pelo endereço usando a API ViaCEP.
    Retorna o CEP formatado (XXXXX-XXX) ou None se não encontrado.
    """
    import httpx
    
    if not endereco or len(endereco) < 10:
        return None
    
    endereco_upper = endereco.upper().strip()
    
    # Lista de cidades conhecidas e seus estados (foco na região de Curitiba)
    cidades_uf = {
        'CURITIBA': 'PR', 'SAO JOSE DOS PINHAIS': 'PR', 'PINHAIS': 'PR', 
        'COLOMBO': 'PR', 'ARAUCARIA': 'PR', 'CAMPO LARGO': 'PR',
        'PIRAQUARA': 'PR', 'ALMIRANTE TAMANDARE': 'PR', 'FAZENDA RIO GRANDE': 'PR',
        'LONDRINA': 'PR', 'MARINGA': 'PR', 'FOZ DO IGUACU': 'PR', 'CASCAVEL': 'PR',
        'SAO PAULO': 'SP', 'RIO DE JANEIRO': 'RJ', 'BELO HORIZONTE': 'MG',
        'PORTO ALEGRE': 'RS', 'FLORIANOPOLIS': 'SC', 'BRASILIA': 'DF',
    }
    
    # Tentar extrair UF do endereço
    uf = None
    cidade = None
    
    # Procurar UF explícita (ex: "- PR" ou "/PR")
    uf_match = re.search(r'[-/,\s]([A-Z]{2})[\s,]*$', endereco_upper)
    if uf_match:
        uf = uf_match.group(1)
    
    # Procurar cidade conhecida no endereço
    for cidade_nome, cidade_uf in cidades_uf.items():
        if cidade_nome in endereco_upper:
            cidade = cidade_nome
            if not uf:
                uf = cidade_uf
            break
    
    if not cidade or not uf:
        logger.info(f"Cidade/UF não identificada no endereço: {endereco}")
        return None
    
    # Extrair logradouro (primeira parte antes da vírgula)
    partes = endereco_upper.split(',')
    logradouro_completo = partes[0].strip() if partes else ''
    
    # Remover prefixos comuns e pegar palavras significativas
    logradouro = re.sub(
        r'^(RUA|AVENIDA|AV\.?|ALAMEDA|AL\.?|TRAVESSA|TV\.?|PRACA|PC\.?|ESTRADA|EST\.?|RODOVIA|ROD\.?)\s+', 
        '', 
        logradouro_completo
    )
    logradouro_busca = ' '.join(logradouro.split()[:2])  # Primeiras 2 palavras
    
    if not logradouro_busca:
        return None
    
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            # Normalizar para URL
            cidade_url = cidade.replace(' ', '%20')
            logradouro_url = logradouro_busca.replace(' ', '%20')
            
            url = f'https://viacep.com.br/ws/{uf}/{cidade_url}/{logradouro_url}/json/'
            logger.info(f"Buscando CEP: {url}")
            
            response = await client.get(url)
            
            if response.status_code == 200:
                data = response.json()
                if isinstance(data, list) and len(data) > 0:
                    cep = data[0].get('cep', '')
                    logger.info(f"CEP encontrado: {cep}")
                    return cep
    except Exception as e:
        logger.error(f"Erro ao buscar CEP: {e}")
    
    return None


async def reverter_uso_estoque(item: dict, po_id: str, numero_oc: str) -> dict:
    """
    Reverte o uso de estoque quando um item é movido de volta para pendente/cotado.
    
    Esta função:
    1. Verifica se o item usou estoque (via estoque_origem)
    2. Para cada fonte de estoque, decrementa a quantidade_usada_estoque
    3. Remove a entrada em estoque_usado_em da OC de origem
    4. Limpa os campos de estoque do item atual
    
    Retorna um dict com informações sobre o que foi revertido.
    """
    resultado = {
        'estoque_revertido': False,
        'quantidade_revertida': 0,
        'fontes_revertidas': []
    }
    
    estoque_origem = item.get('estoque_origem', [])
    if not estoque_origem:
        # Não usou estoque, apenas limpar campos por segurança
        item['quantidade_do_estoque'] = 0
        item['estoque_origem'] = []
        item['parcialmente_atendido_estoque'] = False
        item['atendido_por_estoque'] = False
        # Remover fonte de compra "ESTOQUE INTERNO" se existir
        if item.get('fontes_compra'):
            item['fontes_compra'] = [
                fc for fc in item['fontes_compra'] 
                if fc.get('fornecedor') != 'ESTOQUE INTERNO'
            ]
        return resultado
    
    logger.info(f"Revertendo uso de estoque para item no po_id={po_id}, estoque_origem={estoque_origem}")
    
    # Processar cada fonte de estoque
    for fonte in estoque_origem:
        numero_oc_origem = fonte.get('numero_oc')
        quantidade_usada = fonte.get('quantidade', 0)
        
        if not numero_oc_origem or quantidade_usada <= 0:
            continue
        
        # Encontrar a OC de origem
        po_origem = await db.purchase_orders.find_one(
            {"numero_oc": numero_oc_origem},
            {"_id": 0}
        )
        
        if not po_origem:
            logger.warning(f"OC de origem {numero_oc_origem} não encontrada para reverter estoque")
            continue
        
        # Encontrar o item com o mesmo código na OC de origem
        codigo_item = item.get('codigo_item')
        item_origem_atualizado = False
        
        for item_origem in po_origem.get('items', []):
            if item_origem.get('codigo_item') != codigo_item:
                continue
            
            # Decrementar quantidade_usada_estoque
            qtd_usada_atual = item_origem.get('quantidade_usada_estoque', 0)
            nova_qtd_usada = max(0, qtd_usada_atual - quantidade_usada)
            item_origem['quantidade_usada_estoque'] = nova_qtd_usada
            
            # Remover entrada em estoque_usado_em que corresponde a esta OC destino
            estoque_usado_em = item_origem.get('estoque_usado_em', [])
            item_origem['estoque_usado_em'] = [
                uso for uso in estoque_usado_em 
                if uso.get('po_id') != po_id
            ]
            
            item_origem_atualizado = True
            logger.info(f"Revertido {quantidade_usada} UN do estoque na OC {numero_oc_origem}, nova qtd_usada={nova_qtd_usada}")
            break
        
        if item_origem_atualizado:
            # Salvar a OC de origem atualizada
            await db.purchase_orders.update_one(
                {"id": po_origem['id']},
                {"$set": {"items": po_origem['items']}}
            )
            
            resultado['fontes_revertidas'].append({
                'numero_oc': numero_oc_origem,
                'quantidade': quantidade_usada
            })
            resultado['quantidade_revertida'] += quantidade_usada
    
    # Limpar campos de estoque do item atual
    item['quantidade_do_estoque'] = 0
    item['estoque_origem'] = []
    item['parcialmente_atendido_estoque'] = False
    item['atendido_por_estoque'] = False
    item['preco_estoque_unitario'] = None
    
    # Remover fonte de compra "ESTOQUE INTERNO" se existir
    if item.get('fontes_compra'):
        item['fontes_compra'] = [
            fc for fc in item['fontes_compra'] 
            if fc.get('fornecedor') != 'ESTOQUE INTERNO'
        ]
    
    resultado['estoque_revertido'] = len(resultado['fontes_revertidas']) > 0
    
    return resultado


def atualizar_data_compra(item: dict, novo_status: str) -> None:
    """
    Atualiza a data de compra do item automaticamente:
    - Se mudar para comprado/em_separacao/em_transito/entregue: salva a data atual (se não existir)
    - Se voltar para pendente/cotado: remove a data de compra
    """
    status_comprado_ou_adiante = ['comprado', 'em_separacao', 'em_transito', 'entregue']
    status_antes_compra = ['pendente', 'cotado']
    
    if novo_status in status_comprado_ou_adiante:
        # Se ainda não tem data de compra, salva a data atual
        if not item.get('data_compra'):
            item['data_compra'] = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    elif novo_status in status_antes_compra:
        # Remove a data de compra se voltar para pendente ou cotado
        item['data_compra'] = None


def calcular_lucro_item(item: dict) -> None:
    """
    Calcula o lucro líquido do item usando a QUANTIDADE NECESSÁRIA (não a quantidade comprada).
    
    O lucro é calculado apenas sobre o que será vendido (quantidade necessária da OC),
    mesmo que tenha sido comprado mais (excedente vai para estoque).
    
    Fórmula:
    - Receita = preço_venda × quantidade_necessária
    - Custo = custo_unitário_médio × quantidade_necessária (ou preço_compra × quantidade)
    - Impostos = 11% da receita
    - Lucro = Receita - Custo - Frete_compra_proporcional - Impostos - Frete_envio
    """
    preco_venda = item.get('preco_venda')
    quantidade_necessaria = item.get('quantidade', 0)
    
    if not preco_venda or not quantidade_necessaria:
        return
    
    fontes = item.get('fontes_compra', [])
    receita_total = preco_venda * quantidade_necessaria
    impostos = receita_total * 0.11
    item['imposto'] = round(impostos, 2)
    
    if fontes and len(fontes) > 0:
        # Calcular custo médio ponderado das fontes de compra
        total_quantidade_comprada = sum(fc.get('quantidade', 0) for fc in fontes)
        total_custo_fontes = sum(fc.get('quantidade', 0) * fc.get('preco_unitario', 0) for fc in fontes)
        total_frete_fontes = sum(fc.get('frete', 0) for fc in fontes)
        
        if total_quantidade_comprada > 0:
            # Custo unitário médio ponderado
            custo_unitario_medio = total_custo_fontes / total_quantidade_comprada
            # Custo apenas da quantidade necessária
            custo_para_venda = custo_unitario_medio * quantidade_necessaria
            # Frete proporcional à quantidade necessária
            frete_proporcional = (total_frete_fontes / total_quantidade_comprada) * quantidade_necessaria
        else:
            custo_para_venda = 0
            frete_proporcional = 0
        
        frete_envio = item.get('frete_envio', 0) or 0
        item['lucro_liquido'] = round(receita_total - custo_para_venda - frete_proporcional - impostos - frete_envio, 2)
    elif item.get('preco_compra') is not None:
        # Cálculo tradicional (sem fontes de compra)
        custo_total = item['preco_compra'] * quantidade_necessaria
        frete_compra = item.get('frete_compra', 0) or 0
        frete_envio = item.get('frete_envio', 0) or 0
        item['lucro_liquido'] = round(receita_total - custo_total - frete_compra - impostos - frete_envio, 2)


def extract_text_with_ocr(pdf_bytes: bytes) -> str:
    """Extrair texto de PDF usando OCR (para PDFs escaneados)"""
    if not OCR_AVAILABLE:
        logger.warning("OCR não disponível - pytesseract não instalado")
        return ""
    
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        full_text = ""
        
        for page_num, page in enumerate(doc):
            # Converter página para imagem com alta resolução
            mat = fitz.Matrix(2, 2)  # 2x zoom para melhor qualidade
            pix = page.get_pixmap(matrix=mat)
            
            # Converter para PIL Image
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            
            # Usar OCR para extrair texto (português + inglês)
            try:
                page_text = pytesseract.image_to_string(img, lang='por+eng')
                full_text += page_text + "\n"
                logger.info(f"OCR página {page_num + 1}: {len(page_text)} caracteres extraídos")
            except Exception as ocr_error:
                logger.warning(f"Erro OCR na página {page_num + 1}: {ocr_error}")
        
        doc.close()
        return full_text
        
    except Exception as e:
        logger.error(f"Erro ao extrair texto com OCR: {e}")
        return ""


def buscar_cep_por_endereco(endereco: str) -> Optional[str]:
    """
    Buscar CEP automaticamente pelo endereço usando APIs gratuitas.
    Considera o número do endereço para selecionar o CEP correto.
    """
    if not endereco or len(endereco) < 10:
        return None
    
    # Limpar e formatar endereço
    endereco_limpo = endereco.strip()
    
    # Extrair componentes do endereço (LOGRADOURO, NUMERO, BAIRRO, CIDADE)
    # Formato típico: "AVENIDA AVIAÇÃO, 1851, VILA NOVA, APUCARANA"
    partes = [p.strip() for p in endereco_limpo.split(',')]
    
    logradouro = ""
    numero = 0
    cidade = ""
    
    if len(partes) >= 3:
        logradouro = partes[0]
        cidade = partes[-1].strip()
        # Se tem estado (ex: "APUCARANA - PR"), extrair
        if ' - ' in cidade:
            cidade = cidade.split(' - ')[0].strip()
        
        # Tentar extrair número do endereço
        for parte in partes[1:-1]:
            try:
                numero = int(parte.strip())
                break
            except ValueError:
                pass
    else:
        logradouro = endereco_limpo
    
    # Método 1: API ViaCEP (busca por endereço) - com seleção por número
    try:
        # ViaCEP aceita busca: UF/Cidade/Logradouro
        # Tentar com PR (Paraná) que é o estado da FIEP
        estados = ['PR', 'SC', 'RS', 'SP']
        
        # Tentar primeiro com logradouro completo, depois sem prefixo
        logradouros_tentar = [logradouro]
        
        # Adicionar versão sem prefixo como fallback
        logradouro_sem_prefixo = logradouro
        for prefixo in ['AVENIDA ', 'AV ', 'AV. ', 'RUA ', 'R ', 'R. ', 'TRAVESSA ', 'TV ', 'ALAMEDA ', 'AL ']:
            if logradouro_sem_prefixo.upper().startswith(prefixo):
                logradouro_sem_prefixo = logradouro_sem_prefixo[len(prefixo):]
                break
        if logradouro_sem_prefixo != logradouro:
            logradouros_tentar.append(logradouro_sem_prefixo)
        
        for logradouro_busca in logradouros_tentar:
            for uf in estados:
                url = f"https://viacep.com.br/ws/{uf}/{cidade}/{logradouro_busca}/json/"
                response = requests.get(url, timeout=5)
                
                if response.status_code == 200:
                    data = response.json()
                    if isinstance(data, list) and len(data) > 0:
                        # Se tem número e múltiplos resultados, escolher o correto pelo range
                        if numero > 0 and len(data) > 1:
                            for item in data:
                                complemento = item.get('complemento', '')
                                # Tentar parsear o range (ex: "de 1101/1102 ao fim")
                                if 'até' in complemento.lower():
                                    # Ex: "até 1099/1100"
                                    try:
                                        nums = re.findall(r'\d+', complemento)
                                        if nums:
                                            limite = int(nums[0])
                                            if numero <= limite:
                                                cep = item.get('cep', '').replace('-', '')
                                                if cep and len(cep) == 8:
                                                    logger.info(f"CEP encontrado via ViaCEP (range ≤{limite}): {cep}")
                                                    return f"{cep[:5]}-{cep[5:]}"
                                    except:
                                        pass
                                elif 'de ' in complemento.lower() and 'ao fim' in complemento.lower():
                                    # Ex: "de 1101/1102 ao fim"
                                    try:
                                        nums = re.findall(r'\d+', complemento)
                                        if nums:
                                            inicio = int(nums[0])
                                            if numero >= inicio:
                                                cep = item.get('cep', '').replace('-', '')
                                                if cep and len(cep) == 8:
                                                    logger.info(f"CEP encontrado via ViaCEP (range ≥{inicio}): {cep}")
                                                    return f"{cep[:5]}-{cep[5:]}"
                                    except:
                                        pass
                        
                        # Se não conseguiu pelo range, usar o primeiro
                        cep = data[0].get('cep', '').replace('-', '')
                        if cep and len(cep) == 8:
                            logger.info(f"CEP encontrado via ViaCEP: {cep} para {endereco_limpo}")
                            return f"{cep[:5]}-{cep[5:]}"
    except Exception as e:
        logger.warning(f"Erro ao buscar CEP via ViaCEP: {e}")
    
    # Método 2: API Nominatim (OpenStreetMap) - Geocodificação
    try:
        # Adicionar ", Brasil" para melhor precisão
        query = f"{endereco_limpo}, Brasil"
        url = f"https://nominatim.openstreetmap.org/search"
        params = {
            'q': query,
            'format': 'json',
            'addressdetails': 1,
            'limit': 1,
            'countrycodes': 'br'
        }
        headers = {'User-Agent': 'FIEP-OC-System/1.0'}
        
        response = requests.get(url, params=params, headers=headers, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            if data and len(data) > 0:
                address = data[0].get('address', {})
                cep = address.get('postcode', '')
                if cep:
                    cep = cep.replace('-', '').replace('.', '')
                    if len(cep) == 8:
                        logger.info(f"CEP encontrado via Nominatim: {cep} para {endereco_limpo}")
                        return f"{cep[:5]}-{cep[5:]}"
    except Exception as e:
        logger.warning(f"Erro ao buscar CEP via Nominatim: {e}")
    
    logger.info(f"CEP não encontrado para: {endereco_limpo}")
    return None


def extract_oc_from_pdf(pdf_bytes: bytes) -> dict:
    """Extrair dados de OC de um PDF"""
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        full_text = ""
        
        for page in doc:
            full_text += page.get_text()
        
        doc.close()
        
        # Se o PDF não tem texto (é escaneado), usar OCR
        if not full_text.strip() and OCR_AVAILABLE:
            logger.info("PDF parece ser escaneado - tentando OCR...")
            full_text = extract_text_with_ocr(pdf_bytes)
            if full_text.strip():
                logger.info(f"OCR extraiu {len(full_text)} caracteres")
            else:
                logger.warning("OCR não conseguiu extrair texto")
        
        # Extrair número da OC - procurar por padrões como OC-X.XXXXXX
        oc_patterns = [
            r'OC[- ]?(\d+[\.\d]+)',  # OC-2.121437 ou OC 2.121437
            r'Ordem de Compra[:\s]+(\d+[\.\d]+)',
            r'N[úu]mero[:\s]+(\d+[\.\d]+)'
        ]
        
        numero_oc = None
        for pattern in oc_patterns:
            oc_match = re.search(pattern, full_text, re.IGNORECASE)
            if oc_match:
                numero_oc = f"OC-{oc_match.group(1)}"
                break
        
        if not numero_oc:
            numero_oc = "OC-" + str(uuid.uuid4())[:8]
        
        # Extrair endereço de entrega
        endereco_patterns = [
            r'Endere[çc]o de Entrega[:\s]*(.*?)(?:\n\n|Linha|Item)',
            r'Local de Entrega[:\s]*(.*?)(?:\n\n|Linha|Item)',
            r'Entregar em[:\s]*(.*?)(?:\n\n|Linha|Item)'
        ]
        
        endereco_entrega = ""
        for pattern in endereco_patterns:
            endereco_match = re.search(pattern, full_text, re.IGNORECASE | re.DOTALL)
            if endereco_match:
                endereco_entrega = endereco_match.group(1).strip()
                endereco_entrega = ' '.join(endereco_entrega.split())
                break
        
        # Extrair região de entrega
        regiao = ""
        regiao_match = re.search(r'Regi[ãa]o[:\s]*(.*?)(?:\n|$)', full_text, re.IGNORECASE)
        if regiao_match:
            regiao = regiao_match.group(1).strip()
        
        # Extrair CNPJ do requisitante/cliente (FIEP/SESI)
        # CNPJ padrão: XX.XXX.XXX/XXXX-XX
        cnpj_requisitante = ""
        cnpj_patterns = [
            r'CNPJ[:\s]*(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})',
            r'(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})'
        ]
        
        for pattern in cnpj_patterns:
            cnpj_matches = re.findall(pattern, full_text)
            if cnpj_matches:
                # O primeiro CNPJ geralmente é do cliente/requisitante
                # O CNPJ da ON (fornecedor) é 46.663.556/0001-69, então pegamos outro
                for cnpj in cnpj_matches:
                    if cnpj != "46.663.556/0001-69":  # Ignorar CNPJ do fornecedor (ON)
                        cnpj_requisitante = cnpj
                        break
                break
        
        # Extrair Data de Entrega (formato DD/MM/YYYY)
        # A data pode estar em diferentes formatos e locais no PDF
        data_entrega = None
        data_patterns = [
            r'Data de Entrega[:\s]*(\d{2}/\d{2}/\d{4})',
            r'Data Entrega[:\s]*(\d{2}/\d{2}/\d{4})',
            r'Entrega[:\s]*(\d{2}/\d{2}/\d{4})',
            r'Prazo de Entrega[:\s]*(\d{2}/\d{2}/\d{4})',
            r'Dt\.\s*Entrega[:\s]*(\d{2}/\d{2}/\d{4})'
        ]
        
        for pattern in data_patterns:
            data_match = re.search(pattern, full_text, re.IGNORECASE)
            if data_match:
                try:
                    # Converter de DD/MM/YYYY para ISO format
                    data_str = data_match.group(1)
                    dia, mes, ano = data_str.split('/')
                    data_entrega = f"{ano}-{mes}-{dia}"  # ISO format YYYY-MM-DD
                    break
                except:
                    pass
        
        # Se não encontrou nos padrões acima, buscar na tabela de itens
        # O formato é geralmente: "requisição DD/MM/YYYY" no final de cada linha de item
        if not data_entrega:
            # Buscar todas as datas no formato DD/MM/YYYY após um número de requisição
            date_after_req = re.findall(r'\d{1,2}\.\d{2}\.\d{6,}\s*(\d{2}/\d{2}/\d{4})', full_text)
            if date_after_req:
                try:
                    data_str = date_after_req[0]  # Pegar a primeira data encontrada
                    dia, mes, ano = data_str.split('/')
                    # Validar se é uma data futura ou recente (não muito antiga)
                    if int(ano) >= 2024:
                        data_entrega = f"{ano}-{mes}-{dia}"
                except:
                    pass
        
        # ========== PARSER MELHORADO PARA PDFS FIEP ==========
        items = []
        seen_items = set()
        lines = full_text.split('\n')
        
        # Códigos de produto FIEP começam com 0 ou 1 (ex: 089847, 114720)
        # Códigos NCM começam com 8 ou 9 (ex: 853890, 903180) - ignorar
        
        for i, line in enumerate(lines):
            line_stripped = line.strip()
            
            # Procurar código de 6 dígitos que começa com 0 ou 1
            if re.match(r'^([01]\d{5})$', line_stripped):
                codigo = line_stripped
                
                # Verificar se linha anterior é número de linha (1-50)
                if i > 0:
                    prev = lines[i-1].strip()
                    if re.match(r'^\d{1,2}$', prev):
                        try:
                            linha_num = int(prev)
                            if 1 <= linha_num <= 100:
                                # Procurar quantidade nas próximas linhas (aumentado para 40 linhas para descrições longas)
                                quantidade = 0
                                unidade = "UN"
                                descricao_parts = []
                                preco_pdf = None  # Inicializar aqui
                                ncm_item = None  # NCM do item
                                
                                for j in range(i+1, min(i+40, len(lines))):
                                    check_line = lines[j].strip()
                                    
                                    # Se encontrar outro código de produto, parar
                                    if re.match(r'^([01]\d{5})$', check_line):
                                        break
                                    
                                    # Capturar NCM completo (8 dígitos começando com 1-9)
                                    if re.match(r'^[1-9]\d{7}$', check_line) and ncm_item is None:
                                        ncm_item = check_line
                                        continue
                                    
                                    # Capturar NCM dividido (6 dígitos + 2 dígitos na próxima linha)
                                    if re.match(r'^[1-9]\d{5}$', check_line) and ncm_item is None:
                                        if j+1 < len(lines):
                                            next_line = lines[j+1].strip()
                                            if re.match(r'^\d{2}$', next_line):
                                                ncm_item = check_line + next_line
                                                continue
                                    
                                    # Coletar descrição (até encontrar quantidade)
                                    if len(check_line) > 2 and not re.match(r'^[\d.,]+$', check_line):
                                        if check_line not in ['UN', 'UND', 'UNID', 'KG', 'PC', 'M', 'L', 'CX', 'PAR', 'KIT']:
                                            if 'Descritivo Completo' not in check_line and 'CFOP' not in check_line:
                                                # Não incluir NCM na descrição (6 ou 8 dígitos começando com 1-9)
                                                if not re.match(r'^[1-9]\d{5,7}$', check_line):
                                                    descricao_parts.append(check_line)
                                    
                                    # Procurar quantidade (número isolado seguido de unidade)
                                    qty_match = re.match(r'^(\d+)$', check_line)
                                    if qty_match and quantidade == 0:
                                        qty = int(qty_match.group(1))
                                        # Verificar se próxima linha é unidade
                                        if j+1 < len(lines):
                                            unit_line = lines[j+1].strip().upper()
                                            # Lista expandida de unidades aceitas
                                            valid_units = ['UN', 'UND', 'UNID', 'KG', 'PC', 'PÇA', 'PÇ', 'PCA', 'M', 'L', 'CX', 'PAR', 'PCT', 'KIT', 'JG', 'JOGO', 'RL', 'ROLO', 'MT', 'METRO', 'CT', 'CEN', 'CENTO', 'MILHEIRO', 'MIL']
                                            if unit_line in valid_units:
                                                quantidade = qty
                                                # Normalizar unidades
                                                if unit_line in ['UND', 'UNID', 'PÇA', 'PÇ', 'PCA', 'PC']:
                                                    unidade = 'UN'
                                                elif unit_line in ['MT', 'METRO']:
                                                    unidade = 'M'
                                                elif unit_line in ['JOGO']:
                                                    unidade = 'JG'
                                                elif unit_line in ['ROLO']:
                                                    unidade = 'RL'
                                                else:
                                                    unidade = unit_line
                                                
                                                # EXTRAIR PREÇO: Após a unidade vem o preço unitário (formato: 518,95)
                                                # Estrutura: QTD -> UN -> PREÇO_UNITARIO -> TOTAL
                                                preco_pdf = None
                                                if j+2 < len(lines):
                                                    preco_line = lines[j+2].strip()
                                                    # Preço no formato brasileiro: 518,95 ou 1.234,56
                                                    preco_match = re.match(r'^(\d{1,3}(?:\.\d{3})*,\d{2})$', preco_line)
                                                    if preco_match:
                                                        # Converter formato BR para float
                                                        preco_str = preco_match.group(1).replace('.', '').replace(',', '.')
                                                        try:
                                                            preco_pdf = float(preco_str)
                                                        except:
                                                            pass
                                                
                                                break
                                
                                if quantidade > 0:
                                    key = f"{linha_num}-{codigo}"
                                    if key not in seen_items:
                                        seen_items.add(key)
                                        # Pegar descrição completa (sem limite de partes)
                                        descricao = ' '.join(descricao_parts) if descricao_parts else f"Item {codigo}"
                                        item_data = {
                                            "codigo_item": codigo,
                                            "quantidade": quantidade,
                                            "descricao": descricao,  # Descrição completa sem truncar
                                            "unidade": unidade,
                                            "endereco_entrega": endereco_entrega,
                                            "regiao": regiao
                                        }
                                        # Adicionar preço extraído do PDF se encontrado
                                        try:
                                            if preco_pdf is not None:
                                                item_data["preco_venda_pdf"] = preco_pdf
                                        except NameError:
                                            pass
                                        # Adicionar NCM se encontrado
                                        if ncm_item:
                                            item_data["ncm"] = ncm_item
                                        items.append(item_data)
                        except ValueError:
                            pass
        
        # Método 2 (fallback): Se NENHUM item encontrado, tentar padrão mais genérico
        if len(items) == 0:
            items = []
            seen_codes = set()
            
            for i, line in enumerate(lines):
                # Procurar códigos que começam com 0 ou 1
                codigo_match = re.search(r'\b([01]\d{5})\b', line)
                if codigo_match:
                    codigo = codigo_match.group(1)
                    
                    if codigo in seen_codes:
                        continue
                    
                    # Procurar quantidade
                    quantidade = 0
                    unidade = "UN"
                    
                    qty_match = re.search(r'\b(\d+)\s*(UN|UND|UNID|KG|PC|M|L|CX|KIT|CT|CEN|CENTO|PAR|PCT)\b', line, re.IGNORECASE)
                    if qty_match:
                        quantidade = int(qty_match.group(1))
                        unidade = qty_match.group(2).upper()
                    else:
                        for j in range(i+1, min(i+8, len(lines))):
                            qty_match = re.search(r'\b(\d+)\s*(UN|UND|UNID|KG|PC|M|L|CX|KIT|CT|CEN|CENTO|PAR|PCT)\b', lines[j], re.IGNORECASE)
                            if qty_match:
                                quantidade = int(qty_match.group(1))
                                unidade = qty_match.group(2).upper()
                                break
                    
                    if quantidade > 0:
                        seen_codes.add(codigo)
                        items.append({
                            "codigo_item": codigo,
                            "quantidade": quantidade,
                            "descricao": f"Item {codigo}",
                            "unidade": unidade,
                            "endereco_entrega": endereco_entrega,
                            "regiao": regiao
                        })
        
        return {
            "numero_oc": numero_oc,
            "items": items,
            "endereco_entrega": endereco_entrega,
            "regiao": regiao,
            "cnpj_requisitante": cnpj_requisitante,
            "data_entrega": data_entrega
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao processar PDF: {str(e)}")


async def send_password_reset_email(email: str, reset_token: str):
    """Envia email com link de reset de senha"""
    reset_link = f"{os.environ.get('FRONTEND_URL', 'http://localhost:3000')}/reset-password?token={reset_token}"
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
            .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
            .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; text-align: center; border-radius: 10px 10px 0 0; }}
            .content {{ background: #f9f9f9; padding: 30px; border-radius: 0 0 10px 10px; }}
            .button {{ display: inline-block; padding: 12px 30px; background: #667eea; color: white; text-decoration: none; border-radius: 5px; margin: 20px 0; }}
            .footer {{ text-align: center; margin-top: 20px; color: #666; font-size: 12px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>FIEP | Gestão OC</h1>
            </div>
            <div class="content">
                <h2>Troca de Senha</h2>
                <p>Olá,</p>
                <p>Você recebeu acesso à plataforma de Gestão de Ordens de Compra FIEP.</p>
                <p>Sua senha temporária é: <strong>on123456</strong></p>
                <p>Por favor, clique no botão abaixo para alterar sua senha:</p>
                <a href="{reset_link}" class="button">Alterar Senha</a>
                <p><small>Ou copie e cole este link no navegador:<br>{reset_link}</small></p>
                <p>Este link expira em 24 horas.</p>
            </div>
            <div class="footer">
                <p>Se você não solicitou este email, por favor ignore.</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    try:
        params = {
            "from": SENDER_EMAIL,
            "to": [email],
            "subject": "Bem-vindo à Plataforma FIEP - Troca de Senha",
            "html": html_content
        }
        await asyncio.to_thread(resend.Emails.send, params)
        return True
    except Exception as e:
        logging.error(f"Erro ao enviar email: {str(e)}")
        return False

# ===== ROTAS DE AUTENTICAÇÃO MOVIDAS PARA routes/auth_routes.py =====
# As rotas de autenticação foram refatoradas para o módulo routes/auth_routes.py

# Endpoint de diagnóstico SIMPLES - mostra os itens do usuário e o problema
@api_router.get("/debug/meus-itens")
async def debug_meus_itens(current_user: dict = Depends(get_current_user)):
    """
    Endpoint de diagnóstico simples - acesse logado como Maria ou Fabio
    URL: https://onlicitacoes.com/api/debug/meus-itens
    """
    user_email = current_user.get('sub', '')
    user_role = current_user.get('role', '')
    user_owner_name = current_user.get('owner_name') or ''
    
    # Buscar todas as OCs
    all_pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(1000)
    
    # Encontrar itens que DEVERIAM ser do usuário
    meus_itens = []
    for po in all_pos:
        for idx, item in enumerate(po.get('items', [])):
            item_resp = item.get('responsavel', '')
            # Verificar se o nome bate (case-insensitive)
            if item_resp.strip().upper() == user_owner_name.strip().upper():
                meus_itens.append({
                    "po_id": po['id'],
                    "numero_oc": po['numero_oc'],
                    "item_index": idx,
                    "codigo_item": item.get('codigo_item'),
                    "responsavel_no_item": item_resp,
                    "status": item.get('status'),
                    "comparacao": {
                        "item_responsavel_raw": item_resp,
                        "item_responsavel_upper": item_resp.strip().upper(),
                        "user_owner_name_raw": user_owner_name,
                        "user_owner_name_upper": user_owner_name.strip().upper(),
                        "match": item_resp.strip().upper() == user_owner_name.strip().upper()
                    }
                })
    
    return {
        "usuario_logado": {
            "email": user_email,
            "role": user_role,
            "owner_name": user_owner_name,
            "owner_name_upper": user_owner_name.strip().upper() if user_owner_name else None,
            "owner_name_bytes": [ord(c) for c in user_owner_name] if user_owner_name else []
        },
        "total_itens_encontrados": len(meus_itens),
        "primeiros_5_itens": meus_itens[:5],
        "mensagem": "Se 'total_itens_encontrados' é 0 e você é Maria/Fabio, o owner_name não está correto no seu token/usuário"
    }

# Endpoint de diagnóstico para debug do bug de permissão
@api_router.get("/debug/permission/{po_id}/{item_index}")
async def debug_permission(po_id: str, item_index: int, current_user: dict = Depends(get_current_user)):
    """Endpoint de diagnóstico para verificar problema de permissão"""
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        return {"error": "OC não encontrada"}
    
    if item_index < 0 or item_index >= len(po['items']):
        return {"error": "Índice inválido"}
    
    item = po['items'][item_index]
    
    item_responsavel_raw = item.get('responsavel') or ''
    user_owner_name_raw = current_user.get('owner_name') or ''
    item_responsavel = item_responsavel_raw.strip().upper()
    user_owner_name = user_owner_name_raw.strip().upper()
    user_role = current_user.get('role', '')
    
    return {
        "user_info": {
            "email": current_user.get('sub'),
            "role": user_role,
            "owner_name_raw": user_owner_name_raw,
            "owner_name_normalized": user_owner_name,
            "owner_name_bytes": [ord(c) for c in user_owner_name_raw]
        },
        "item_info": {
            "codigo_item": item.get('codigo_item'),
            "responsavel_raw": item_responsavel_raw,
            "responsavel_normalized": item_responsavel,
            "responsavel_bytes": [ord(c) for c in item_responsavel_raw],
            "status": item.get('status')
        },
        "permission_check": {
            "is_admin": user_role == 'admin',
            "names_match": item_responsavel == user_owner_name,
            "would_allow": user_role == 'admin' or item_responsavel == user_owner_name
        }
    }

# Routes
@api_router.get("/")
async def root():
    return {"message": "Sistema de Gestão de Ordens de Compra FIEP"}

# ENDPOINT DE VERIFICAÇÃO DE VERSÃO - USE PARA CONFIRMAR QUE O DEPLOY FOI FEITO
@api_router.get("/version")
async def get_version():
    return {
        "version": "2.2.0",
        "deploy_date": "2025-01-12",
        "fix": "Checkbox carrinho, mover em lote para Comprado, campo observação",
        "status": "OK"
    }

# ENDPOINT DE DEBUG DO UPDATE - MOSTRA EXATAMENTE O QUE ACONTECE
@api_router.patch("/debug-update/{po_id}/{item_index}")
async def debug_update(
    po_id: str, 
    item_index: int, 
    update: ItemStatusUpdate, 
    current_user: dict = Depends(get_current_user)
):
    """Debug do update - retorna detalhes do que aconteceu"""
    debug_info = {
        "user": current_user.get('sub'),
        "role": current_user.get('role'),
        "po_id": po_id,
        "item_index": item_index,
        "update_recebido": {
            "status": update.status,
            "preco_venda": update.preco_venda,
            "fontes_compra": len(update.fontes_compra) if update.fontes_compra else 0
        }
    }
    
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        debug_info["erro"] = "OC não encontrada"
        return debug_info
    
    if item_index < 0 or item_index >= len(po['items']):
        debug_info["erro"] = f"Índice inválido. Total items: {len(po['items'])}"
        return debug_info
    
    item = po['items'][item_index]
    
    debug_info["item_antes"] = {
        "codigo": item.get('codigo_item'),
        "preco_compra": item.get('preco_compra'),
        "preco_venda": item.get('preco_venda'),
        "fontes_compra": len(item.get('fontes_compra', []))
    }
    
    # APLICAR TODAS AS ALTERAÇÕES SEM RESTRIÇÃO
    item['status'] = update.status
    atualizar_data_compra(item, update.status)  # Atualiza data de compra automaticamente
    
    if update.fontes_compra is not None:
        item['fontes_compra'] = [fc.model_dump() for fc in update.fontes_compra]
        total_custo = sum(fc.quantidade * fc.preco_unitario for fc in update.fontes_compra)
        total_frete = sum(fc.frete for fc in update.fontes_compra)
        total_qtd = sum(fc.quantidade for fc in update.fontes_compra)
        if total_qtd > 0:
            item['preco_compra'] = round(total_custo / total_qtd, 2)
        item['frete_compra'] = total_frete
    
    if update.preco_venda is not None:
        item['preco_venda'] = update.preco_venda
    
    if update.imposto is not None:
        item['imposto'] = update.imposto
    
    if update.frete_envio is not None:
        item['frete_envio'] = update.frete_envio
    
    debug_info["item_depois"] = {
        "codigo": item.get('codigo_item'),
        "preco_compra": item.get('preco_compra'),
        "preco_venda": item.get('preco_venda'),
        "fontes_compra": len(item.get('fontes_compra', []))
    }
    
    # Salvar
    result = await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": po['items']}}
    )
    
    debug_info["mongodb"] = {
        "matched": result.matched_count,
        "modified": result.modified_count
    }
    
    # Verificar se salvou
    po_verificar = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    item_verificar = po_verificar['items'][item_index]
    
    debug_info["item_verificado"] = {
        "preco_compra": item_verificar.get('preco_compra'),
        "preco_venda": item_verificar.get('preco_venda'),
        "fontes_compra": len(item_verificar.get('fontes_compra', []))
    }
    
    debug_info["sucesso"] = True
    return debug_info

# ENDPOINT DE TESTE DIRETO - PARA DEBUG DEFINITIVO
@api_router.post("/test-update-direto/{po_id}/{item_index}")
async def test_update_direto(po_id: str, item_index: int, current_user: dict = Depends(get_current_user)):
    """
    Teste direto de update - faz uma edição simples e retorna o resultado
    """
    try:
        # 1. Buscar a OC
        po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
        
        if not po:
            return {"erro": "OC não encontrada", "po_id": po_id}
        
        if item_index < 0 or item_index >= len(po['items']):
            return {"erro": "Índice inválido", "item_index": item_index, "total": len(po['items'])}
        
        # 2. Pegar o item
        item = po['items'][item_index]
        valor_antigo = item.get('preco_compra')
        
        # 3. Fazer uma edição de teste
        item['preco_compra'] = 12345.67
        item['_teste_update'] = "FUNCIONOU"
        
        # 4. Salvar
        result = await db.purchase_orders.update_one(
            {"id": po_id},
            {"$set": {"items": po['items']}}
        )
        
        # 5. Buscar novamente para confirmar
        po_depois = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
        item_depois = po_depois['items'][item_index]
        
        return {
            "sucesso": True,
            "usuario": current_user.get('sub'),
            "role": current_user.get('role'),
            "valor_antigo": valor_antigo,
            "valor_novo": item_depois.get('preco_compra'),
            "teste_campo": item_depois.get('_teste_update'),
            "mongodb_matched": result.matched_count,
            "mongodb_modified": result.modified_count
        }
    except Exception as e:
        return {"erro": str(e)}

# ENDPOINT DE TESTE DE EDIÇÃO - PARA DEBUG
@api_router.post("/test-edit/{po_id}/{item_index}")
async def test_edit(po_id: str, item_index: int, current_user: dict = Depends(get_current_user)):
    """Endpoint de teste para verificar se edição funciona"""
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        return {"error": "OC não encontrada", "po_id": po_id}
    
    if item_index < 0 or item_index >= len(po['items']):
        return {"error": "Índice inválido", "item_index": item_index, "total_items": len(po['items'])}
    
    item = po['items'][item_index]
    
    # Tentar atualizar
    try:
        item['_test_edit'] = "OK"
        await db.purchase_orders.update_one(
            {"id": po_id},
            {"$set": {"items": po['items']}}
        )
        return {
            "success": True,
            "message": "Edição funcionou!",
            "user": current_user.get('sub'),
            "role": current_user.get('role'),
            "item_codigo": item.get('codigo_item')
        }
    except Exception as e:
        return {"error": str(e)}

@api_router.post("/reference-items/seed")
async def seed_reference_items(current_user: dict = Depends(require_admin)):
    """Popula banco com itens de referência do Excel (ADMIN ONLY)"""
    try:
        file_path = Path("/app/items_data.xlsx")
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="Arquivo Excel não encontrado")
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        items = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            lote_str = str(row[0]) if row[0] else ""
            codigo_item = str(row[12]) if row[12] else ""
            
            if not lote_str or not codigo_item:
                continue
            
            try:
                lot_number = int(lote_str.replace("Lote", "").replace("lote", "").strip())
            except:
                continue
            
            responsavel = get_responsible_by_lot(lot_number)
            
            # Extrair preço unitário (coluna F - index 5)
            preco_venda = None
            if row[5] is not None:
                try:
                    preco_venda = float(row[5])
                except:
                    pass
            
            item = ReferenceItem(
                lote=lote_str,
                lot_number=lot_number,
                regiao=str(row[1]) if row[1] else "",
                descricao=str(row[2]) if row[2] else "",
                unidade=str(row[3]) if row[3] else "",
                marca_modelo=str(row[11]) if row[11] else "",
                codigo_item=codigo_item,
                responsavel=responsavel,
                preco_venda_unitario=preco_venda
            )
            items.append(item.model_dump())
        
        await db.reference_items.delete_many({})
        
        if items:
            for item in items:
                item['created_at'] = item['created_at'].isoformat()
            await db.reference_items.insert_many(items)
        
        return {"message": f"{len(items)} itens de referência carregados com sucesso"}
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao carregar itens: {str(e)}")

@api_router.get("/reference-items", response_model=List[ReferenceItem])
async def get_reference_items(codigo: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if codigo:
        query["codigo_item"] = codigo
    
    items = await db.reference_items.find(query, {"_id": 0}).to_list(5000)
    for item in items:
        if isinstance(item['created_at'], str):
            item['created_at'] = datetime.fromisoformat(item['created_at'])
    return items


@api_router.get("/items/historico-cotacoes")
async def get_historico_cotacoes(
    codigo_item: Optional[str] = None,
    descricao: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Buscar histórico de cotações anteriores para um item.
    Retorna links e fornecedores de itens já cotados/comprados com mesmo código ou descrição similar.
    """
    if not codigo_item and not descricao:
        return {"historico": [], "encontrado": False}
    
    # Buscar todas as OCs
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(5000)
    
    historico = []
    seen_combinations = set()  # Para evitar duplicatas
    
    # Status que indicam que o item já foi cotado
    status_cotados = ['cotado', 'comprado', 'em_separacao', 'em_transito', 'entregue']
    
    for po in pos:
        for item in po.get('items', []):
            item_status = item.get('status', '').lower()
            
            # Verificar se o item já foi cotado
            if item_status not in status_cotados:
                continue
            
            # Verificar se é o mesmo item (por código ou descrição similar)
            match_codigo = codigo_item and item.get('codigo_item', '').upper() == codigo_item.upper()
            match_descricao = descricao and descricao.upper() in (item.get('descricao', '') or '').upper()
            
            if not match_codigo and not match_descricao:
                continue
            
            # Extrair fontes de compra do item
            fontes_compra = item.get('fontes_compra', [])
            
            for fonte in fontes_compra:
                fornecedor = fonte.get('fornecedor', '').strip()
                link = fonte.get('link', '').strip()
                preco = fonte.get('preco_unitario', 0)
                frete = fonte.get('frete', 0)
                
                # Criar chave única para evitar duplicatas
                key = f"{fornecedor}_{link}_{preco}"
                if key in seen_combinations:
                    continue
                seen_combinations.add(key)
                
                if fornecedor or link:
                    historico.append({
                        "numero_oc": po.get('numero_oc', ''),
                        "codigo_item": item.get('codigo_item', ''),
                        "descricao": item.get('descricao', ''),
                        "status": item_status,
                        "fornecedor": fornecedor,
                        "link": link,
                        "preco_unitario": preco,
                        "frete": frete,
                        "data_compra": item.get('updated_at') or item.get('created_at', '')
                    })
    
    # Ordenar por data mais recente
    historico.sort(key=lambda x: x.get('data_compra', ''), reverse=True)
    
    # Limitar a 10 resultados mais recentes
    historico = historico[:10]
    
    return {
        "historico": historico,
        "encontrado": len(historico) > 0,
        "total": len(historico)
    }

@api_router.post("/purchase-orders/preview-pdf")
async def preview_pdf_purchase_order(file: UploadFile = File(...), current_user: dict = Depends(require_admin)):
    """Preview de PDF de Ordem de Compra - retorna itens sem criar OC (ADMIN ONLY)"""
    
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Apenas arquivos PDF são aceitos")
    
    # Ler conteúdo do PDF
    pdf_content = await file.read()
    
    # Extrair dados
    oc_data = extract_oc_from_pdf(pdf_content)
    
    if not oc_data["items"]:
        raise HTTPException(status_code=400, detail="Nenhum item encontrado no PDF. Verifique o formato do arquivo.")
    
    # Processar itens com dados de referência (preview apenas)
    preview_items = []
    items_without_ref = []
    
    import random
    
    for item in oc_data["items"]:
        # Buscar TODAS as ocorrências deste código (em todos os lotes)
        ref_items = await db.reference_items.find(
            {"codigo_item": item["codigo_item"]},
            {"_id": 0}
        ).to_list(100)
        
        preview_item = {
            "codigo_item": item["codigo_item"],
            "quantidade": item["quantidade"],
            "unidade": item.get("unidade", "UN"),
            "descricao": item.get("descricao", ""),
            "endereco_entrega": item.get("endereco_entrega", ""),
            "regiao": item.get("regiao", "")
        }
        
        # PRIORIDADE: Usar preço do PDF se disponível
        if item.get("preco_venda_pdf"):
            preview_item["preco_venda"] = item["preco_venda_pdf"]
        
        if ref_items:
            if len(ref_items) > 1:
                non_admin_items = [ri for ri in ref_items if ri['responsavel'] in ['Maria', 'Mylena', 'Fabio']]
                selected_ref = random.choice(non_admin_items) if non_admin_items else ref_items[0]
            else:
                selected_ref = ref_items[0]
            
            preview_item["responsavel"] = selected_ref['responsavel']
            preview_item["lote"] = selected_ref['lote']
            preview_item["marca_modelo"] = selected_ref.get('marca_modelo', '')
            # SEMPRE usar a descrição do Excel (mais completa)
            preview_item["descricao"] = selected_ref['descricao']
            
            # Só usa preço da planilha se não tiver preço do PDF
            if not preview_item.get("preco_venda") and selected_ref.get('preco_venda_unitario'):
                preview_item["preco_venda"] = selected_ref['preco_venda_unitario']
        else:
            items_without_ref.append(item["codigo_item"])
            preview_item["responsavel"] = "⚠️ NÃO ENCONTRADO"
            preview_item["lote"] = "⚠️ NÃO ENCONTRADO"
            preview_item["marca_modelo"] = ""
        
        preview_items.append(preview_item)
    
    return {
        "numero_oc": oc_data["numero_oc"],
        "endereco_entrega": oc_data.get("endereco_entrega", ""),
        "data_entrega": oc_data.get("data_entrega"),
        "cnpj_requisitante": oc_data.get("cnpj_requisitante", ""),
        "items": preview_items,
        "total_items": len(preview_items),
        "items_without_ref": items_without_ref
    }

@api_router.post("/purchase-orders/upload-pdf")
async def upload_pdf_purchase_order(file: UploadFile = File(...), current_user: dict = Depends(require_admin)):
    """Upload de PDF de Ordem de Compra (ADMIN ONLY)"""
    
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Apenas arquivos PDF são aceitos")
    
    # Ler conteúdo do PDF
    pdf_content = await file.read()
    
    # Extrair dados
    oc_data = extract_oc_from_pdf(pdf_content)
    
    if not oc_data["items"]:
        raise HTTPException(status_code=400, detail="Nenhum item encontrado no PDF. Verifique o formato do arquivo.")
    
    # Verificar se OC já existe
    existing_po = await db.purchase_orders.find_one({"numero_oc": oc_data["numero_oc"]}, {"_id": 0})
    if existing_po:
        raise HTTPException(status_code=409, detail=f"Ordem de Compra {oc_data['numero_oc']} já existe no sistema")
    
    # Processar itens com dados de referência
    processed_items = []
    items_without_ref = []
    
    import random
    
    # OTIMIZAÇÃO: Batch query - buscar todas as referências de uma vez
    all_codigo_items = [item["codigo_item"] for item in oc_data["items"]]
    all_ref_items = await db.reference_items.find(
        {"codigo_item": {"$in": all_codigo_items}},
        {"_id": 0}
    ).to_list(5000)
    
    # Criar lookup dictionary
    ref_lookup = {}
    for ref in all_ref_items:
        codigo = ref["codigo_item"]
        if codigo not in ref_lookup:
            ref_lookup[codigo] = []
        ref_lookup[codigo].append(ref)
    
    for item in oc_data["items"]:
        # Buscar do lookup
        ref_items = ref_lookup.get(item["codigo_item"], [])
        
        if ref_items:
            # Se existem múltiplas ocorrências (item em vários lotes)
            if len(ref_items) > 1:
                # Pegar apenas responsáveis não-admin
                non_admin_items = [ri for ri in ref_items if ri['responsavel'] in ['Maria', 'Mylena', 'Fabio']]
                
                if non_admin_items:
                    # Escolher aleatoriamente entre os não-admins
                    selected_ref = random.choice(non_admin_items)
                else:
                    # Se não houver não-admins, usar o primeiro disponível
                    selected_ref = ref_items[0]
            else:
                # Se existe apenas uma ocorrência, usar ela
                selected_ref = ref_items[0]
            
            # Preencher dados
            item["responsavel"] = selected_ref['responsavel']
            item["lote"] = selected_ref['lote']
            item["lot_number"] = selected_ref['lot_number']
            item["regiao"] = selected_ref['regiao']
            # SEMPRE usar a descrição do Excel (mais completa)
            item["descricao"] = selected_ref['descricao']
            if not item.get("marca_modelo"):
                item["marca_modelo"] = selected_ref.get('marca_modelo', '')
            
            # Preencher preço de venda
            if selected_ref.get('preco_venda_unitario'):
                item["preco_venda"] = selected_ref['preco_venda_unitario']
                # Calcular imposto (11%)
                item["imposto"] = round(item["preco_venda"] * item["quantidade"] * 0.11, 2)
        else:
            # Item não encontrado na referência
            items_without_ref.append(item["codigo_item"])
            item["responsavel"] = "⚠️ NÃO ENCONTRADO"
            item["lote"] = "⚠️ NÃO ENCONTRADO"
            item["lot_number"] = 0
            item["regiao"] = item.get("regiao", "")
            item["marca_modelo"] = ""
            if not item.get("descricao"):
                item["descricao"] = f"Item {item['codigo_item']}"
        
        # Garantir que todos os campos obrigatórios existem
        item.setdefault("status", ItemStatus.PENDENTE)
        item.setdefault("preco_compra", None)
        item.setdefault("preco_venda", None)
        item.setdefault("imposto", None)
        item.setdefault("custo_frete", None)
        item.setdefault("lucro_liquido", None)
        
        po_item = POItem(**item)
        processed_items.append(po_item)
    
    # Criar OC com data de entrega extraída do PDF
    po = PurchaseOrder(
        numero_oc=oc_data["numero_oc"],
        items=processed_items,
        created_by=current_user.get('sub')
    )
    
    doc = po.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    # Adicionar data de entrega e CNPJ do requisitante
    if oc_data.get("data_entrega"):
        doc['data_entrega'] = oc_data["data_entrega"]
    if oc_data.get("cnpj_requisitante"):
        doc['cnpj_requisitante'] = oc_data["cnpj_requisitante"]
    
    await db.purchase_orders.insert_one(doc)
    
    warning = ""
    if items_without_ref:
        warning = f" ATENÇÃO: {len(items_without_ref)} itens não encontrados no banco de referência: {', '.join(items_without_ref[:5])}"
    
    return {
        "success": True,
        "message": f"OC criada com sucesso a partir do PDF{warning}",
        "po_id": po.id,
        "numero_oc": po.numero_oc,
        "total_items": len(processed_items),
        "items_without_ref": items_without_ref
    }

@api_router.post("/purchase-orders/upload-multiple-pdfs")
async def upload_multiple_pdfs(files: List[UploadFile] = File(...), current_user: dict = Depends(require_admin)):
    """Upload de múltiplos PDFs de Ordens de Compra (ADMIN ONLY)"""
    import random
    
    results = {
        "success": [],
        "failed": [],
        "duplicates": [],
        "total_processed": 0,
        "total_items_created": 0
    }
    
    for file in files:
        try:
            if not file.filename.endswith('.pdf'):
                results["failed"].append({
                    "filename": file.filename,
                    "error": "Arquivo não é PDF"
                })
                continue
            
            # Ler conteúdo do PDF
            pdf_content = await file.read()
            
            # Extrair dados
            oc_data = extract_oc_from_pdf(pdf_content)
            
            if not oc_data.get("items"):
                results["failed"].append({
                    "filename": file.filename,
                    "error": "Nenhum item encontrado no PDF"
                })
                continue
            
            # Verificar se OC já existe
            existing_po = await db.purchase_orders.find_one({"numero_oc": oc_data["numero_oc"]}, {"_id": 0})
            if existing_po:
                results["duplicates"].append({
                    "filename": file.filename,
                    "numero_oc": oc_data["numero_oc"],
                    "existing_id": existing_po["id"]
                })
                continue
            
            # Processar itens com dados de referência
            processed_items = []
            
            # OTIMIZAÇÃO: Batch query - buscar todas as referências de uma vez
            all_codigo_items = [item["codigo_item"] for item in oc_data["items"]]
            all_ref_items = await db.reference_items.find(
                {"codigo_item": {"$in": all_codigo_items}},
                {"_id": 0}
            ).to_list(5000)
            
            # Criar lookup dictionary
            ref_lookup = {}
            for ref in all_ref_items:
                codigo = ref["codigo_item"]
                if codigo not in ref_lookup:
                    ref_lookup[codigo] = []
                ref_lookup[codigo].append(ref)
            
            for item in oc_data["items"]:
                # Buscar do lookup
                ref_items = ref_lookup.get(item["codigo_item"], [])
                
                responsavel = item.get("responsavel", "")
                lote = item.get("lote", "")
                lot_number = 0
                
                if ref_items:
                    ref_item = ref_items[0]
                    if len(ref_items) > 1:
                        ref_item = random.choice(ref_items)
                    responsavel = ref_item.get("quem_cotou", "")
                    lote = ref_item.get("lote", "")
                    try:
                        lot_number = int(''.join(filter(str.isdigit, lote))) if lote else 0
                    except:
                        lot_number = 0
                
                preco_venda = item.get("preco_venda_pdf") or item.get("preco_venda")
                
                processed_items.append(POItem(
                    codigo_item=item["codigo_item"],
                    quantidade=int(item.get("quantidade", 1)),
                    unidade=item.get("unidade", "UN"),
                    descricao=item.get("descricao", ""),
                    endereco_entrega=oc_data.get("endereco_entrega", ""),
                    responsavel=responsavel,
                    lote=lote,
                    lot_number=lot_number,
                    regiao=item.get("regiao", ""),
                    status="pendente",
                    preco_venda=preco_venda
                ))
            
            # Criar OC com data de entrega
            po = PurchaseOrder(
                numero_oc=oc_data["numero_oc"],
                cnpj_requisitante=oc_data.get("cnpj_requisitante", ""),
                items=processed_items,
                created_by=current_user.get('sub')
            )
            
            doc = po.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            
            # Adicionar data de entrega se extraída do PDF
            if oc_data.get("data_entrega"):
                doc['data_entrega'] = oc_data["data_entrega"]
            
            await db.purchase_orders.insert_one(doc)
            
            results["success"].append({
                "filename": file.filename,
                "numero_oc": po.numero_oc,
                "po_id": po.id,
                "total_items": len(processed_items)
            })
            results["total_items_created"] += len(processed_items)
            
        except Exception as e:
            results["failed"].append({
                "filename": file.filename,
                "error": str(e)
            })
    
    results["total_processed"] = len(files)
    
    return results

@api_router.get("/purchase-orders/check-duplicate/{numero_oc}")
async def check_duplicate_purchase_order(numero_oc: str, current_user: dict = Depends(require_admin)):
    """Verificar se OC já existe (ADMIN ONLY)"""
    existing_po = await db.purchase_orders.find_one({"numero_oc": numero_oc}, {"_id": 0})
    
    if existing_po:
        return {
            "exists": True,
            "message": f"Ordem de Compra {numero_oc} já existe no sistema",
            "existing_po": {
                "id": existing_po["id"],
                "numero_oc": existing_po["numero_oc"],
                "created_at": existing_po["created_at"],
                "total_items": len(existing_po["items"])
            }
        }
    else:
        return {
            "exists": False,
            "message": f"Ordem de Compra {numero_oc} não existe no sistema"
        }

@api_router.post("/purchase-orders", response_model=PurchaseOrder)
async def create_purchase_order(po_create: PurchaseOrderCreate, current_user: dict = Depends(require_admin)):
    """Criar nova Ordem de Compra (ADMIN ONLY)"""
    
    # Verificar se OC já existe
    existing_po = await db.purchase_orders.find_one({"numero_oc": po_create.numero_oc}, {"_id": 0})
    if existing_po:
        raise HTTPException(status_code=409, detail=f"Ordem de Compra {po_create.numero_oc} já existe no sistema")
    
    processed_items = []
    items_not_found = []
    
    # OTIMIZAÇÃO: Batch query - buscar todas as referências de uma vez
    all_codigo_items = [item.codigo_item for item in po_create.items]
    all_ref_items = await db.reference_items.find(
        {"codigo_item": {"$in": all_codigo_items}},
        {"_id": 0}
    ).to_list(5000)
    
    # Criar lookup dictionary
    ref_lookup = {}
    for ref in all_ref_items:
        codigo = ref["codigo_item"]
        if codigo not in ref_lookup:
            ref_lookup[codigo] = []
        ref_lookup[codigo].append(ref)
    
    import random
    
    for item in po_create.items:
        # Buscar do lookup
        ref_items = ref_lookup.get(item.codigo_item, [])
        
        if ref_items:
            # Se existem múltiplas ocorrências (item em vários lotes)
            if len(ref_items) > 1:
                # Pegar apenas responsáveis não-admin
                non_admin_items = [ri for ri in ref_items if ri['responsavel'] in ['Maria', 'Mylena', 'Fabio']]
                
                if non_admin_items:
                    # Escolher aleatoriamente entre os não-admins
                    selected_ref = random.choice(non_admin_items)
                else:
                    # Se não houver não-admins, usar o primeiro disponível
                    selected_ref = ref_items[0]
            else:
                # Se existe apenas uma ocorrência, usar ela
                selected_ref = ref_items[0]
            
            # Preencher com dados da referência selecionada
            item.responsavel = selected_ref['responsavel']
            item.lote = selected_ref['lote']
            item.lot_number = selected_ref['lot_number']
            item.regiao = selected_ref.get('regiao', item.regiao or '')
            # SEMPRE usar a descrição do Excel (mais completa)
            item.descricao = selected_ref['descricao']
            if not item.marca_modelo:
                item.marca_modelo = selected_ref.get('marca_modelo', '')
            
            # Preencher preço de venda automaticamente
            if not item.preco_venda and selected_ref.get('preco_venda_unitario'):
                item.preco_venda = selected_ref['preco_venda_unitario']
            
            # Calcular imposto automaticamente (11% do preço de venda total)
            if item.preco_venda and not item.imposto:
                item.imposto = round(item.preco_venda * item.quantidade * 0.11, 2)
        else:
            # Item não encontrado - marcar claramente
            items_not_found.append(item.codigo_item)
            if not item.responsavel:
                item.responsavel = "⚠️ NÃO ENCONTRADO"
            if not item.lote:
                item.lote = "⚠️ NÃO ENCONTRADO"
            if not item.regiao:
                item.regiao = item.regiao or "Verificar PDF"
            item.lot_number = 0
            if not item.descricao:
                item.descricao = f"Item {item.codigo_item} - VERIFICAR MANUALMENTE"
        
        processed_items.append(item)
    
    po = PurchaseOrder(
        numero_oc=po_create.numero_oc,
        data_entrega=po_create.data_entrega,
        endereco_entrega=po_create.endereco_entrega or "",
        items=processed_items,
        created_by=current_user.get('sub')
    )
    
    doc = po.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    # Garantir que endereco_entrega está no documento
    if po_create.endereco_entrega:
        doc['endereco_entrega'] = po_create.endereco_entrega
    
    await db.purchase_orders.insert_one(doc)
    
    return po

@api_router.get("/purchase-orders")
async def get_purchase_orders(
    current_user: dict = Depends(get_current_user),
    page: int = 1,
    limit: int = 5,
    search: str = None,
    responsavel: str = None
):
    """Listar Ordens de Compra com paginação server-side"""
    query = {}
    
    # Busca total primeiro (para paginação)
    total = await db.purchase_orders.count_documents(query)
    
    # Paginação
    skip = (page - 1) * limit
    
    # Se limit = 0, retornar todos (para compatibilidade)
    if limit == 0:
        pos = await db.purchase_orders.find(query, {"_id": 0}).to_list(1000)
    else:
        pos = await db.purchase_orders.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    
    for po in pos:
        if 'created_at' in po and po['created_at']:
            if isinstance(po['created_at'], str):
                try:
                    po['created_at'] = datetime.fromisoformat(po['created_at'])
                except:
                    pass
        
        # Adicionar índice original a cada item ANTES de filtrar
        for idx, item in enumerate(po.get('items', [])):
            item['_originalIndex'] = idx
        
        # Se não for admin, filtrar apenas itens do responsável (case-insensitive)
        if current_user['role'] != 'admin' and current_user.get('owner_name'):
            user_name = current_user['owner_name'].strip().upper()
            po['items'] = [item for item in po.get('items', []) if (item.get('responsavel') or '').strip().upper() == user_name]
    
    return {
        "data": pos,
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": (total + limit - 1) // limit if limit > 0 else 1
    }


@api_router.get("/purchase-orders/list/simple")
async def get_purchase_orders_simple(
    current_user: dict = Depends(get_current_user),
    search_oc: str = None,
    search_codigo: str = None,
    search_descricao: str = None,
    search_responsavel: str = None,
    date_from: str = None,
    date_to: str = None
):
    """Listar Ordens de Compra de forma simplificada (para carregamento rápido)
    Retorna apenas dados essenciais sem os itens completos
    Suporta filtros server-side para performance"""
    query = {}
    
    pos = await db.purchase_orders.find(query, {
        "_id": 0,
        "id": 1,
        "numero_oc": 1,
        "created_at": 1,
        "data_entrega": 1,  # Data de entrega
        "endereco_entrega": 1,  # Endereço de entrega
        "cnpj_requisitante": 1,
        "items": 1  # Necessário para contagem e filtro
    }).to_list(1000)
    
    result = []
    for po in pos:
        if 'created_at' in po and po.get('created_at'):
            if isinstance(po['created_at'], str):
                try:
                    po['created_at'] = datetime.fromisoformat(po['created_at'])
                except:
                    po['created_at'] = datetime.now(timezone.utc)
        else:
            po['created_at'] = datetime.now(timezone.utc)
        
        # Filtro por número da OC
        if search_oc:
            if search_oc.lower() not in po.get('numero_oc', '').lower():
                continue
        
        # Filtro por data
        if date_from or date_to:
            po_date = po.get('created_at')
            if isinstance(po_date, str):
                try:
                    po_date = datetime.fromisoformat(po_date)
                except:
                    po_date = datetime.now(timezone.utc)
            if not po_date:
                po_date = datetime.now(timezone.utc)
            
            if date_from:
                from_date = datetime.fromisoformat(date_from)
                if po_date.date() < from_date.date():
                    continue
            
            if date_to:
                to_date = datetime.fromisoformat(date_to)
                if po_date.date() > to_date.date():
                    continue
        
        items = po.get('items', [])
        
        # Se não for admin, filtrar apenas itens do responsável
        if current_user['role'] != 'admin' and current_user.get('owner_name'):
            user_name = current_user['owner_name'].strip().upper()
            items = [item for item in items if (item.get('responsavel') or '').strip().upper() == user_name]
        
        # Filtro por código do item
        if search_codigo:
            search_codigo_upper = search_codigo.upper()
            items = [item for item in items if search_codigo_upper in (item.get('codigo_item') or '').upper()]
            if not items:
                continue  # Pular OC se nenhum item corresponder
        
        # Filtro por descrição/nome do item
        if search_descricao:
            search_descricao_upper = search_descricao.upper()
            items = [item for item in items if search_descricao_upper in (item.get('descricao') or '').upper()]
            if not items:
                continue  # Pular OC se nenhum item corresponder
        
        # Filtro por responsável
        if search_responsavel:
            if search_responsavel == 'nao_atribuido':
                items = [item for item in items if 
                    not item.get('responsavel') or 
                    item.get('responsavel', '').strip() == '' or
                    'NÃO ENCONTRADO' in item.get('responsavel', '') or
                    'Não atribuído' in item.get('responsavel', '')
                ]
            else:
                items = [item for item in items if item.get('responsavel') == search_responsavel]
            
            if not items:
                continue  # Pular OC se nenhum item corresponder
        
        # Calcular resumo dos itens
        total_items = len(items)
        if total_items == 0:
            continue  # Pular OCs sem itens para o usuário
            
        valor_total = sum(
            (item.get('preco_venda') or 0) * (item.get('quantidade') or 1) 
            for item in items
        )
        
        result.append({
            "id": po['id'],
            "numero_oc": po['numero_oc'],
            "created_at": po['created_at'],
            "data_entrega": po.get('data_entrega'),  # Data de entrega extraída do PDF
            "endereco_entrega": po.get('endereco_entrega'),  # Endereço de entrega extraído do PDF
            "cnpj_requisitante": po.get('cnpj_requisitante', ''),
            "total_items": total_items,
            "valor_total": valor_total,
            # Incluir itens simplificados para filtro local (com quantidade, status e marca_modelo para resumo)
            "items": [{"codigo_item": i.get('codigo_item'), "descricao": i.get('descricao'), "responsavel": i.get('responsavel'), "quantidade": i.get('quantidade', 1), "status": i.get('status'), "marca_modelo": i.get('marca_modelo', '')} for i in items],
            # Resumo por status
            "status_count": {
                "pendente": sum(1 for i in items if i.get('status') == 'pendente'),
                "cotado": sum(1 for i in items if i.get('status') == 'cotado'),
                "comprado": sum(1 for i in items if i.get('status') == 'comprado'),
                "em_separacao": sum(1 for i in items if i.get('status') == 'em_separacao'),
                "em_transito": sum(1 for i in items if i.get('status') == 'em_transito'),
                "entregue": sum(1 for i in items if i.get('status') == 'entregue'),
            }
        })
    
    return result

@api_router.get("/purchase-orders/{po_id}")
async def get_purchase_order(po_id: str, current_user: dict = Depends(get_current_user)):
    """Obter detalhes de uma OC"""
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    if isinstance(po['created_at'], str):
        po['created_at'] = datetime.fromisoformat(po['created_at'])
    
    # Adicionar índice original a cada item ANTES de filtrar
    for idx, item in enumerate(po['items']):
        item['_originalIndex'] = idx
    
    # Se não for admin, filtrar apenas itens do responsável (case-insensitive)
    if current_user['role'] != 'admin' and current_user.get('owner_name'):
        user_name = current_user['owner_name'].strip().upper()
        po['items'] = [item for item in po['items'] if (item.get('responsavel') or '').strip().upper() == user_name]
    
    return po

@api_router.delete("/purchase-orders/{po_id}")
async def delete_purchase_order(po_id: str, current_user: dict = Depends(require_admin)):
    """Deletar uma OC (ADMIN ONLY)"""
    result = await db.purchase_orders.delete_one({"id": po_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    return {"message": "Ordem de Compra deletada com sucesso"}

@api_router.put("/purchase-orders/{po_id}")
async def update_purchase_order(po_id: str, po_update: PurchaseOrderCreate, current_user: dict = Depends(require_admin)):
    """Atualizar uma OC (ADMIN ONLY)"""
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    # OTIMIZAÇÃO: Batch query - buscar todas as referências de uma vez
    all_codigo_items = [item.codigo_item for item in po_update.items]
    all_ref_items = await db.reference_items.find(
        {"codigo_item": {"$in": all_codigo_items}},
        {"_id": 0}
    ).to_list(5000)
    
    # Criar lookup dictionary
    ref_lookup = {ref["codigo_item"]: ref for ref in all_ref_items}
    
    # Processar itens
    processed_items = []
    for item in po_update.items:
        ref_item = ref_lookup.get(item.codigo_item)
        
        if ref_item:
            item.responsavel = ref_item['responsavel']
            item.lote = ref_item['lote']
            item.lot_number = ref_item['lot_number']
            item.regiao = ref_item['regiao']
            # SEMPRE usar a descrição do Excel (mais completa)
            item.descricao = ref_item['descricao']
            if not item.marca_modelo:
                item.marca_modelo = ref_item['marca_modelo']
        
        processed_items.append(item)
    
    # Atualizar
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {
            "numero_oc": po_update.numero_oc,
            "data_entrega": po_update.data_entrega,
            "items": [item.model_dump() for item in processed_items]
        }}
    )
    
    return {"message": "Ordem de Compra atualizada com sucesso"}


@api_router.patch("/purchase-orders/{po_id}/data-entrega")
async def update_data_entrega(
    po_id: str,
    data_entrega: str = Body(..., embed=True),
    current_user: dict = Depends(require_admin)
):
    """Atualizar data de entrega de uma OC (ADMIN ONLY)"""
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    # Validar formato da data (YYYY-MM-DD ou DD/MM/YYYY)
    try:
        if '/' in data_entrega:
            # Converter de DD/MM/YYYY para YYYY-MM-DD
            dia, mes, ano = data_entrega.split('/')
            data_entrega = f"{ano}-{mes}-{dia}"
        # Validar que é uma data válida
        datetime.strptime(data_entrega, '%Y-%m-%d')
    except ValueError:
        raise HTTPException(status_code=400, detail="Data inválida. Use o formato DD/MM/AAAA ou AAAA-MM-DD")
    
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"data_entrega": data_entrega}}
    )
    
    return {"message": "Data de entrega atualizada com sucesso", "data_entrega": data_entrega}


@api_router.patch("/purchase-orders/{po_id}/items/{codigo_item}")
async def update_item_status(po_id: str, codigo_item: str, update: ItemStatusUpdate, current_user: dict = Depends(get_current_user)):
    """Atualizar status de um item"""
    logger.info(f"update_item_status chamado: po_id={po_id}, codigo_item={codigo_item}, user={current_user.get('sub')}")
    
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    item_updated = False
    user_role = current_user.get('role', '')
    user_email = current_user.get('sub', '')
    
    for item in po['items']:
        if item['codigo_item'] == codigo_item:
            # PERMISSÃO SIMPLIFICADA: Qualquer usuário autenticado pode editar
            status_anterior = item.get('status', 'pendente')
            logger.info(f"Usuário {user_email} (role={user_role}) editando item {codigo_item}, status_anterior={status_anterior}, novo_status={update.status}")
            
            # VERIFICAR SE PRECISA REVERTER USO DE ESTOQUE
            status_antes_compra = ['pendente', 'cotado']
            status_apos_compra = ['comprado', 'em_separacao', 'em_transito', 'entregue']
            
            if update.status in status_antes_compra and status_anterior in status_apos_compra:
                if item.get('estoque_origem') or item.get('atendido_por_estoque'):
                    logger.info(f"Item voltando para {update.status}, revertendo uso de estoque...")
                    resultado_reversao = await reverter_uso_estoque(item, po_id, po.get('numero_oc', ''))
                    logger.info(f"Resultado da reversão de estoque: {resultado_reversao}")
            
            item['status'] = update.status
            atualizar_data_compra(item, update.status)  # Atualiza data de compra automaticamente
            
            # Atualizar fontes de compra (todos podem editar)
            if update.fontes_compra is not None:
                item['fontes_compra'] = [fc.model_dump() for fc in update.fontes_compra]
                
                # Calcular totais das fontes de compra
                total_custo = 0
                total_frete = 0
                total_qtd = 0
                for fc in update.fontes_compra:
                    total_custo += fc.quantidade * fc.preco_unitario
                    total_frete += fc.frete
                    total_qtd += fc.quantidade
                
                # Atualizar campos consolidados baseados nas fontes
                if total_qtd > 0:
                    item['preco_compra'] = round(total_custo / total_qtd, 2)  # Preço médio unitário
                item['frete_compra'] = total_frete
            
            # Campos simples - todos podem atualizar
            if update.link_compra is not None:
                item['link_compra'] = update.link_compra
            
            if update.preco_compra is not None and not update.fontes_compra:
                item['preco_compra'] = update.preco_compra
            
            if update.frete_compra is not None and not update.fontes_compra:
                item['frete_compra'] = update.frete_compra
            
            # Preço de venda pode ser editado por qualquer usuário
            if update.preco_venda is not None:
                item['preco_venda'] = update.preco_venda
            
            # Imposto e frete de envio são restritos a admins
            if current_user['role'] == 'admin':
                if update.imposto is not None:
                    item['imposto'] = update.imposto
                if update.frete_envio is not None:
                    item['frete_envio'] = update.frete_envio
            
            # Calcular lucro líquido usando a função centralizada
            calcular_lucro_item(item)
            
            # Atualizar datas
            now = datetime.now(timezone.utc).isoformat()
            if update.status == ItemStatus.COTADO:
                item['data_cotacao'] = now
            elif update.status == ItemStatus.COMPRADO:
                item['data_compra'] = now
            elif update.status == ItemStatus.EM_TRANSITO:
                item['data_envio'] = now
            elif update.status == ItemStatus.ENTREGUE:
                item['data_entrega'] = now
            
            # Atualizar código de rastreio se fornecido
            if update.codigo_rastreio is not None:
                item['codigo_rastreio'] = update.codigo_rastreio.strip().upper() if update.codigo_rastreio else None
            
            item_updated = True
            break
    
    if not item_updated:
        raise HTTPException(status_code=404, detail="Item não encontrado")
    
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": po['items']}}
    )
    
    return {"message": "Item atualizado com sucesso"}

@api_router.patch("/purchase-orders/{po_id}/items/by-index/{item_index}")
async def update_item_by_index_status(
    po_id: str, 
    item_index: int, 
    update: ItemStatusUpdate, 
    current_user: dict = Depends(get_current_user)
):
    """Atualizar item por índice - QUALQUER USUÁRIO PODE EDITAR"""
    logger.info(f"update_item_by_index_status: po_id={po_id}, item_index={item_index}, user={current_user.get('sub')}")
    
    # Buscar OC diretamente sem filtro
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    if item_index < 0 or item_index >= len(po['items']):
        raise HTTPException(status_code=404, detail="Índice de item inválido")
    
    item = po['items'][item_index]
    status_anterior = item.get('status', 'pendente')
    logger.info(f"Editando item {item.get('codigo_item')} - user: {current_user.get('sub')}, status_anterior={status_anterior}, novo_status={update.status}")
    
    # VERIFICAR SE PRECISA REVERTER USO DE ESTOQUE
    # Se o item está sendo movido de volta para pendente/cotado E tinha usado estoque
    status_antes_compra = ['pendente', 'cotado']
    status_apos_compra = ['comprado', 'em_separacao', 'em_transito', 'entregue']
    
    if update.status in status_antes_compra and status_anterior in status_apos_compra:
        # Verificar se o item usou estoque
        if item.get('estoque_origem') or item.get('atendido_por_estoque'):
            logger.info(f"Item voltando para {update.status}, revertendo uso de estoque...")
            resultado_reversao = await reverter_uso_estoque(item, po_id, po.get('numero_oc', ''))
            logger.info(f"Resultado da reversão de estoque: {resultado_reversao}")
    
    # APLICAR TODAS AS ALTERAÇÕES - SEM RESTRIÇÕES
    item['status'] = update.status
    atualizar_data_compra(item, update.status)  # Atualiza data de compra automaticamente
    
    # Fontes de compra
    if update.fontes_compra is not None:
        item['fontes_compra'] = [fc.model_dump() for fc in update.fontes_compra]
        total_custo = sum(fc.quantidade * fc.preco_unitario for fc in update.fontes_compra)
        total_frete = sum(fc.frete for fc in update.fontes_compra)
        total_qtd = sum(fc.quantidade for fc in update.fontes_compra)
        if total_qtd > 0:
            item['preco_compra'] = round(total_custo / total_qtd, 2)
        item['frete_compra'] = total_frete
    
    # Preço de venda - TODOS podem editar
    if update.preco_venda is not None:
        item['preco_venda'] = update.preco_venda
    
    # Imposto - TODOS podem editar
    if update.imposto is not None:
        item['imposto'] = update.imposto
    
    # Frete envio - TODOS podem editar
    if update.frete_envio is not None:
        item['frete_envio'] = update.frete_envio
    
    # Link de compra
    if update.link_compra is not None:
        item['link_compra'] = update.link_compra
    
    # Preço de compra manual
    if update.preco_compra is not None and not update.fontes_compra:
        item['preco_compra'] = update.preco_compra
    
    # Frete de compra manual
    if update.frete_compra is not None and not update.fontes_compra:
        item['frete_compra'] = update.frete_compra
    
    # Código de rastreio
    if update.codigo_rastreio is not None:
        item['codigo_rastreio'] = update.codigo_rastreio.strip().upper() if update.codigo_rastreio else None
    
    # No carrinho - checkbox
    if update.no_carrinho is not None:
        item['no_carrinho'] = update.no_carrinho
    
    # Observação
    if update.observacao is not None:
        item['observacao'] = update.observacao
    
    # Calcular lucro usando a função centralizada
    calcular_lucro_item(item)
    
    # Atualizar datas
    now = datetime.now(timezone.utc).isoformat()
    if update.status == ItemStatus.COTADO:
        item['data_cotacao'] = now
    elif update.status == ItemStatus.COMPRADO:
        item['data_compra'] = now
    elif update.status == ItemStatus.EM_TRANSITO:
        item['data_envio'] = now
    elif update.status == ItemStatus.ENTREGUE:
        item['data_entrega'] = now
    
    # SALVAR
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": po['items']}}
    )
    
    return {"message": "Item atualizado com sucesso"}


# ============== ENDPOINT PARA COMPRA PARCIAL ==============

class CompraParcialRequest(BaseModel):
    """Request para compra parcial de um item"""
    quantidade_comprar: int  # Quantidade a comprar (vai para 'comprado')
    fontes_compra: Optional[List[FonteCompra]] = None  # Fontes de compra (preço, fornecedor, etc)
    preco_unitario: Optional[float] = None
    frete: Optional[float] = None
    fornecedor: Optional[str] = None
    link: Optional[str] = None

@api_router.post("/purchase-orders/{po_id}/items/by-index/{item_index}/compra-parcial")
async def comprar_parcialmente(
    po_id: str,
    item_index: int,
    request: CompraParcialRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Comprar parcialmente um item cotado.
    
    Divide o item original em dois:
    1. A quantidade comprada vai para status 'comprado'
    2. A quantidade restante permanece em status 'cotado'
    
    Exemplo:
    - Item original: 4 unidades em 'cotado'
    - Comprar 2 unidades
    - Resultado: 2 itens no banco
      - Item 1: 2 unidades em 'comprado' (com dados da compra)
      - Item 2: 2 unidades em 'cotado' (mantém dados da cotação)
    """
    logger.info(f"compra_parcial: po_id={po_id}, item_index={item_index}, qtd_comprar={request.quantidade_comprar}")
    
    # Buscar OC
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    if item_index < 0 or item_index >= len(po['items']):
        raise HTTPException(status_code=404, detail="Índice de item inválido")
    
    item_original = po['items'][item_index]
    quantidade_total = item_original.get('quantidade', 0)
    
    # Validações
    if item_original.get('status') != 'cotado':
        raise HTTPException(status_code=400, detail="Apenas itens com status 'cotado' podem ser comprados parcialmente")
    
    if request.quantidade_comprar <= 0:
        raise HTTPException(status_code=400, detail="Quantidade a comprar deve ser maior que zero")
    
    if request.quantidade_comprar >= quantidade_total:
        raise HTTPException(status_code=400, detail=f"Para comprar todas as {quantidade_total} unidades, use a função de compra normal. Este endpoint é para compra parcial.")
    
    quantidade_restante = quantidade_total - request.quantidade_comprar
    now = datetime.now(timezone.utc).isoformat()
    
    # CRIAR ITEM COMPRADO (cópia do original com quantidade comprada)
    item_comprado = item_original.copy()
    item_comprado['quantidade'] = request.quantidade_comprar
    item_comprado['status'] = 'comprado'
    item_comprado['data_compra'] = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    item_comprado['compra_parcial'] = True  # Flag para indicar que veio de compra parcial
    item_comprado['compra_parcial_de'] = quantidade_total  # Quantidade original
    
    # Adicionar fonte de compra
    if request.fontes_compra:
        item_comprado['fontes_compra'] = [fc.model_dump() for fc in request.fontes_compra]
    elif request.preco_unitario is not None:
        item_comprado['fontes_compra'] = [{
            'quantidade': request.quantidade_comprar,
            'preco_unitario': request.preco_unitario,
            'frete': request.frete or 0,
            'fornecedor': request.fornecedor or '',
            'link': request.link or ''
        }]
    
    # Calcular preço de compra médio
    if item_comprado.get('fontes_compra'):
        total_custo = sum(fc.get('quantidade', 0) * fc.get('preco_unitario', 0) for fc in item_comprado['fontes_compra'])
        total_qtd = sum(fc.get('quantidade', 0) for fc in item_comprado['fontes_compra'])
        if total_qtd > 0:
            item_comprado['preco_compra'] = round(total_custo / total_qtd, 2)
        item_comprado['frete_compra'] = sum(fc.get('frete', 0) for fc in item_comprado['fontes_compra'])
    
    # Calcular lucro
    calcular_lucro_item(item_comprado)
    
    # ATUALIZAR ITEM ORIGINAL (mantém em cotado com quantidade restante)
    item_original['quantidade'] = quantidade_restante
    item_original['quantidade_original_antes_compra_parcial'] = quantidade_total
    
    # Inserir item comprado logo após o original
    po['items'].insert(item_index + 1, item_comprado)
    
    # Salvar
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": po['items']}}
    )
    
    logger.info(f"Compra parcial realizada: {request.quantidade_comprar} de {quantidade_total} unidades do item {item_original.get('codigo_item')}")
    
    return {
        "success": True,
        "message": f"Compra parcial realizada com sucesso!",
        "detalhes": {
            "codigo_item": item_original.get('codigo_item'),
            "quantidade_original": quantidade_total,
            "quantidade_comprada": request.quantidade_comprar,
            "quantidade_restante_cotado": quantidade_restante
        }
    }


class EnvioParcialRequest(BaseModel):
    """Request para envio parcial de um item"""
    quantidade_enviar: int  # Quantidade a enviar (vai para 'em_transito')
    codigo_rastreio: Optional[str] = None
    frete_envio: Optional[float] = None


@api_router.post("/purchase-orders/{po_id}/items/by-index/{item_index}/envio-parcial")
async def enviar_parcialmente(
    po_id: str,
    item_index: int,
    request: EnvioParcialRequest,
    current_user: dict = Depends(require_admin)
):
    """
    Enviar parcialmente um item em separação.
    
    Divide o item original em dois:
    1. A quantidade enviada vai para status 'em_transito'
    2. A quantidade restante permanece em status 'em_separacao'
    
    Exemplo:
    - Item original: 10 unidades em 'em_separacao'
    - Enviar 4 unidades
    - Resultado: 2 itens no banco
      - Item 1: 6 unidades em 'em_separacao' (quantidade restante)
      - Item 2: 4 unidades em 'em_transito' (com rastreio e frete)
    """
    logger.info(f"envio_parcial: po_id={po_id}, item_index={item_index}, qtd_enviar={request.quantidade_enviar}")
    
    # Buscar OC
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    if item_index < 0 or item_index >= len(po['items']):
        raise HTTPException(status_code=404, detail="Índice de item inválido")
    
    item_original = po['items'][item_index]
    quantidade_total = item_original.get('quantidade', 0)
    
    # Validações
    if item_original.get('status') != 'em_separacao':
        raise HTTPException(status_code=400, detail="Apenas itens com status 'em_separacao' podem ser enviados parcialmente")
    
    if request.quantidade_enviar <= 0:
        raise HTTPException(status_code=400, detail="Quantidade a enviar deve ser maior que zero")
    
    if request.quantidade_enviar >= quantidade_total:
        raise HTTPException(status_code=400, detail=f"Para enviar todas as {quantidade_total} unidades, use a função de envio normal. Este endpoint é para envio parcial.")
    
    quantidade_restante = quantidade_total - request.quantidade_enviar
    
    # CRIAR ITEM ENVIADO (cópia do original com quantidade enviada)
    item_enviado = item_original.copy()
    item_enviado['quantidade'] = request.quantidade_enviar
    item_enviado['status'] = 'em_transito'
    item_enviado['envio_parcial'] = True  # Flag para indicar que veio de envio parcial
    item_enviado['envio_parcial_de'] = quantidade_total  # Quantidade original
    
    # Adicionar código de rastreio se informado
    if request.codigo_rastreio:
        item_enviado['codigo_rastreio'] = request.codigo_rastreio.strip().upper()
    
    # Adicionar frete se informado
    if request.frete_envio is not None and request.frete_envio > 0:
        item_enviado['frete_envio'] = request.frete_envio
    
    # ATUALIZAR ITEM ORIGINAL (mantém em separação com quantidade restante)
    item_original['quantidade'] = quantidade_restante
    item_original['quantidade_original_antes_envio_parcial'] = quantidade_total
    
    # Inserir item enviado logo após o original
    po['items'].insert(item_index + 1, item_enviado)
    
    # Salvar
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": po['items']}}
    )
    
    logger.info(f"Envio parcial realizado: {request.quantidade_enviar} de {quantidade_total} unidades do item {item_original.get('codigo_item')}")
    
    return {
        "success": True,
        "message": f"Envio parcial realizado com sucesso!",
        "detalhes": {
            "codigo_item": item_original.get('codigo_item'),
            "quantidade_original": quantidade_total,
            "quantidade_enviada": request.quantidade_enviar,
            "quantidade_restante_separacao": quantidade_restante,
            "codigo_rastreio": request.codigo_rastreio
        }
    }


# Endpoint para mover múltiplos itens do carrinho para Comprado
@api_router.post("/purchase-orders/mover-carrinho-para-comprado")
async def mover_carrinho_para_comprado(
    items: List[dict],
    current_user: dict = Depends(get_current_user)
):
    """
    Move múltiplos itens marcados como 'no_carrinho' para status 'comprado'
    Recebe lista de: [{"po_id": "xxx", "item_index": 0}, ...]
    """
    logger.info(f"mover_carrinho_para_comprado: {len(items)} itens, user={current_user.get('sub')}")
    
    updated_count = 0
    errors = []
    now = datetime.now(timezone.utc).isoformat()
    
    # Agrupar por po_id para otimizar updates
    items_by_po = {}
    for item_info in items:
        po_id = item_info.get('po_id')
        item_index = item_info.get('item_index')
        if po_id and item_index is not None:
            if po_id not in items_by_po:
                items_by_po[po_id] = []
            items_by_po[po_id].append(item_index)
    
    for po_id, indices in items_by_po.items():
        try:
            po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
            if not po:
                errors.append(f"OC {po_id} não encontrada")
                continue
            
            for item_index in indices:
                if 0 <= item_index < len(po['items']):
                    item = po['items'][item_index]
                    item['status'] = 'comprado'
                    item['no_carrinho'] = False
                    atualizar_data_compra(item, 'comprado')  # Usa a função padronizada
                    updated_count += 1
                else:
                    errors.append(f"Índice {item_index} inválido na OC {po_id}")
            
            await db.purchase_orders.update_one(
                {"id": po_id},
                {"$set": {"items": po['items']}}
            )
        except Exception as e:
            errors.append(f"Erro ao atualizar OC {po_id}: {str(e)}")
    
    return {
        "message": f"{updated_count} itens movidos para Comprado",
        "updated_count": updated_count,
        "errors": errors if errors else None
    }

@api_router.patch("/purchase-orders/{po_id}/items/{codigo_item}/full")
async def update_item_full(
    po_id: str, 
    codigo_item: str, 
    update: ItemFullUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Atualização completa do item - apenas admin"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Apenas administradores podem fazer edição completa")
    
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    
    item_updated = False
    for item in po['items']:
        if item['codigo_item'] == codigo_item:
            # Atualizar campos se fornecidos
            if update.descricao is not None:
                item['descricao'] = update.descricao
            if update.quantidade is not None:
                item['quantidade'] = update.quantidade
            if update.unidade is not None:
                item['unidade'] = update.unidade
            if update.responsavel is not None:
                item['responsavel'] = update.responsavel
            if update.lote is not None:
                item['lote'] = update.lote
            if update.marca_modelo is not None:
                item['marca_modelo'] = update.marca_modelo
            
            # Recalcular imposto e lucro usando a função centralizada
            calcular_lucro_item(item)
            
            item_updated = True
            break
    
    if not item_updated:
        raise HTTPException(status_code=404, detail="Item não encontrado")
    
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": po['items']}}
    )
    
    return {"message": "Item atualizado com sucesso"}

@api_router.patch("/purchase-orders/{po_id}/items/by-index/{item_index}/full")
async def update_item_by_index(
    po_id: str, 
    item_index: int, 
    update: ItemFullUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Atualização completa do item por índice - apenas admin"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Apenas administradores podem fazer edição completa")
    
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    
    if item_index < 0 or item_index >= len(po['items']):
        raise HTTPException(status_code=404, detail="Índice de item inválido")
    
    item = po['items'][item_index]
    status_anterior = item.get('status', 'pendente')
    
    # Atualizar campos se fornecidos
    if update.descricao is not None:
        item['descricao'] = update.descricao
    if update.quantidade is not None:
        item['quantidade'] = update.quantidade
    if update.unidade is not None:
        item['unidade'] = update.unidade
    if update.responsavel is not None:
        item['responsavel'] = update.responsavel
    if update.lote is not None:
        item['lote'] = update.lote
    if update.marca_modelo is not None:
        item['marca_modelo'] = update.marca_modelo
    if update.status is not None:
        # VERIFICAR SE PRECISA REVERTER USO DE ESTOQUE
        status_antes_compra = ['pendente', 'cotado']
        status_apos_compra = ['comprado', 'em_separacao', 'em_transito', 'entregue']
        
        if update.status in status_antes_compra and status_anterior in status_apos_compra:
            if item.get('estoque_origem') or item.get('atendido_por_estoque'):
                logger.info(f"Item voltando para {update.status} via full update, revertendo uso de estoque...")
                resultado_reversao = await reverter_uso_estoque(item, po_id, po.get('numero_oc', ''))
                logger.info(f"Resultado da reversão de estoque: {resultado_reversao}")
        
        item['status'] = update.status
        atualizar_data_compra(item, update.status)  # Atualiza data de compra automaticamente
    if update.preco_venda is not None:
        item['preco_venda'] = update.preco_venda
    if update.quantidade_comprada is not None:
        item['quantidade_comprada'] = update.quantidade_comprada
    
    # Recalcular imposto e lucro usando a função centralizada
    calcular_lucro_item(item)
    
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": po['items']}}
    )
    
    return {"message": "Item atualizado com sucesso"}

@api_router.get("/dashboard", response_model=DashboardStats)
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    """Estatísticas do dashboard"""
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(1000)
    
    all_items = []
    ocs_with_user_items = 0
    
    for po in pos:
        # Filtrar itens baseado no role (case-insensitive)
        if current_user['role'] != 'admin' and current_user.get('owner_name'):
            user_name = current_user['owner_name'].strip().upper()
            filtered_items = [item for item in po['items'] if (item.get('responsavel') or '').strip().upper() == user_name]
            if filtered_items:  # Só contar OCs que têm itens do usuário
                ocs_with_user_items += 1
            all_items.extend(filtered_items)
        else:
            all_items.extend(po['items'])
    
    # Para admins, total_ocs é o total de OCs. Para usuários, é apenas OCs com seus itens
    total_ocs = len(pos) if current_user['role'] == 'admin' else ocs_with_user_items
    
    total_items = len(all_items)
    items_pendentes = sum(1 for item in all_items if item['status'] == ItemStatus.PENDENTE)
    items_cotados = sum(1 for item in all_items if item['status'] == ItemStatus.COTADO)
    items_comprados = sum(1 for item in all_items if item['status'] == ItemStatus.COMPRADO)
    items_em_separacao = sum(1 for item in all_items if item['status'] == ItemStatus.EM_SEPARACAO)
    items_pronto_envio = sum(1 for item in all_items if item['status'] == ItemStatus.PRONTO_ENVIO)
    items_em_transito = sum(1 for item in all_items if item['status'] == ItemStatus.EM_TRANSITO)
    items_entregues = sum(1 for item in all_items if item['status'] == ItemStatus.ENTREGUE)
    
    # Para usuários não-admin, mostrar apenas seus próprios itens no breakdown por responsável
    items_por_responsavel = {}
    if current_user['role'] == 'admin':
        # Admin vê todos os responsáveis com breakdown por status
        for owner in ['Maria', 'Mateus', 'João', 'Mylena', 'Fabio']:
            owner_items = [item for item in all_items if (item.get('responsavel') or '').strip().upper() == owner.upper()]
            items_por_responsavel[owner] = {
                "total": len(owner_items),
                "pendente": sum(1 for item in owner_items if item['status'] == ItemStatus.PENDENTE),
                "cotado": sum(1 for item in owner_items if item['status'] == ItemStatus.COTADO),
                "comprado": sum(1 for item in owner_items if item['status'] == ItemStatus.COMPRADO),
                "em_separacao": sum(1 for item in owner_items if item['status'] == ItemStatus.EM_SEPARACAO),
                "pronto_envio": sum(1 for item in owner_items if item['status'] == ItemStatus.PRONTO_ENVIO),
                "em_transito": sum(1 for item in owner_items if item['status'] == ItemStatus.EM_TRANSITO),
                "entregue": sum(1 for item in owner_items if item['status'] == ItemStatus.ENTREGUE)
            }
    else:
        # Usuário não-admin vê apenas seus próprios itens com breakdown
        owner_name = current_user.get('owner_name')
        if owner_name:
            user_name = owner_name.strip().upper()
            owner_items = [item for item in all_items if (item.get('responsavel') or '').strip().upper() == user_name]
            items_por_responsavel[owner_name] = {
                "total": len(owner_items),
                "pendente": sum(1 for item in owner_items if item['status'] == ItemStatus.PENDENTE),
                "cotado": sum(1 for item in owner_items if item['status'] == ItemStatus.COTADO),
                "comprado": sum(1 for item in owner_items if item['status'] == ItemStatus.COMPRADO),
                "em_separacao": sum(1 for item in owner_items if item['status'] == ItemStatus.EM_SEPARACAO),
                "pronto_envio": sum(1 for item in owner_items if item['status'] == ItemStatus.PRONTO_ENVIO),
                "em_transito": sum(1 for item in owner_items if item['status'] == ItemStatus.EM_TRANSITO),
                "entregue": sum(1 for item in owner_items if item['status'] == ItemStatus.ENTREGUE)
            }
    
    return DashboardStats(
        total_ocs=total_ocs,
        total_items=total_items,
        items_pendentes=items_pendentes,
        items_cotados=items_cotados,
        items_comprados=items_comprados,
        items_em_separacao=items_em_separacao,
        items_pronto_envio=items_pronto_envio,
        items_em_transito=items_em_transito,
        items_entregues=items_entregues,
        items_por_responsavel=items_por_responsavel
    )

@api_router.get("/admin/summary", response_model=List[AdminSummary])
async def get_admin_summary(current_user: dict = Depends(require_admin)):
    """Resumo financeiro para admins (ADMIN ONLY)"""
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(1000)
    
    summaries = []
    for po in pos:
        for item in po['items']:
            if item.get('preco_compra') or item.get('preco_venda'):
                summary = AdminSummary(
                    numero_oc=po['numero_oc'],
                    codigo_item=item['codigo_item'],
                    nome_item=item['descricao'][:50] + "..." if len(item['descricao']) > 50 else item['descricao'],
                    quem_cotou=item.get('responsavel', 'N/A'),
                    preco_compra=item.get('preco_compra'),
                    preco_venda=item.get('preco_venda'),
                    imposto=item.get('imposto'),
                    frete_compra=item.get('frete_compra'),
                    frete_envio=item.get('frete_envio'),
                    lucro_liquido=item.get('lucro_liquido'),
                    status=item['status']
                )
                summaries.append(summary)
    
    return summaries

@api_router.get("/items/duplicates")
async def get_duplicate_items(current_user: dict = Depends(get_current_user)):
    """Encontrar itens duplicados por código"""
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(1000)
    
    codigo_count = {}
    for po in pos:
        items_to_check = po['items']
        
        # Filtrar por responsável se não for admin (case-insensitive)
        if current_user['role'] != 'admin' and current_user.get('owner_name'):
            user_name = current_user['owner_name'].strip().upper()
            items_to_check = [item for item in po['items'] if (item.get('responsavel') or '').strip().upper() == user_name]
        
        for item in items_to_check:
            codigo = item['codigo_item']
            if codigo not in codigo_count:
                codigo_count[codigo] = []
            codigo_count[codigo].append({
                'numero_oc': po['numero_oc'],
                'quantidade': item['quantidade'],
                'status': item['status']
            })
    
    duplicates = {codigo: occurrences for codigo, occurrences in codigo_count.items() if len(occurrences) > 1}
    
    return {
        "total_duplicados": len(duplicates),
        "duplicados": duplicates
    }

@api_router.post("/purchase-orders/fix-responsaveis")
async def fix_responsaveis(current_user: dict = Depends(require_admin)):
    """Corrigir responsáveis faltantes nas OCs existentes (ADMIN ONLY)"""
    import random
    
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(1000)
    total_fixed = 0
    
    # OTIMIZAÇÃO: Coletar todos os codigos primeiro
    all_codigos = set()
    for po in pos:
        for item in po['items']:
            if not item.get('responsavel'):
                all_codigos.add(item["codigo_item"])
    
    # Buscar todas as referências de uma vez
    all_ref_items = await db.reference_items.find(
        {"codigo_item": {"$in": list(all_codigos)}},
        {"_id": 0}
    ).to_list(10000)
    
    # Criar lookup
    ref_lookup = {}
    for ref in all_ref_items:
        codigo = ref["codigo_item"]
        if codigo not in ref_lookup:
            ref_lookup[codigo] = []
        ref_lookup[codigo].append(ref)
    
    for po in pos:
        updated = False
        for item in po['items']:
            if not item.get('responsavel'):
                ref_items = ref_lookup.get(item["codigo_item"], [])
                
                if ref_items:
                    # Preferir não-admins
                    non_admin_items = [ri for ri in ref_items if ri.get('responsavel') in ['Maria', 'Mylena', 'Fabio']]
                    
                    if non_admin_items:
                        selected_ref = random.choice(non_admin_items)
                    else:
                        selected_ref = ref_items[0]
                    
                    item['responsavel'] = selected_ref.get('responsavel', '')
                    item['lote'] = selected_ref.get('lote', '')
                    item['lot_number'] = selected_ref.get('lot_number', 0)
                    total_fixed += 1
                    updated = True
        
        if updated:
            await db.purchase_orders.update_one(
                {"id": po['id']},
                {"$set": {"items": po['items']}}
            )
    
    return {"message": f"{total_fixed} itens corrigidos com sucesso"}

@api_router.post("/purchase-orders/fix-marca-modelo")
async def fix_marca_modelo(current_user: dict = Depends(require_admin)):
    """Corrigir marca/modelo faltantes nas OCs existentes (ADMIN ONLY)"""
    import random
    
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(1000)
    total_fixed = 0
    
    # OTIMIZAÇÃO: Coletar todos os codigos primeiro
    all_codigos = set()
    for po in pos:
        for item in po['items']:
            if not item.get('marca_modelo'):
                all_codigos.add(item["codigo_item"])
    
    # Buscar todas as referências de uma vez
    all_ref_items = await db.reference_items.find(
        {"codigo_item": {"$in": list(all_codigos)}},
        {"_id": 0}
    ).to_list(10000)
    
    # Criar lookup
    ref_lookup = {}
    for ref in all_ref_items:
        codigo = ref["codigo_item"]
        if codigo not in ref_lookup:
            ref_lookup[codigo] = []
        ref_lookup[codigo].append(ref)
    
    for po in pos:
        updated = False
        for item in po['items']:
            if not item.get('marca_modelo'):
                ref_items = ref_lookup.get(item["codigo_item"], [])
                
                if ref_items:
                    # Procurar uma referência com marca_modelo preenchido
                    for ref in ref_items:
                        marca = ref.get('marca_modelo', '')
                        if marca and marca != '#N/A':
                            item['marca_modelo'] = marca
                            total_fixed += 1
                            updated = True
                            break
        
        if updated:
            await db.purchase_orders.update_one(
                {"id": po['id']},
                {"$set": {"items": po['items']}}
            )
    
    return {"message": f"{total_fixed} itens corrigidos com marca/modelo"}

@api_router.post("/purchase-orders/normalize-fornecedores")
async def normalize_fornecedores(current_user: dict = Depends(require_admin)):
    """Normalizar fornecedores (maiúsculas, sem acentos) e unificar duplicados (ADMIN ONLY)"""
    import unicodedata
    
    def normalize_text(text):
        if not text:
            return ''
        # Remove acentos e converte para maiúsculas
        normalized = unicodedata.normalize('NFD', text)
        without_accents = ''.join(c for c in normalized if unicodedata.category(c) != 'Mn')
        return without_accents.upper().strip()
    
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(1000)
    total_normalized = 0
    
    for po in pos:
        updated = False
        for item in po['items']:
            fontes = item.get('fontes_compra', [])
            for fonte in fontes:
                fornecedor_original = fonte.get('fornecedor', '')
                fornecedor_normalizado = normalize_text(fornecedor_original)
                if fornecedor_original != fornecedor_normalizado:
                    fonte['fornecedor'] = fornecedor_normalizado
                    total_normalized += 1
                    updated = True
        
        if updated:
            await db.purchase_orders.update_one(
                {"id": po['id']},
                {"$set": {"items": po['items']}}
            )
    
    return {"message": f"{total_normalized} fornecedores normalizados com sucesso"}

@api_router.post("/purchase-orders/update-descriptions")
async def update_all_descriptions(current_user: dict = Depends(require_admin)):
    """Atualizar descrições de todos os itens com as descrições completas do Excel (ADMIN ONLY)"""
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(1000)
    total_updated = 0
    
    for po in pos:
        updated = False
        for item in po['items']:
            codigo_item = item.get('codigo_item')
            if codigo_item:
                # Buscar descrição completa no Excel
                ref_item = await db.reference_items.find_one(
                    {"codigo_item": codigo_item},
                    {"_id": 0, "descricao": 1}
                )
                if ref_item and ref_item.get('descricao'):
                    old_desc = item.get('descricao', '')
                    new_desc = ref_item['descricao']
                    # Atualizar se a descrição do Excel for diferente/mais completa
                    if old_desc != new_desc:
                        item['descricao'] = new_desc
                        total_updated += 1
                        updated = True
        
        if updated:
            await db.purchase_orders.update_one(
                {"id": po['id']},
                {"$set": {"items": po['items']}}
            )
    
    return {"message": f"{total_updated} descrições atualizadas com sucesso"}

@api_router.get("/backup/export")
async def export_backup(current_user: dict = Depends(require_admin)):
    """Exportar backup completo do sistema (ADMIN ONLY)"""
    from datetime import datetime
    
    # Buscar todos os dados COMPLETOS
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(10000)
    users = await db.users.find({}, {"_id": 0}).to_list(1000)  # Incluir tudo para restauração
    reference_items = await db.reference_items.find({}, {"_id": 0}).to_list(10000)
    notifications = await db.notifications.find({}, {"_id": 0}).to_list(10000)
    
    # Calcular estatísticas detalhadas
    total_items = 0
    status_counts = {}
    items_com_cotacao = 0
    items_com_link = 0
    total_valor_venda = 0
    
    for po in pos:
        for item in po.get('items', []):
            total_items += 1
            status = item.get('status', 'pendente')
            status_counts[status] = status_counts.get(status, 0) + 1
            
            # Contar cotações
            fontes = item.get('fontes_compra', [])
            if fontes and len(fontes) > 0:
                items_com_cotacao += 1
                for f in fontes:
                    if f.get('link'):
                        items_com_link += 1
            
            # Somar valor de venda
            preco_venda = item.get('preco_venda', 0) or 0
            quantidade = item.get('quantidade', 0) or 0
            total_valor_venda += preco_venda * quantidade
    
    backup = {
        "backup_info": {
            "data_export": datetime.now().isoformat(),
            "versao": "2.0",
            "sistema": "FIEP - Sistema de Gestão de OCs",
            "estatisticas": {
                "total_ocs": len(pos),
                "total_itens": total_items,
                "total_usuarios": len(users),
                "total_itens_referencia": len(reference_items),
                "total_notificacoes": len(notifications),
                "items_com_cotacao": items_com_cotacao,
                "items_com_link": items_com_link,
                "valor_total_venda": round(total_valor_venda, 2),
                "status_itens": status_counts
            }
        },
        "purchase_orders": pos,
        "users": users,
        "reference_items": reference_items,
        "notifications": notifications
    }
    
    return backup

@api_router.post("/backup/restore")
async def restore_backup(current_user: dict = Depends(require_admin)):
    """Restaurar backup do sistema - CUIDADO: substitui todos os dados! (ADMIN ONLY)"""
    # Este endpoint espera o JSON do backup no body
    # Será chamado via upload de arquivo no frontend
    pass

@api_router.post("/backup/restore-data")
async def restore_backup_data(backup_data: dict, current_user: dict = Depends(require_admin)):
    """Restaurar dados do backup (ADMIN ONLY) - SUBSTITUI TODOS OS DADOS!"""
    from datetime import datetime
    
    try:
        # Verificar se é um backup válido
        if "backup_info" not in backup_data or "purchase_orders" not in backup_data:
            raise HTTPException(status_code=400, detail="Arquivo de backup inválido")
        
        # Estatísticas antes
        ocs_antes = await db.purchase_orders.count_documents({})
        
        # Restaurar Purchase Orders
        if backup_data.get("purchase_orders"):
            # Limpar coleção existente
            await db.purchase_orders.delete_many({})
            # Inserir dados do backup
            if len(backup_data["purchase_orders"]) > 0:
                await db.purchase_orders.insert_many(backup_data["purchase_orders"])
        
        # Restaurar Reference Items (se existir no backup)
        if backup_data.get("reference_items") and len(backup_data["reference_items"]) > 0:
            await db.reference_items.delete_many({})
            await db.reference_items.insert_many(backup_data["reference_items"])
        
        # Restaurar Notifications (se existir no backup)
        if backup_data.get("notifications") and len(backup_data["notifications"]) > 0:
            await db.notifications.delete_many({})
            await db.notifications.insert_many(backup_data["notifications"])
        
        # NÃO restaurar usuários automaticamente (segurança)
        # Os usuários devem ser restaurados manualmente se necessário
        
        # Estatísticas depois
        ocs_depois = await db.purchase_orders.count_documents({})
        refs_depois = await db.reference_items.count_documents({})
        
        return {
            "success": True,
            "message": "Backup restaurado com sucesso!",
            "detalhes": {
                "data_backup": backup_data.get("backup_info", {}).get("data_export", "N/A"),
                "ocs_antes": ocs_antes,
                "ocs_restauradas": ocs_depois,
                "itens_referencia": refs_depois
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao restaurar backup: {str(e)}")

# ================== RASTREAMENTO CORREIOS ==================

import httpx
from services.correios_service import rastrear_objeto_correios, verificar_status_evento

async def buscar_rastreio_api(codigo: str) -> dict:
    """
    Busca rastreio usando a API oficial dos Correios (com fallback para API pública).
    """
    result = await rastrear_objeto_correios(codigo)
    
    if result.get('success'):
        return {
            "success": True,
            "eventos": result.get('eventos', []),
            "entregue": result.get('entregue', False),
            "saiu_para_entrega": result.get('saiu_para_entrega', False),
            "tentativa_entrega": result.get('tentativa_entrega', False),
            "fonte": result.get('fonte', 'correios'),
            "link_detalhes": result.get('link_detalhes', '')
        }
    
    # Retornar informações úteis quando falhar
    return {
        "success": False, 
        "eventos": [], 
        "message": result.get('error', "Não foi possível consultar o rastreio."),
        "rastreamento_manual": result.get('rastreamento_manual', False),
        "link_correios": f"https://rastreamento.correios.com.br/app/resultado.php?objeto={codigo}"
    }

# Endpoint para consultar rastreio (usando API dos Correios)
@api_router.get("/rastreio/{codigo}")
async def consultar_rastreio(codigo: str, current_user: dict = Depends(get_current_user)):
    """Consulta o rastreamento de um objeto na API dos Correios."""
    result = await buscar_rastreio_api(codigo)
    return {
        "codigo": codigo,
        **result
    }

# Endpoint para forçar verificação de todos os rastreios (admin only)
@api_router.post("/rastreio/verificar-todos")
async def verificar_todos_rastreios(current_user: dict = Depends(get_current_user)):
    """Força a verificação de todos os itens em trânsito. Apenas admin."""
    if current_user.get('role') != 'admin':
        raise HTTPException(status_code=403, detail="Apenas administradores podem executar esta ação")
    
    logger = logging.getLogger(__name__)
    logger.info("Verificação manual de rastreios solicitada por admin")
    
    # Executar verificação em background
    asyncio.create_task(_executar_verificacao_rastreios_uma_vez())
    
    return {"message": "Verificação de rastreios iniciada em background. As notificações aparecerão quando houver atualizações."}


async def _executar_verificacao_rastreios_uma_vez():
    """Executa uma única verificação de rastreios (para chamadas manuais)."""
    logger = logging.getLogger(__name__)
    
    try:
        logger.info("Iniciando verificação manual de rastreios (API Correios)...")
        
        cursor = db.purchase_orders.find(
            {"items.status": "em_transito"},
            {"_id": 0}
        )
        
        stats = {"verificados": 0, "entregues": 0, "saiu_entrega": 0, "tentativa": 0, "erros": 0}
        
        async for po in cursor:
            po_atualizado = False
            
            for item in po['items']:
                if item.get('status') == 'em_transito' and item.get('codigo_rastreio'):
                    codigo = item['codigo_rastreio']
                    stats["verificados"] += 1
                    
                    try:
                        result = await buscar_rastreio_api(codigo)
                        
                        if result.get('success') and result.get('eventos'):
                            eventos = result['eventos']
                            eventos_anteriores = item.get('rastreio_eventos', [])
                            qtd_anterior = len(eventos_anteriores)
                            item['rastreio_eventos'] = eventos
                            novos_eventos = len(eventos) > qtd_anterior
                            now = datetime.now(timezone.utc)
                            
                            if result.get('entregue'):
                                item['status'] = ItemStatus.ENTREGUE.value
                                item['data_entrega'] = now.isoformat()
                                stats["entregues"] += 1
                                po_atualizado = True
                                await _criar_notificacao_rastreio(po, item, "entrega", "✅ Item Entregue", f"O item {item['codigo_item']} foi entregue.")
                                logger.info(f"✅ Item {item['codigo_item']} ENTREGUE")
                            elif result.get('saiu_para_entrega') and novos_eventos and not item.get('notificado_saiu_entrega'):
                                stats["saiu_entrega"] += 1
                                item['notificado_saiu_entrega'] = True
                                po_atualizado = True
                                await _criar_notificacao_rastreio(po, item, "saiu_entrega", "🚚 Saiu para Entrega", f"O item {item['codigo_item']} saiu para entrega.")
                            elif result.get('tentativa_entrega') and novos_eventos:
                                stats["tentativa"] += 1
                                item['notificado_tentativa'] = True
                                po_atualizado = True
                                await _criar_notificacao_rastreio(po, item, "tentativa", "⚠️ Tentativa de Entrega", f"Tentativa de entrega: {item['codigo_item']}")
                            else:
                                po_atualizado = True
                                
                    except Exception as e:
                        stats["erros"] += 1
                        logger.warning(f"Erro ao verificar {codigo}: {e}")
            
            if po_atualizado:
                await db.purchase_orders.update_one({"id": po['id']}, {"$set": {"items": po['items']}})
        
        logger.info(f"Verificação manual concluída. Stats: {stats}")
        
    except Exception as e:
        logger.error(f"Erro na verificação manual: {e}")

# ===== ROTAS DE RASTREIO MOVIDAS PARA routes/rastreio_routes.py =====
# As rotas /rastreio/{codigo}, /items/{codigo_item}/rastreio, /items/{codigo_item}/atualizar-rastreio
# e /items/{codigo_item}/marcar-entregue foram refatoradas para o módulo routes/rastreio_routes.py

# ===== ROTAS DE NOTIFICAÇÕES MOVIDAS PARA routes/notificacao_routes.py =====

# ================== VERIFICAÇÃO AUTOMÁTICA DE RASTREIO ==================

async def verificar_rastreios_em_transito():
    """
    Verifica todos os itens em trânsito e atualiza status automaticamente.
    Cria notificações para: saiu para entrega, tentativa de entrega, entregue.
    Executado 1x ao dia.
    """
    logger = logging.getLogger(__name__)
    
    while True:
        try:
            logger.info("Iniciando verificação automática de rastreios (API Correios)...")
            
            # Buscar todas as OCs com itens em trânsito
            cursor = db.purchase_orders.find(
                {"items.status": "em_transito"},
                {"_id": 0}
            )
            
            stats = {
                "verificados": 0,
                "entregues": 0,
                "saiu_entrega": 0,
                "tentativa": 0,
                "erros": 0
            }
            
            async for po in cursor:
                po_atualizado = False
                
                for item in po['items']:
                    if item.get('status') == 'em_transito' and item.get('codigo_rastreio'):
                        codigo = item['codigo_rastreio']
                        stats["verificados"] += 1
                        
                        try:
                            # Buscar rastreio na API dos Correios
                            result = await buscar_rastreio_api(codigo)
                            
                            if result.get('success') and result.get('eventos'):
                                eventos = result['eventos']
                                
                                # Comparar com eventos anteriores para detectar novidades
                                eventos_anteriores = item.get('rastreio_eventos', [])
                                qtd_anterior = len(eventos_anteriores)
                                
                                # Atualizar eventos no item
                                item['rastreio_eventos'] = eventos
                                
                                # Verificar se há novos eventos
                                novos_eventos = len(eventos) > qtd_anterior
                                
                                now = datetime.now(timezone.utc)
                                
                                # Verificar se foi entregue
                                if result.get('entregue'):
                                    item['status'] = ItemStatus.ENTREGUE.value
                                    item['data_entrega'] = now.isoformat()
                                    stats["entregues"] += 1
                                    po_atualizado = True
                                    
                                    # Criar notificação de entrega
                                    await _criar_notificacao_rastreio(
                                        po, item, "entrega", "✅ Item Entregue",
                                        f"O item {item['codigo_item']} foi entregue ao destinatário."
                                    )
                                    
                                    logger.info(f"✅ Item {item['codigo_item']} da OC {po['numero_oc']} ENTREGUE")
                                
                                # Verificar se saiu para entrega (apenas se houver novos eventos)
                                elif result.get('saiu_para_entrega') and novos_eventos:
                                    if not item.get('notificado_saiu_entrega'):
                                        stats["saiu_entrega"] += 1
                                        item['notificado_saiu_entrega'] = True
                                        po_atualizado = True
                                        
                                        await _criar_notificacao_rastreio(
                                            po, item, "saiu_entrega", "🚚 Saiu para Entrega",
                                            f"O item {item['codigo_item']} saiu para entrega ao destinatário."
                                        )
                                        
                                        logger.info(f"🚚 Item {item['codigo_item']} da OC {po['numero_oc']} SAIU PARA ENTREGA")
                                
                                # Verificar tentativa de entrega (apenas se houver novos eventos)
                                elif result.get('tentativa_entrega') and novos_eventos:
                                    if not item.get('notificado_tentativa') or item.get('tentativas_count', 0) < len([e for e in eventos if 'ausente' in (e.get('status', '') or '').lower() or 'não entregue' in (e.get('status', '') or '').lower()]):
                                        stats["tentativa"] += 1
                                        item['notificado_tentativa'] = True
                                        item['tentativas_count'] = item.get('tentativas_count', 0) + 1
                                        po_atualizado = True
                                        
                                        ultimo_evento = eventos[0] if eventos else {}
                                        await _criar_notificacao_rastreio(
                                            po, item, "tentativa", "⚠️ Tentativa de Entrega",
                                            f"Tentativa de entrega para o item {item['codigo_item']}: {ultimo_evento.get('status', 'Destinatário ausente')}"
                                        )
                                        
                                        logger.info(f"⚠️ Item {item['codigo_item']} da OC {po['numero_oc']} TENTATIVA DE ENTREGA")
                                
                                # Atualizar eventos mesmo sem mudança de status
                                po_atualizado = True
                                
                        except Exception as e:
                            stats["erros"] += 1
                            logger.warning(f"Erro ao verificar rastreio {codigo}: {str(e)}")
                
                # Atualizar OC no banco se houve alterações
                if po_atualizado:
                    await db.purchase_orders.update_one(
                        {"id": po['id']},
                        {"$set": {"items": po['items']}}
                    )
            
            logger.info(
                f"Verificação concluída. "
                f"Verificados: {stats['verificados']}, "
                f"Entregues: {stats['entregues']}, "
                f"Saiu p/ Entrega: {stats['saiu_entrega']}, "
                f"Tentativas: {stats['tentativa']}, "
                f"Erros: {stats['erros']}"
            )
                
        except Exception as e:
            logger.error(f"Erro na verificação automática de rastreios: {str(e)}")
        
        # Aguardar 24 horas antes da próxima verificação (1x ao dia)
        await asyncio.sleep(86400)


async def _criar_notificacao_rastreio(po: dict, item: dict, tipo: str, titulo: str, mensagem: str):
    """Cria uma notificação de rastreio no banco de dados."""
    now = datetime.now(timezone.utc)
    
    notificacao = {
        "id": str(uuid.uuid4()),
        "tipo": tipo,
        "titulo": titulo,
        "mensagem": mensagem,
        "numero_oc": po.get('numero_oc', ''),
        "codigo_item": item.get('codigo_item', ''),
        "codigo_rastreio": item.get('codigo_rastreio', ''),
        "descricao_item": (item.get('descricao', '')[:50] + '...') if len(item.get('descricao', '')) > 50 else item.get('descricao', ''),
        "lida": False,
        "created_at": now.isoformat()
    }
    
    await db.notificacoes.insert_one(notificacao)

# ============== FUNÇÕES DE NOTAS FISCAIS ==============

def extract_ncm_from_xml(xml_content: str) -> Optional[str]:
    """Extrair TODOS os NCMs de um XML de Nota Fiscal Eletrônica (NFe)"""
    import xml.etree.ElementTree as ET
    try:
        # Remover BOM se existir
        if xml_content.startswith('\ufeff'):
            xml_content = xml_content[1:]
        
        root = ET.fromstring(xml_content)
        
        # Namespace padrão da NFe
        namespaces = {
            'nfe': 'http://www.portalfiscal.inf.br/nfe'
        }
        
        ncm_list = set()  # Usar set para evitar duplicatas
        
        # Tentar encontrar NCM com namespace
        ncm_elements = root.findall('.//nfe:NCM', namespaces)
        for elem in ncm_elements:
            if elem.text:
                ncm_list.add(elem.text.strip())
        
        # Tentar sem namespace
        if not ncm_list:
            ncm_elements = root.findall('.//NCM')
            for elem in ncm_elements:
                if elem.text:
                    ncm_list.add(elem.text.strip())
        
        # Tentar padrão alternativo
        if not ncm_list:
            for elem in root.iter():
                if elem.tag.endswith('NCM') and elem.text:
                    ncm_list.add(elem.text.strip())
        
        if ncm_list:
            return ', '.join(sorted(ncm_list))
        
        return None
    except Exception as e:
        logging.error(f"Erro ao extrair NCM do XML: {str(e)}")
        return None

def extract_ncm_from_pdf(pdf_bytes: bytes) -> Optional[str]:
    """Tentar extrair TODOS os NCMs de um PDF de Nota Fiscal"""
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        full_text = ""
        
        for page in doc:
            full_text += page.get_text()
        
        doc.close()
        
        ncm_list = set()  # Usar set para evitar duplicatas
        
        # Lista de valores que NÃO são NCM (CEPs, códigos de barras, etc.)
        invalid_ncms = {
            '17582008', '83005430',  # CEPs
            '83704614',  # Código de vendedor/outro
        }
        
        # Estratégia 1: Procurar padrão "NCM/SH" seguido diretamente por 8 dígitos (mais confiável)
        # Em DANFEs, o NCM aparece na coluna NCM/SH
        ncm_direct_pattern = r'NCM[/\s]*SH[:\s]*(\d{8})'
        matches = re.findall(ncm_direct_pattern, full_text, re.IGNORECASE)
        for m in matches:
            if len(m) == 8 and m.isdigit() and m not in invalid_ncms:
                ncm_list.add(m)
        
        # Estratégia 2: Procurar na estrutura típica de DANFE
        # O NCM geralmente aparece após o código do produto e antes do CFOP
        # Padrão: linha com NCM/SH seguida de 8 dígitos em formato específico
        ncm_table_pattern = r'(?:NCM|NCM/SH)\s*\n?\s*(\d{8})'
        matches = re.findall(ncm_table_pattern, full_text, re.IGNORECASE | re.MULTILINE)
        for m in matches:
            if len(m) == 8 and m.isdigit() and m not in invalid_ncms:
                # Verificar se os primeiros 2 dígitos são capítulos válidos de NCM
                capitulo = int(m[:2])
                # Capítulos comuns: 84, 85 (máquinas/elétricos), 73 (ferro/aço), etc.
                # Capítulos válidos: 01-97
                if 1 <= capitulo <= 97:
                    ncm_list.add(m)
        
        # Estratégia 3: Buscar números de 8 dígitos que começam com capítulos comuns
        # Capítulos mais comuns em compras: 84, 85, 73, 39, 40, 48, 90, etc.
        common_chapters = ['84', '85', '73', '39', '40', '48', '90', '94', '72', '76', '83']
        all_8digit = re.findall(r'\b(\d{8})\b', full_text)
        for num in all_8digit:
            if num[:2] in common_chapters and num not in invalid_ncms:
                ncm_list.add(num)
        
        # Estratégia 4: Padrões tradicionais com formatação
        patterns = [
            r'NCM[:\s]*(\d{4}[\.\s]?\d{2}[\.\s]?\d{2})',
            r'Classifica[çc][ãa]o Fiscal[:\s]*(\d{4}[\.\s]?\d{2}[\.\s]?\d{2})',
            r'(\d{4}\.\d{2}\.\d{2})',  # Padrão com pontos (mais específico)
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, full_text, re.IGNORECASE)
            for match in matches:
                ncm = re.sub(r'[\.\s]', '', match)
                if len(ncm) == 8 and ncm.isdigit() and ncm not in invalid_ncms:
                    capitulo = int(ncm[:2])
                    if 1 <= capitulo <= 97:
                        ncm_list.add(ncm)
        
        if ncm_list:
            # Ordenar e retornar
            return ', '.join(sorted(ncm_list))
        
        return None
    except Exception as e:
        logging.error(f"Erro ao extrair NCM do PDF: {str(e)}")
        return None


def extract_items_with_ncm_from_pdf(pdf_bytes: bytes) -> List[dict]:
    """Extrair lista de itens com seus respectivos NCMs do PDF de NF"""
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        full_text = ""
        for page in doc:
            full_text += page.get_text()
        doc.close()
        
        items = []
        invalid_ncms = {'17582008', '83005430', '83704614'}
        
        # Procurar seção de produtos
        start = full_text.find("DADOS DO PRODUTO")
        if start == -1:
            start = full_text.find("DADOS DOS PRODUTOS")
        
        if start != -1:
            # Pegar seção de produtos (até CÁLCULO ou DADOS ADICIONAIS)
            end_markers = ["CÁLCULO DO ISSQN", "DADOS ADICIONAIS", "INFORMAÇÕES COMPLEMENTARES"]
            end = len(full_text)
            for marker in end_markers:
                idx = full_text.find(marker, start)
                if idx != -1 and idx < end:
                    end = idx
            
            product_section = full_text[start:end]
            
            # Encontrar todos os NCMs de 8 dígitos na seção de produtos
            ncm_matches = list(re.finditer(r'\b(\d{8})\b', product_section))
            
            for match in ncm_matches:
                ncm = match.group(1)
                capitulo = int(ncm[:2])
                
                # Verificar se é um NCM válido (capítulos 01-97)
                if 1 <= capitulo <= 97 and ncm not in invalid_ncms:
                    # A descrição pode estar ANTES ou DEPOIS do NCM dependendo do layout
                    # Tentar primeiro DEPOIS do NCM (padrão mais comum em DANFEs)
                    after_text = product_section[match.end():match.end()+500]
                    lines_after = after_text.split('\n')[:8]
                    
                    descricao = ""
                    for line in lines_after:
                        line = line.strip()
                        # Descrição tem mais de 15 chars, contém letras e não é cabeçalho
                        if len(line) > 15 and re.search(r'[A-Za-z]', line):
                            if not any(header in line.upper() for header in ['DESCRIÇÃO', 'NCM/SH', 'CFOP', 'QUANT', 'VALOR', 'DADOS DO', 'CST', 'ALÍQUOTA']):
                                # Limpar a descrição (remover | Ped: etc)
                                descricao = re.sub(r'\s*\|\s*Ped:.*', '', line)
                                descricao = descricao[:80].strip()
                                break
                    
                    # Se não encontrou depois, tentar antes
                    if not descricao:
                        before_text = product_section[:match.start()]
                        lines_before = before_text.split('\n')[-10:]
                        
                        for line in reversed(lines_before):
                            line = line.strip()
                            if len(line) > 15 and re.search(r'[A-Za-z]', line):
                                if not any(header in line.upper() for header in ['DESCRIÇÃO', 'NCM/SH', 'CFOP', 'QUANT', 'VALOR', 'DADOS DO']):
                                    descricao = re.sub(r'\s*\|\s*Ped:.*', '', line)
                                    descricao = descricao[:80].strip()
                                    break
                    
                    if descricao:
                        items.append({
                            'descricao': descricao.upper(),
                            'ncm': ncm
                        })
        
        # Remover duplicatas mantendo ordem
        seen = set()
        unique_items = []
        for item in items:
            key = (item['descricao'], item['ncm'])
            if key not in seen:
                seen.add(key)
                unique_items.append(item)
        
        return unique_items
    except Exception as e:
        logging.error(f"Erro ao extrair itens do PDF: {str(e)}")
        return []


import base64

def extract_numero_nf_from_xml(xml_content: str) -> Optional[str]:
    """Extrair número da Nota Fiscal do XML"""
    import xml.etree.ElementTree as ET
    try:
        if xml_content.startswith('\ufeff'):
            xml_content = xml_content[1:]
        
        root = ET.fromstring(xml_content)
        
        namespaces = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
        
        # Tentar encontrar número da NF com namespace
        # nNF = número da nota fiscal
        for tag in ['nNF', 'nf', 'numero']:
            elements = root.findall(f'.//nfe:{tag}', namespaces)
            if elements and elements[0].text:
                return elements[0].text.strip()
            elements = root.findall(f'.//{tag}')
            if elements and elements[0].text:
                return elements[0].text.strip()
        
        # Tentar encontrar no texto
        for elem in root.iter():
            if elem.tag.endswith('nNF') and elem.text:
                return elem.text.strip()
        
        return None
    except Exception as e:
        logging.error(f"Erro ao extrair número da NF do XML: {str(e)}")
        return None

def extract_numero_nf_from_pdf(pdf_bytes: bytes) -> Optional[str]:
    """Tentar extrair número da NF de um PDF"""
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        full_text = ""
        
        for page in doc:
            full_text += page.get_text()
        
        doc.close()
        
        # Padrões comuns para número de NF
        patterns = [
            r'N[°º]?\s*(?:da\s*)?(?:NF|Nota|Nota Fiscal)[:\s]*(\d{6,9})',
            r'(?:NF|Nota Fiscal)\s*N[°º]?\s*(\d{6,9})',
            r'N[úu]mero[:\s]*(\d{6,9})',
            r'(\d{9})',  # Número de 9 dígitos solto
        ]
        
        for pattern in patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                num = match.group(1)
                if len(num) >= 6:
                    return num
        
        return None
    except Exception as e:
        logging.error(f"Erro ao extrair número da NF do PDF: {str(e)}")
        return None

class NFUploadRequest(BaseModel):
    """Request para upload de nota fiscal"""
    filename: str
    content_type: str
    file_data: str  # Base64
    ncm_manual: Optional[str] = None
    tipo: str  # "fornecedor" ou "revenda"

class NFUpdateNCMRequest(BaseModel):
    """Request para atualizar NCM manualmente"""
    ncm: str

class NFEmitidaRequest(BaseModel):
    """Request para marcar NF como emitida/pronto para despacho"""
    nf_emitida_pronto_despacho: bool

@api_router.post("/purchase-orders/{po_id}/items/by-index/{item_index}/notas-fiscais")
async def upload_nota_fiscal(
    po_id: str,
    item_index: int,
    request: NFUploadRequest,
    current_user: dict = Depends(get_current_user)
):
    """Upload de nota fiscal para um item"""
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    if item_index < 0 or item_index >= len(po['items']):
        raise HTTPException(status_code=404, detail="Índice de item inválido")
    
    item = po['items'][item_index]
    
    # Tentar extrair NCM, número da NF e itens
    ncm = request.ncm_manual
    numero_nf = None
    itens_nf = []  # Lista de itens com NCMs
    try:
        file_bytes = base64.b64decode(request.file_data)
        
        if request.content_type == 'text/xml' or request.filename.endswith('.xml'):
            xml_content = file_bytes.decode('utf-8')
            if not ncm:
                ncm = extract_ncm_from_xml(xml_content)
            numero_nf = extract_numero_nf_from_xml(xml_content)
        elif request.content_type == 'application/pdf' or request.filename.endswith('.pdf'):
            if not ncm:
                ncm = extract_ncm_from_pdf(file_bytes)
            numero_nf = extract_numero_nf_from_pdf(file_bytes)
            # Extrair itens com NCMs
            itens_nf = extract_items_with_ncm_from_pdf(file_bytes)
    except Exception as e:
        logging.error(f"Erro ao processar arquivo: {str(e)}")
    
    # Criar documento da NF
    nf_doc = {
        "id": str(uuid.uuid4()),
        "filename": request.filename,
        "content_type": request.content_type,
        "file_data": request.file_data,
        "ncm": ncm,
        "numero_nf": numero_nf,  # Número da NF extraído
        "itens_nf": itens_nf,  # Lista de itens com NCMs
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "uploaded_by": current_user.get('sub')
    }
    
    if request.tipo == "fornecedor":
        # Adicionar às NFs de fornecedor (múltiplas)
        if 'notas_fiscais_fornecedor' not in item:
            item['notas_fiscais_fornecedor'] = []
        item['notas_fiscais_fornecedor'].append(nf_doc)
        logging.info(f"NF Fornecedor adicionada - po_id: {po_id}, item_index: {item_index}, total NFs: {len(item['notas_fiscais_fornecedor'])}")
    elif request.tipo == "revenda":
        # Substituir NF de revenda (única)
        item['nota_fiscal_revenda'] = nf_doc
        logging.info(f"NF Revenda adicionada - po_id: {po_id}, item_index: {item_index}")
    else:
        raise HTTPException(status_code=400, detail="Tipo deve ser 'fornecedor' ou 'revenda'")
    
    # Atualizar o item específico no array usando arrayFilters ou posicional
    result = await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": po['items']}}
    )
    logging.info(f"Update result - matched: {result.matched_count}, modified: {result.modified_count}")
    
    return {
        "success": True,
        "message": "Nota fiscal adicionada com sucesso",
        "nf_id": nf_doc["id"],
        "ncm": ncm or "NCM NAO ENCONTRADO",
        "itens_nf": itens_nf
    }


# ============== ENDPOINTS PARA NF DE VENDA DA OC (não do item) ==============

class NFVendaOCRequest(BaseModel):
    filename: str
    content_type: str
    file_data: str  # Base64
    itens_indices: Optional[List[int]] = None  # Índices dos itens incluídos na NF


@api_router.post("/purchase-orders/{po_id}/nf-venda")
async def add_nf_venda_oc(
    po_id: str,
    request: NFVendaOCRequest,
    current_user: dict = Depends(get_current_user)
):
    """Adicionar NF de Venda para a OC (pode ser parcial com itens selecionados)"""
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    # Extrair número da NF do PDF se aplicável
    numero_nf = None
    if request.filename.lower().endswith('.pdf'):
        try:
            import base64
            pdf_bytes = base64.b64decode(request.file_data)
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            text = ""
            for page in doc:
                text += page.get_text()
            doc.close()
            
            nf_patterns = [
                r'NF[- ]?e?[:\s]*(\d{6,})',
                r'N[úu]mero[:\s]*(\d{6,})',
                r'NOTA FISCAL[:\s]*(\d{6,})'
            ]
            for pattern in nf_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    numero_nf = match.group(1)
                    break
        except Exception as e:
            logger.warning(f"Erro ao extrair número da NF: {e}")
    
    # Se itens_indices foi fornecido, usar; senão, incluir todos os itens
    items = po.get('items', [])
    total_itens = len(items)
    
    # Determinar quais itens serão incluídos nesta NF
    if request.itens_indices is not None and len(request.itens_indices) > 0:
        itens_nf = request.itens_indices
    else:
        # Incluir todos os itens que estão em "em_separacao"
        itens_nf = [i for i, item in enumerate(items) if item.get('status') == 'em_separacao']
    
    nf_doc = {
        "id": str(uuid.uuid4()),
        "filename": request.filename,
        "content_type": request.content_type,
        "file_data": request.file_data,
        "numero_nf": numero_nf,
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "uploaded_by": current_user.get('sub'),
        "itens_indices": itens_nf  # Guardar quais itens estão na NF
    }
    
    # Mesclar com NF existente se houver (para permitir múltiplas NFs parciais)
    existing_nfs = po.get('notas_fiscais_venda', [])
    existing_nfs.append(nf_doc)
    
    # Também manter retrocompatibilidade com nota_fiscal_venda (última NF)
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {
            "nota_fiscal_venda": nf_doc,
            "notas_fiscais_venda": existing_nfs
        }}
    )
    
    return {
        "success": True,
        "message": f"NF de Venda adicionada para {len(itens_nf)} item(s)",
        "nf_id": nf_doc["id"],
        "numero_nf": numero_nf,
        "itens_incluidos": len(itens_nf),
        "total_itens_oc": total_itens
    }


@api_router.get("/purchase-orders/{po_id}/nf-venda/download")
async def download_nf_venda_oc(
    po_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Download da NF de Venda da OC"""
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    nf_venda = po.get('nota_fiscal_venda')
    if not nf_venda:
        raise HTTPException(status_code=404, detail="NF de Venda não encontrada")
    
    return {
        "filename": nf_venda['filename'],
        "content_type": nf_venda['content_type'],
        "file_data": nf_venda['file_data']
    }


@api_router.delete("/purchase-orders/{po_id}/nf-venda")
async def delete_nf_venda_oc(
    po_id: str,
    nf_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Remover NF de Venda da OC (específica ou todas)"""
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    if nf_id:
        # Remover NF específica da lista
        existing_nfs = po.get('notas_fiscais_venda', [])
        existing_nfs = [nf for nf in existing_nfs if nf.get('id') != nf_id]
        
        # Atualizar nota_fiscal_venda para a última NF ou None
        last_nf = existing_nfs[-1] if existing_nfs else None
        
        await db.purchase_orders.update_one(
            {"id": po_id},
            {"$set": {
                "notas_fiscais_venda": existing_nfs,
                "nota_fiscal_venda": last_nf
            }}
        )
    else:
        # Remover todas as NFs
        await db.purchase_orders.update_one(
            {"id": po_id},
            {"$unset": {"nota_fiscal_venda": "", "notas_fiscais_venda": ""}}
        )
    
    return {"success": True, "message": "NF de Venda removida"}


class ProntoDespachoOCRequest(BaseModel):
    pronto_despacho: bool


@api_router.patch("/purchase-orders/{po_id}/pronto-despacho")
async def toggle_pronto_despacho_oc(
    po_id: str,
    request: ProntoDespachoOCRequest,
    current_user: dict = Depends(get_current_user)
):
    """Marcar OC inteira como Pronta para Despacho"""
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"pronto_despacho": request.pronto_despacho}}
    )
    
    return {"success": True, "pronto_despacho": request.pronto_despacho}


@api_router.get("/purchase-orders/{po_id}/items/by-index/{item_index}/notas-fiscais/{nf_id}/download")
async def download_nota_fiscal(
    po_id: str,
    item_index: int,
    nf_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Download de uma nota fiscal específica"""
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    if item_index < 0 or item_index >= len(po['items']):
        raise HTTPException(status_code=404, detail="Índice de item inválido")
    
    item = po['items'][item_index]
    
    # Procurar NF de fornecedor
    for nf in item.get('notas_fiscais_fornecedor', []):
        if nf['id'] == nf_id:
            return {
                "filename": nf['filename'],
                "content_type": nf['content_type'],
                "file_data": nf['file_data']
            }
    
    # Procurar NF de revenda
    nf_revenda = item.get('nota_fiscal_revenda')
    if nf_revenda and nf_revenda['id'] == nf_id:
        return {
            "filename": nf_revenda['filename'],
            "content_type": nf_revenda['content_type'],
            "file_data": nf_revenda['file_data']
        }
    
    raise HTTPException(status_code=404, detail="Nota fiscal não encontrada")


class BulkDownloadRequest(BaseModel):
    nfs: List[dict]  # Lista de {po_id, item_index, nf_id, tipo}


@api_router.post("/admin/notas-fiscais/bulk-download")
async def bulk_download_notas_fiscais(
    request: BulkDownloadRequest,
    current_user: dict = Depends(require_admin)
):
    """Download de múltiplas notas fiscais em formato ZIP"""
    import zipfile
    import io
    import base64
    
    if not request.nfs:
        raise HTTPException(status_code=400, detail="Nenhuma NF selecionada")
    
    # Criar arquivo ZIP em memória
    zip_buffer = io.BytesIO()
    user_name = current_user.get('owner_name', current_user.get('email', 'admin'))
    downloaded_at = datetime.now(timezone.utc).isoformat()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for nf_info in request.nfs:
            po_id = nf_info.get('po_id')
            item_index = nf_info.get('item_index')
            nf_id = nf_info.get('nf_id')
            tipo = nf_info.get('tipo', 'fornecedor')
            
            po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
            if not po or item_index >= len(po['items']):
                continue
            
            item = po['items'][item_index]
            nf_data = None
            
            if tipo == 'fornecedor':
                for nf in item.get('notas_fiscais_fornecedor', []):
                    if nf['id'] == nf_id:
                        nf_data = nf
                        break
            else:
                nf_revenda = item.get('nota_fiscal_revenda')
                if nf_revenda and nf_revenda['id'] == nf_id:
                    nf_data = nf_revenda
            
            if nf_data:
                # Decodificar arquivo base64
                file_bytes = base64.b64decode(nf_data['file_data'])
                filename = f"{po['numero_oc']}_{item.get('codigo_item', 'item')}_{nf_data['filename']}"
                zip_file.writestr(filename, file_bytes)
                
                # Marcar NF como baixada
                if tipo == 'fornecedor':
                    for idx, nf in enumerate(item.get('notas_fiscais_fornecedor', [])):
                        if nf['id'] == nf_id:
                            await db.purchase_orders.update_one(
                                {"id": po_id},
                                {"$set": {
                                    f"items.{item_index}.notas_fiscais_fornecedor.{idx}.baixado_por": user_name,
                                    f"items.{item_index}.notas_fiscais_fornecedor.{idx}.baixado_em": downloaded_at
                                }}
                            )
                            break
                else:
                    await db.purchase_orders.update_one(
                        {"id": po_id},
                        {"$set": {
                            f"items.{item_index}.nota_fiscal_revenda.baixado_por": user_name,
                            f"items.{item_index}.nota_fiscal_revenda.baixado_em": downloaded_at
                        }}
                    )
    
    zip_buffer.seek(0)
    zip_base64 = base64.b64encode(zip_buffer.read()).decode('utf-8')
    
    return {
        "filename": f"notas_fiscais_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
        "content_type": "application/zip",
        "file_data": zip_base64,
        "total_nfs": len(request.nfs)
    }


@api_router.delete("/purchase-orders/{po_id}/items/by-index/{item_index}/notas-fiscais/{nf_id}")
async def delete_nota_fiscal(
    po_id: str,
    item_index: int,
    nf_id: str,
    tipo: str,  # Query param: "fornecedor" ou "revenda"
    current_user: dict = Depends(get_current_user)
):
    """Deletar uma nota fiscal específica"""
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    if item_index < 0 or item_index >= len(po['items']):
        raise HTTPException(status_code=404, detail="Índice de item inválido")
    
    item = po['items'][item_index]
    
    if tipo == "fornecedor":
        # Remover das NFs de fornecedor
        if 'notas_fiscais_fornecedor' in item:
            item['notas_fiscais_fornecedor'] = [
                nf for nf in item['notas_fiscais_fornecedor'] if nf['id'] != nf_id
            ]
    elif tipo == "revenda":
        # Remover NF de revenda
        if item.get('nota_fiscal_revenda') and item['nota_fiscal_revenda']['id'] == nf_id:
            item['nota_fiscal_revenda'] = None
    else:
        raise HTTPException(status_code=400, detail="Tipo deve ser 'fornecedor' ou 'revenda'")
    
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": po['items']}}
    )
    
    return {"success": True, "message": "Nota fiscal removida com sucesso"}

@api_router.patch("/purchase-orders/{po_id}/items/by-index/{item_index}/notas-fiscais/{nf_id}/ncm")
async def update_nf_ncm(
    po_id: str,
    item_index: int,
    nf_id: str,
    request: NFUpdateNCMRequest,
    current_user: dict = Depends(get_current_user)
):
    """Atualizar NCM de uma nota fiscal manualmente"""
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    if item_index < 0 or item_index >= len(po['items']):
        raise HTTPException(status_code=404, detail="Índice de item inválido")
    
    item = po['items'][item_index]
    updated = False
    
    # Procurar e atualizar NF de fornecedor
    for nf in item.get('notas_fiscais_fornecedor', []):
        if nf['id'] == nf_id:
            nf['ncm'] = request.ncm.upper()
            updated = True
            break
    
    # Procurar e atualizar NF de revenda
    nf_revenda = item.get('nota_fiscal_revenda')
    if not updated and nf_revenda and nf_revenda['id'] == nf_id:
        nf_revenda['ncm'] = request.ncm.upper()
        updated = True
    
    if not updated:
        raise HTTPException(status_code=404, detail="Nota fiscal não encontrada")
    
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": po['items']}}
    )
    
    return {"success": True, "message": "NCM atualizado com sucesso"}

@api_router.patch("/purchase-orders/{po_id}/items/by-index/{item_index}/nf-emitida")
async def update_nf_emitida_status(
    po_id: str,
    item_index: int,
    request: NFEmitidaRequest,
    current_user: dict = Depends(get_current_user)
):
    """Atualizar status de NF emitida/pronto para despacho"""
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    if item_index < 0 or item_index >= len(po['items']):
        raise HTTPException(status_code=404, detail="Índice de item inválido")
    
    item = po['items'][item_index]
    item['nf_emitida_pronto_despacho'] = request.nf_emitida_pronto_despacho
    
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": po['items']}}
    )
    
    return {"success": True, "message": "Status atualizado com sucesso"}

@api_router.patch("/purchase-orders/{po_id}/items/by-index/{item_index}/endereco-entrega")
async def update_endereco_entrega(
    po_id: str,
    item_index: int,
    endereco: str = "",
    current_user: dict = Depends(get_current_user)
):
    """Atualizar endereço de entrega de um item manualmente"""
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    if item_index < 0 or item_index >= len(po['items']):
        raise HTTPException(status_code=404, detail="Índice de item inválido")
    
    item = po['items'][item_index]
    item['endereco_entrega'] = endereco.upper()
    
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": po['items']}}
    )
    
    return {"success": True, "message": "Endereço atualizado com sucesso"}

class EnderecoRequest(BaseModel):
    endereco: str

@api_router.patch("/purchase-orders/{po_id}/items/by-index/{item_index}/endereco")
async def update_endereco_entrega_v2(
    po_id: str,
    item_index: int,
    request: EnderecoRequest,
    current_user: dict = Depends(get_current_user)
):
    """Atualizar endereço de entrega de um item manualmente (v2 com body)"""
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    if item_index < 0 or item_index >= len(po['items']):
        raise HTTPException(status_code=404, detail="Índice de item inválido")
    
    item = po['items'][item_index]
    item['endereco_entrega'] = request.endereco.upper()
    
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": po['items']}}
    )
    
    return {"success": True, "message": "Endereço atualizado com sucesso"}


# ============== ENDPOINTS DE IMAGEM DE ITEM ==============

UPLOAD_DIR = Path(__file__).parent / "uploads" / "item_images"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# Tipos de imagem permitidos
ALLOWED_IMAGE_TYPES = ["image/jpeg", "image/png", "image/webp", "image/gif"]
MAX_IMAGE_SIZE = 5 * 1024 * 1024  # 5MB


@api_router.post("/purchase-orders/{po_id}/items/by-index/{item_index}/imagem")
async def upload_item_image(
    po_id: str,
    item_index: int,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Upload de imagem para um item.
    Aceita: JPEG, PNG, WebP, GIF (máx 5MB)
    A imagem é salva no servidor e vinculada ao CÓDIGO do item.
    Todos os itens com o mesmo código compartilharão a mesma imagem.
    """
    logger.info(f"Upload de imagem recebido: po_id={po_id}, item_index={item_index}, filename={file.filename}, content_type={file.content_type}")
    
    # Verificar tipo de arquivo
    if file.content_type not in ALLOWED_IMAGE_TYPES:
        logger.warning(f"Tipo de arquivo rejeitado: {file.content_type}")
        raise HTTPException(
            status_code=400, 
            detail=f"Tipo de arquivo não permitido. Use: {', '.join(ALLOWED_IMAGE_TYPES)}"
        )
    
    # Ler arquivo
    contents = await file.read()
    logger.info(f"Arquivo lido: {len(contents)} bytes")
    
    # Verificar tamanho mínimo (evitar imagens corrompidas/vazias)
    if len(contents) < 1000:
        logger.warning(f"Arquivo muito pequeno: {len(contents)} bytes - possível imagem corrompida")
        raise HTTPException(status_code=400, detail="Arquivo muito pequeno ou corrompido. Envie uma imagem válida.")
    
    # Verificar tamanho máximo
    if len(contents) > MAX_IMAGE_SIZE:
        raise HTTPException(status_code=400, detail="Arquivo muito grande. Máximo: 5MB")
    
    # Buscar OC
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    if item_index < 0 or item_index >= len(po['items']):
        raise HTTPException(status_code=404, detail="Índice de item inválido")
    
    item = po['items'][item_index]
    codigo_item = item.get('codigo_item', 'unknown')
    
    # Converter imagem para base64 para salvar no MongoDB (PERSISTENTE)
    import base64
    imagem_base64 = base64.b64encode(contents).decode('utf-8')
    
    # Determinar content type
    ext = file.filename.split('.')[-1].lower() if '.' in file.filename else 'jpg'
    content_types = {
        'jpg': 'image/jpeg',
        'jpeg': 'image/jpeg',
        'png': 'image/png',
        'webp': 'image/webp',
        'gif': 'image/gif'
    }
    content_type = content_types.get(ext, 'image/jpeg')
    
    # Gerar URL única para servir a imagem
    unique_id = str(uuid.uuid4())[:8]
    imagem_url = f"/api/item-images-db/{codigo_item}"
    
    # Salvar na coleção de imagens por código (PERSISTENTE NO MONGODB)
    await db.imagens_itens.update_one(
        {"codigo_item": codigo_item},
        {
            "$set": {
                "codigo_item": codigo_item,
                "imagem_url": imagem_url,
                "imagem_base64": imagem_base64,  # SALVAR BASE64 NO BANCO
                "content_type": content_type,
                "tamanho_bytes": len(contents),
                "data_upload": datetime.now(timezone.utc).isoformat(),
                "uploaded_by": current_user.get('email')
            }
        },
        upsert=True
    )
    
    # Atualizar TODOS os itens com este código em TODAS as OCs
    result = await db.purchase_orders.update_many(
        {"items.codigo_item": codigo_item},
        {
            "$set": {
                "items.$[elem].imagem_url": imagem_url
            }
        },
        array_filters=[{"elem.codigo_item": codigo_item}]
    )
    
    logger.info(f"Imagem salva no MongoDB para código {codigo_item}: {len(contents)} bytes - {result.modified_count} OCs atualizadas")
    
    return {
        "success": True,
        "imagem_url": imagem_url,
        "tamanho_bytes": len(contents),
        "ocs_atualizadas": result.modified_count,
        "message": f"Imagem salva permanentemente para todos os itens com código {codigo_item}"
    }


@api_router.patch("/purchase-orders/{po_id}/items/by-index/{item_index}/copiar-imagem")
async def copiar_imagem_item(
    po_id: str,
    item_index: int,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Copia uma imagem já existente para outro item.
    Usado para propagar imagem entre itens com mesmo código em diferentes OCs.
    """
    imagem_url = data.get('imagem_url')
    imagem_filename = data.get('imagem_filename')
    
    if not imagem_url:
        raise HTTPException(status_code=400, detail="imagem_url é obrigatório")
    
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    if item_index < 0 or item_index >= len(po['items']):
        raise HTTPException(status_code=404, detail="Índice de item inválido")
    
    item = po['items'][item_index]
    item['imagem_url'] = imagem_url
    if imagem_filename:
        item['imagem_filename'] = imagem_filename
    
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": po['items']}}
    )
    
    logger.info(f"Imagem copiada para item {item.get('codigo_item')} na OC {po.get('numero_oc')}")
    
    return {"success": True, "message": "Imagem copiada com sucesso"}


@api_router.get("/itens/imagens-disponiveis")
async def get_itens_imagens_disponiveis(current_user: dict = Depends(get_current_user)):
    """
    Retorna lista de códigos de itens que possuem imagem.
    """
    cursor = db.imagens_itens.find({}, {"_id": 0, "codigo_item": 1})
    imagens = await cursor.to_list(10000)
    
    codigos = [item['codigo_item'] for item in imagens if item.get('codigo_item')]
    
    return {"codigos": codigos}


@api_router.get("/imagens-itens/mapa")
async def get_mapa_imagens_itens(current_user: dict = Depends(get_current_user)):
    """
    Retorna um mapa de código_item -> imagem_url para todos os itens com imagem.
    Usado pelo frontend para exibir imagens rapidamente.
    """
    cursor = db.imagens_itens.find({}, {"_id": 0, "codigo_item": 1, "imagem_url": 1})
    imagens = await cursor.to_list(10000)
    
    return {item['codigo_item']: item['imagem_url'] for item in imagens}


@api_router.delete("/purchase-orders/{po_id}/items/by-index/{item_index}/imagem")
async def delete_item_image(
    po_id: str,
    item_index: int,
    current_user: dict = Depends(get_current_user)
):
    """
    Remove a imagem de um item.
    A imagem é removida da coleção imagens_itens e de TODOS os itens com o mesmo código.
    """
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Ordem de Compra não encontrada")
    
    if item_index < 0 or item_index >= len(po['items']):
        raise HTTPException(status_code=404, detail="Índice de item inválido")
    
    item = po['items'][item_index]
    codigo_item = item.get('codigo_item', 'unknown')
    
    # Buscar imagem da coleção imagens_itens
    imagem_info = await db.imagens_itens.find_one({"codigo_item": codigo_item}, {"_id": 0})
    
    # Remover arquivo físico se existir
    filename = item.get('imagem_filename') or (imagem_info.get('imagem_filename') if imagem_info else None)
    if filename:
        filepath = UPLOAD_DIR / filename
        if filepath.exists():
            filepath.unlink()
            logger.info(f"Arquivo de imagem removido: {filename}")
    
    # Remover da coleção imagens_itens
    await db.imagens_itens.delete_one({"codigo_item": codigo_item})
    
    # Limpar imagem de TODOS os itens com este código em TODAS as OCs
    result = await db.purchase_orders.update_many(
        {"items.codigo_item": codigo_item},
        {
            "$set": {
                "items.$[elem].imagem_url": None,
                "items.$[elem].imagem_filename": None
            }
        },
        array_filters=[{"elem.codigo_item": codigo_item}]
    )
    
    logger.info(f"Imagem removida para código {codigo_item}: {result.modified_count} OCs atualizadas")
    
    return {
        "success": True, 
        "message": f"Imagem removida de todos os itens com código {codigo_item}",
        "ocs_atualizadas": result.modified_count
    }


# NOVO ENDPOINT: Servir imagens do MongoDB (PERSISTENTE)
@api_router.get("/item-images-db/{codigo_item}")
async def get_item_image_from_db(codigo_item: str):
    """Servir imagem de item diretamente do MongoDB (persistente)"""
    import base64
    from fastapi.responses import Response
    
    # Buscar imagem no MongoDB
    imagem_doc = await db.imagens_itens.find_one(
        {"codigo_item": codigo_item},
        {"_id": 0, "imagem_base64": 1, "content_type": 1}
    )
    
    if not imagem_doc or not imagem_doc.get('imagem_base64'):
        raise HTTPException(status_code=404, detail="Imagem não encontrada")
    
    # Decodificar base64
    imagem_bytes = base64.b64decode(imagem_doc['imagem_base64'])
    content_type = imagem_doc.get('content_type', 'image/jpeg')
    
    return Response(
        content=imagem_bytes,
        media_type=content_type,
        headers={
            "Cache-Control": "public, max-age=31536000",  # Cache por 1 ano (imagem não muda)
            "Content-Length": str(len(imagem_bytes))
        }
    )


# Manter endpoint antigo para compatibilidade (arquivos em disco)
@api_router.get("/item-images/{filename}")
@api_router.head("/item-images/{filename}")
async def get_item_image(filename: str):
    """Servir imagem de item do disco (legado - para compatibilidade)"""
    filepath = UPLOAD_DIR / filename
    
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Imagem não encontrada")
    
    # Determinar content type
    ext = filename.split('.')[-1].lower()
    content_types = {
        'jpg': 'image/jpeg',
        'jpeg': 'image/jpeg',
        'png': 'image/png',
        'webp': 'image/webp',
        'gif': 'image/gif'
    }
    content_type = content_types.get(ext, 'image/jpeg')
    
    # Adicionar headers de cache para garantir atualização
    return FileResponse(
        filepath, 
        media_type=content_type,
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0"
        }
    )



# ============== ENDPOINTS DE LIMITES DO CONTRATO (PLANILHA) ==============

@api_router.post("/admin/importar-limites-contrato")
async def importar_limites_contrato(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_admin)
):
    """
    Importa a planilha de limites do contrato FIEP.
    A planilha deve ter:
    - Coluna J (10): Código do item
    - Coluna H (8): Quantidade máxima permitida no contrato
    
    Os limites são armazenados na coleção 'limites_contrato'.
    """
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser Excel (.xlsx ou .xls)")
    
    try:
        import openpyxl
        from io import BytesIO
        
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        sheet = wb.active
        
        # Processar planilha
        limites_por_codigo = {}
        linhas_processadas = 0
        
        for row in range(2, sheet.max_row + 1):
            codigo = sheet.cell(row=row, column=10).value  # Coluna J
            qtd_maxima = sheet.cell(row=row, column=8).value  # Coluna H
            
            if codigo and qtd_maxima:
                codigo_str = str(codigo).strip()
                try:
                    qtd = int(qtd_maxima)
                    if codigo_str not in limites_por_codigo:
                        limites_por_codigo[codigo_str] = 0
                    limites_por_codigo[codigo_str] += qtd
                    linhas_processadas += 1
                except (ValueError, TypeError):
                    continue
        
        # Limpar coleção existente e inserir novos limites
        await db.limites_contrato.delete_many({})
        
        documentos = [
            {
                'codigo_item': codigo,
                'quantidade_maxima_contrato': quantidade,
                'data_importacao': datetime.now(timezone.utc).isoformat()
            }
            for codigo, quantidade in limites_por_codigo.items()
        ]
        
        if documentos:
            await db.limites_contrato.insert_many(documentos)
        
        logger.info(f"Importados limites para {len(limites_por_codigo)} códigos de itens")
        
        return {
            "success": True,
            "itens_importados": len(limites_por_codigo),
            "linhas_processadas": linhas_processadas,
            "mensagem": f"Limites do contrato importados para {len(limites_por_codigo)} códigos de itens"
        }
        
    except Exception as e:
        logger.error(f"Erro ao importar limites: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao processar planilha: {str(e)}")


@api_router.get("/limites-contrato")
async def listar_limites_contrato(current_user: dict = Depends(get_current_user)):
    """Lista todos os limites de contrato importados"""
    limites = await db.limites_contrato.find({}, {"_id": 0}).to_list(10000)
    return {
        "total": len(limites),
        "limites": limites
    }


@api_router.get("/limites-contrato/mapa")
async def get_limites_contrato_mapa(current_user: dict = Depends(get_current_user)):
    """
    Retorna um mapa de código_item -> quantidade_maxima_contrato
    Usado pelo frontend para exibir os limites rapidamente
    """
    limites = await db.limites_contrato.find({}, {"_id": 0}).to_list(10000)
    mapa = {item['codigo_item']: item['quantidade_maxima_contrato'] for item in limites}
    return mapa


# ============== ENDPOINTS ADMIN - COMISSÕES E NOTAS FISCAIS ==============

class ComissaoUpdate(BaseModel):
    """Atualizar comissão de um responsável"""
    percentual: float
    pago: bool = False

@api_router.get("/admin/comissoes")
async def get_comissoes(current_user: dict = Depends(require_admin)):
    """Obter dados de comissões por responsável
    
    Sistema de Comissões:
    - Comissão fixa de 1.5% sobre o VALOR TOTAL DA VENDA
    - Apenas para itens com status "entregue" ou "em_transito"
    - Lógica híbrida:
      1. Itens com lote numérico (ex: "Lote 42") → Usa mapeamento fixo de lotes por pessoa
      2. Itens sem lote numérico (ex: "CHAMAMENTO PÚBLICO...") → Usa campo responsavel do item
    """
    
    # Mapeamento fixo de LOTES por pessoa (baseado na cotação original)
    # Formato: "Lote XX" onde XX é o número
    LOTES_POR_PESSOA = {
        'MARIA': [1,2,3,4,5,6,7,8,9,10,11,12,43,44,45,46,47,48,49,50,51,52,53],
        'MYLENA': [80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97],
        'FABIO': [32,33,34,35,36,37,38,39,40,41,42],
    }
    
    # OCs que foram cotadas por ADMIN (João/Mateus) e não devem gerar comissão
    # mesmo que o lote esteja na lista de alguém
    OCS_EXCLUIDAS_COMISSAO = ['OC-2.118938', 'OC-2.118941']
    
    # Comissão fixa de 1.5%
    PERCENTUAL_COMISSAO = 1.5
    
    def extrair_numero_lote(lote_str):
        """Extrai o número do lote de strings como 'Lote 36' ou '36'"""
        if not lote_str:
            return None
        import re
        # Só considera como lote numérico se tiver o padrão "Lote XX" ou similar
        match = re.search(r'[Ll]ote\s*(\d+)', str(lote_str))
        if match:
            return int(match.group(1))
        # Também aceita só o número se for curto (ex: "42")
        if len(str(lote_str).strip()) <= 5:
            match = re.search(r'^(\d+)$', str(lote_str).strip())
            if match:
                return int(match.group(1))
        return None
    
    def normalizar_responsavel(resp):
        """Normaliza o nome do responsável para maiúsculas"""
        if not resp:
            return None
        return resp.strip().upper()
    
    # Obter todas as OCs
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(length=1000)
    
    # Dicionários para acumular valores por pessoa
    valor_venda_por_pessoa = {}
    itens_por_pessoa = {}
    lotes_por_pessoa = {}
    
    # Inicializar com as pessoas que têm lotes fixos
    for nome in LOTES_POR_PESSOA.keys():
        valor_venda_por_pessoa[nome] = 0
        itens_por_pessoa[nome] = []
        lotes_por_pessoa[nome] = LOTES_POR_PESSOA[nome].copy()
    
    for po in pos:
        numero_oc = po.get('numero_oc', '')
        
        # Pular OCs que foram cotadas por admin
        if numero_oc in OCS_EXCLUIDAS_COMISSAO:
            continue
            
        for item in po.get('items', []):
            item_status = item.get('status', '')
            # Apenas itens "entregue" ou "em_transito" geram comissão
            if item_status not in ['entregue', 'em_transito']:
                continue
            
            # Calcular valor total de venda do item
            preco_venda = item.get('preco_venda', 0) or 0
            quantidade = item.get('quantidade', 1) or 1
            valor_total_venda = preco_venda * quantidade
            
            lote_str = item.get('lote', '')
            numero_lote = extrair_numero_lote(lote_str)
            
            pessoa_responsavel = None
            
            # LÓGICA 1: Se tem lote numérico, usar mapeamento fixo
            if numero_lote is not None:
                for pessoa, lotes in LOTES_POR_PESSOA.items():
                    if numero_lote in lotes:
                        pessoa_responsavel = pessoa
                        break
            
            # LÓGICA 2: Se não tem lote numérico, usar campo responsavel
            if pessoa_responsavel is None:
                responsavel_item = normalizar_responsavel(item.get('responsavel', ''))
                if responsavel_item and responsavel_item not in ['JOÃO', 'MATEUS', 'ADMIN']:
                    pessoa_responsavel = responsavel_item
            
            # Se encontrou um responsável, contabilizar
            if pessoa_responsavel:
                # Inicializar se for pessoa nova
                if pessoa_responsavel not in valor_venda_por_pessoa:
                    valor_venda_por_pessoa[pessoa_responsavel] = 0
                    itens_por_pessoa[pessoa_responsavel] = []
                    lotes_por_pessoa[pessoa_responsavel] = []
                
                valor_venda_por_pessoa[pessoa_responsavel] += valor_total_venda
                itens_por_pessoa[pessoa_responsavel].append({
                    'numero_oc': numero_oc,
                    'codigo_item': item.get('codigo_item'),
                    'lote': lote_str,
                    'valor_venda': valor_total_venda,
                    'status': item_status,
                    'fonte': 'lote' if numero_lote is not None else 'responsavel'
                })
    
    # Montar resposta
    resultado = []
    for pessoa, valor_venda in valor_venda_por_pessoa.items():
        valor_comissao = valor_venda * (PERCENTUAL_COMISSAO / 100)
        resultado.append({
            'responsavel': pessoa,
            'email': None,
            'valor_venda_total': valor_venda,
            'percentual_comissao': PERCENTUAL_COMISSAO,
            'valor_comissao': valor_comissao,
            'lotes_atribuidos': lotes_por_pessoa.get(pessoa, []),
            'qtd_itens': len(itens_por_pessoa[pessoa])
        })
    
    return sorted(resultado, key=lambda x: x['valor_venda_total'], reverse=True)

@api_router.patch("/admin/comissoes/{responsavel}")
async def update_comissao(
    responsavel: str,
    request: ComissaoUpdate,
    current_user: dict = Depends(require_admin)
):
    """Atualizar comissão de um responsável"""
    
    await db.comissoes.update_one(
        {"responsavel": responsavel.upper()},
        {"$set": {
            "responsavel": responsavel.upper(),
            "percentual": request.percentual,
            "pago": request.pago
        }},
        upsert=True
    )
    
    return {"success": True, "message": "Comissão atualizada com sucesso"}

@api_router.get("/admin/notas-fiscais")
async def get_todas_notas_fiscais(current_user: dict = Depends(require_admin)):
    """Obter todas as notas fiscais do sistema (compra e venda)"""
    
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(length=1000)
    
    nfs_compra = []  # NFs de fornecedor
    nfs_venda = []   # NFs de revenda (ON)
    
    # Para detectar duplicatas, usamos o filename como chave
    nf_por_filename = {}  # filename -> lista de itens que usam essa NF
    
    for po in pos:
        for idx, item in enumerate(po.get('items', [])):
            # NFs de compra (fornecedor)
            for nf in item.get('notas_fiscais_fornecedor', []):
                filename = nf.get('filename', '')
                
                # Rastrear quais itens usam cada NF
                if filename not in nf_por_filename:
                    nf_por_filename[filename] = []
                nf_por_filename[filename].append({
                    'numero_oc': po.get('numero_oc'),
                    'codigo_item': item.get('codigo_item')
                })
                
                nfs_compra.append({
                    'id': nf.get('id'),
                    'filename': filename,
                    'content_type': nf.get('content_type'),
                    'ncm': nf.get('ncm'),
                    'numero_nf': nf.get('numero_nf'),  # Número da NF
                    'uploaded_at': nf.get('uploaded_at'),
                    'numero_oc': po.get('numero_oc'),
                    'codigo_item': item.get('codigo_item'),
                    'descricao': item.get('descricao', '')[:50],
                    'po_id': po.get('id'),
                    'item_index': idx,
                    'baixado_por': nf.get('baixado_por'),
                    'baixado_em': nf.get('baixado_em')
                })
            
            # NF de venda (revenda)
            nf_revenda = item.get('nota_fiscal_revenda')
            if nf_revenda:
                nfs_venda.append({
                    'id': nf_revenda.get('id'),
                    'filename': nf_revenda.get('filename'),
                    'content_type': nf_revenda.get('content_type'),
                    'ncm': nf_revenda.get('ncm'),
                    'numero_nf': nf_revenda.get('numero_nf'),  # Número da NF
                    'uploaded_at': nf_revenda.get('uploaded_at'),
                    'numero_oc': po.get('numero_oc'),
                    'codigo_item': item.get('codigo_item'),
                    'descricao': item.get('descricao', '')[:50],
                    'po_id': po.get('id'),
                    'item_index': idx,
                    'baixado_por': nf_revenda.get('baixado_por'),
                    'baixado_em': nf_revenda.get('baixado_em')
                })
    
    # Marcar NFs que são usadas em múltiplos itens
    for nf in nfs_compra:
        filename = nf.get('filename', '')
        itens_usando = nf_por_filename.get(filename, [])
        nf['duplicada'] = len(itens_usando) > 1
        nf['itens_usando'] = itens_usando if len(itens_usando) > 1 else []
        nf['qtd_usos'] = len(itens_usando)
    
    # Criar lista de NFs duplicadas (únicas por filename)
    nfs_duplicadas = []
    filenames_vistos = set()
    for nf in nfs_compra:
        if nf.get('duplicada') and nf['filename'] not in filenames_vistos:
            filenames_vistos.add(nf['filename'])
            nfs_duplicadas.append({
                'filename': nf['filename'],
                'numero_nf': nf.get('numero_nf'),
                'ncm': nf.get('ncm'),
                'qtd_usos': nf['qtd_usos'],
                'itens': nf['itens_usando']
            })
    
    # Ordenar por data de upload (mais recente primeiro)
    nfs_compra.sort(key=lambda x: x.get('uploaded_at', ''), reverse=True)
    nfs_venda.sort(key=lambda x: x.get('uploaded_at', ''), reverse=True)
    
    return {
        'notas_compra': nfs_compra,
        'notas_venda': nfs_venda,
        'total_compra': len(nfs_compra),
        'total_venda': len(nfs_venda),
        'notas_duplicadas': nfs_duplicadas,
        'total_duplicadas': len(nfs_duplicadas)
    }

@api_router.get("/admin/itens-responsavel/{responsavel}")
async def get_itens_responsavel(responsavel: str, current_user: dict = Depends(require_admin)):
    """Obter todos os itens entregues/em_transito de uma pessoa
    
    Sistema de Comissões com lógica híbrida:
    1. Itens com lote numérico (ex: "Lote 42") → Usa mapeamento fixo de lotes por pessoa
    2. Itens sem lote numérico (ex: "CHAMAMENTO PÚBLICO...") → Usa campo responsavel do item
    """
    
    # Mapeamento fixo de LOTES por pessoa (baseado na cotação original)
    LOTES_POR_PESSOA = {
        'MARIA': [1,2,3,4,5,6,7,8,9,10,11,12,43,44,45,46,47,48,49,50,51,52,53],
        'MYLENA': [80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97],
        'FABIO': [32,33,34,35,36,37,38,39,40,41,42],
    }
    
    # OCs que foram cotadas por ADMIN (João/Mateus) e não devem gerar comissão
    OCS_EXCLUIDAS_COMISSAO = ['OC-2.118938', 'OC-2.118941']
    
    # Comissão fixa de 1.5%
    PERCENTUAL_COMISSAO = 1.5
    
    def extrair_numero_lote(lote_str):
        """Extrai o número do lote de strings como 'Lote 36' ou '36'"""
        if not lote_str:
            return None
        import re
        # Só considera como lote numérico se tiver o padrão "Lote XX" ou similar
        match = re.search(r'[Ll]ote\s*(\d+)', str(lote_str))
        if match:
            return int(match.group(1))
        # Também aceita só o número se for curto (ex: "42")
        if len(str(lote_str).strip()) <= 5:
            match = re.search(r'^(\d+)$', str(lote_str).strip())
            if match:
                return int(match.group(1))
        return None
    
    def normalizar_responsavel(resp):
        """Normaliza o nome do responsável para maiúsculas"""
        if not resp:
            return None
        return resp.strip().upper()
    
    responsavel_upper = responsavel.upper().strip()
    
    # Verificar se é um responsável do mapeamento de lotes
    lotes_do_responsavel = LOTES_POR_PESSOA.get(responsavel_upper, [])
    
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(length=1000)
    
    # Obter IDs de itens já pagos
    pagamentos = await db.pagamentos.find({}, {"_id": 0}).to_list(length=1000)
    itens_pagos = set()
    for pag in pagamentos:
        for item_id in pag.get('itens_ids', []):
            itens_pagos.add(item_id)
    
    itens = []
    for po in pos:
        numero_oc = po.get('numero_oc', '')
        
        # Pular OCs que foram cotadas por admin
        if numero_oc in OCS_EXCLUIDAS_COMISSAO:
            continue
            
        for idx, item in enumerate(po.get('items', [])):
            item_status = item.get('status', '')
            # Apenas itens "entregue" ou "em_transito" geram comissão
            if item_status not in ['entregue', 'em_transito']:
                continue
            
            lote_str = item.get('lote', '')
            numero_lote = extrair_numero_lote(lote_str)
            
            pertence_ao_responsavel = False
            
            # LÓGICA 1: Se tem lote numérico, usar mapeamento fixo
            if numero_lote is not None:
                if numero_lote in lotes_do_responsavel:
                    pertence_ao_responsavel = True
            else:
                # LÓGICA 2: Se não tem lote numérico, usar campo responsavel
                responsavel_item = normalizar_responsavel(item.get('responsavel', ''))
                if responsavel_item == responsavel_upper and responsavel_upper not in ['JOÃO', 'MATEUS', 'ADMIN']:
                    pertence_ao_responsavel = True
            
            if not pertence_ao_responsavel:
                continue
            
            # Calcular valor total de venda
            preco_venda = item.get('preco_venda', 0) or 0
            quantidade = item.get('quantidade', 1) or 1
            valor_total_venda = preco_venda * quantidade
            valor_comissao = valor_total_venda * (PERCENTUAL_COMISSAO / 100)
            
            item_id = f"{po.get('id')}_{idx}"
            itens.append({
                'id': item_id,
                'po_id': po.get('id'),
                'item_index': idx,
                'numero_oc': numero_oc,
                'codigo_item': item.get('codigo_item'),
                'descricao': (item.get('descricao', '') or '')[:30],
                'lote': item.get('lote'),
                'valor_venda': valor_total_venda,
                'valor_comissao': valor_comissao,
                'status': item_status,
                'data_entrega': item.get('data_entrega'),
                'pago': item_id in itens_pagos
            })
    
    # Ordenar por data de entrega (mais recente primeiro)
    itens.sort(key=lambda x: x.get('data_entrega', '') or '', reverse=True)
    
    return itens

class PagamentoCreate(BaseModel):
    """Criar pagamento de comissão"""
    responsavel: str
    itens_ids: List[str]
    percentual: float
    valor_comissao: float
    total_venda: float  # Valor total de venda (1.5% sobre este valor)

@api_router.get("/admin/pagamentos")
async def get_pagamentos(current_user: dict = Depends(require_admin)):
    """Obter histórico de pagamentos"""
    pagamentos = await db.pagamentos.find({}, {"_id": 0}).to_list(length=1000)
    # Ordenar por data (mais recente primeiro)
    pagamentos.sort(key=lambda x: x.get('data', ''), reverse=True)
    return pagamentos

@api_router.post("/admin/pagamentos")
async def create_pagamento(request: PagamentoCreate, current_user: dict = Depends(require_admin)):
    """Registrar pagamento de comissão"""
    
    pagamento = {
        "id": str(uuid.uuid4()),
        "responsavel": request.responsavel.upper(),
        "itens_ids": request.itens_ids,
        "percentual": request.percentual,
        "valor_comissao": request.valor_comissao,
        "total_venda": request.total_venda,
        "qtd_itens": len(request.itens_ids),
        "data": datetime.now(timezone.utc).isoformat(),
        "pago_por": current_user.get('sub')
    }
    
    await db.pagamentos.insert_one(pagamento)
    
    return {"success": True, "message": "Pagamento registrado com sucesso", "pagamento_id": pagamento["id"]}

class PagamentoUpdate(BaseModel):
    """Atualizar pagamento"""
    valor_comissao: Optional[float] = None
    percentual: Optional[float] = None

@api_router.patch("/admin/pagamentos/{pagamento_id}")
async def update_pagamento(
    pagamento_id: str,
    request: PagamentoUpdate,
    current_user: dict = Depends(require_admin)
):
    """Editar um pagamento existente"""
    
    update_data = {}
    if request.valor_comissao is not None:
        update_data["valor_comissao"] = request.valor_comissao
    if request.percentual is not None:
        update_data["percentual"] = request.percentual
    
    if not update_data:
        raise HTTPException(status_code=400, detail="Nenhum campo para atualizar")
    
    result = await db.pagamentos.update_one(
        {"id": pagamento_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Pagamento não encontrado")
    
    return {"success": True, "message": "Pagamento atualizado com sucesso"}

@api_router.delete("/admin/pagamentos/{pagamento_id}")
async def delete_pagamento(
    pagamento_id: str,
    current_user: dict = Depends(require_admin)
):
    """Deletar um pagamento"""
    
    result = await db.pagamentos.delete_one({"id": pagamento_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pagamento não encontrado")
    
    return {"success": True, "message": "Pagamento deletado com sucesso"}


@api_router.get("/fornecedores")
async def get_fornecedores(current_user: dict = Depends(get_current_user)):
    """Buscar lista de fornecedores únicos do sistema"""
    
    pos = await db.purchase_orders.find({}, {"_id": 0, "items.fontes_compra.fornecedor": 1}).to_list(1000)
    
    fornecedores = set()
    for po in pos:
        for item in po.get('items', []):
            for fonte in item.get('fontes_compra', []):
                fornecedor = fonte.get('fornecedor', '').strip()
                if fornecedor:
                    fornecedores.add(fornecedor)
    
    return {"fornecedores": sorted(list(fornecedores))}


@api_router.patch("/purchase-orders/{po_id}/endereco-entrega")
async def update_po_endereco_entrega(
    po_id: str,
    data: dict,
    current_user: dict = Depends(require_admin)
):
    """Atualizar endereço de entrega da OC (apenas admin)"""
    
    endereco = data.get("endereco_entrega", "").strip().upper()
    
    result = await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"endereco_entrega": endereco}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    
    return {"success": True, "endereco_entrega": endereco}


@api_router.post("/purchase-orders/{po_id}/status-em-massa")
async def atualizar_status_em_massa(
    po_id: str,
    data: dict,
    current_user: dict = Depends(require_admin)
):
    """
    Atualizar o status de itens selecionados de uma OC.
    
    Body:
    {
        "novo_status": "entregue",  // pendente, cotado, comprado, em_separacao, em_transito, entregue
        "item_indices": [0, 1, 2]   // Opcional: Se não fornecido, atualiza todos os itens
    }
    """
    
    novo_status = data.get("novo_status", "").strip().lower()
    item_indices = data.get("item_indices")  # Lista de índices ou None
    
    status_validos = ["pendente", "cotado", "comprado", "em_separacao", "em_transito", "entregue"]
    
    if novo_status not in status_validos:
        raise HTTPException(status_code=400, detail=f"Status inválido. Use: {', '.join(status_validos)}")
    
    # Buscar OC
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    
    # Atualizar status dos itens (selecionados ou todos)
    items = po.get('items', [])
    itens_atualizados = 0
    
    # Definir categorias de status
    status_antes_compra = ['pendente', 'cotado']
    status_apos_compra = ['comprado', 'em_separacao', 'em_transito', 'entregue']
    
    # Se item_indices foi fornecido, usar apenas esses índices
    if item_indices is not None:
        indices_set = set(item_indices)
        for idx, item in enumerate(items):
            if idx in indices_set:
                status_anterior = item.get('status', 'pendente')
                if status_anterior != novo_status:
                    # VERIFICAR SE PRECISA REVERTER USO DE ESTOQUE
                    if novo_status in status_antes_compra and status_anterior in status_apos_compra:
                        if item.get('estoque_origem') or item.get('atendido_por_estoque'):
                            logger.info(f"Item {idx} voltando para {novo_status} via status em massa, revertendo uso de estoque...")
                            await reverter_uso_estoque(item, po_id, po.get('numero_oc', ''))
                    
                    item['status'] = novo_status
                    atualizar_data_compra(item, novo_status)  # Atualiza data de compra automaticamente
                    itens_atualizados += 1
    else:
        # Atualizar todos os itens
        for idx, item in enumerate(items):
            status_anterior = item.get('status', 'pendente')
            if status_anterior != novo_status:
                # VERIFICAR SE PRECISA REVERTER USO DE ESTOQUE
                if novo_status in status_antes_compra and status_anterior in status_apos_compra:
                    if item.get('estoque_origem') or item.get('atendido_por_estoque'):
                        logger.info(f"Item {idx} voltando para {novo_status} via status em massa (todos), revertendo uso de estoque...")
                        await reverter_uso_estoque(item, po_id, po.get('numero_oc', ''))
                
                item['status'] = novo_status
                atualizar_data_compra(item, novo_status)  # Atualiza data de compra automaticamente
                itens_atualizados += 1
    
    # Salvar alterações
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": items}}
    )
    
    return {
        "success": True,
        "numero_oc": po.get('numero_oc'),
        "novo_status": novo_status,
        "itens_atualizados": itens_atualizados,
        "total_itens": len(items)
    }


@api_router.post("/purchase-orders/{po_id}/frete-envio-multiplo")
async def aplicar_frete_envio_multiplo(
    po_id: str,
    data: dict,
    current_user: dict = Depends(require_admin)
):
    """
    Aplicar frete de envio dividido entre múltiplos itens.
    
    Body:
    {
        "item_indices": [0, 1, 2],  // Índices dos itens
        "frete_total": 150.00       // Valor total do frete a ser dividido
    }
    
    O frete será dividido igualmente entre os itens selecionados.
    """
    
    item_indices = data.get("item_indices", [])
    frete_total = data.get("frete_total", 0)
    
    if not item_indices:
        raise HTTPException(status_code=400, detail="Nenhum item selecionado")
    
    if frete_total <= 0:
        raise HTTPException(status_code=400, detail="Valor do frete deve ser maior que zero")
    
    # Buscar OC
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    
    # Calcular frete por item (dividido igualmente)
    frete_por_item = round(frete_total / len(item_indices), 2)
    
    # Atualizar cada item
    items = po.get('items', [])
    itens_atualizados = []
    
    for idx in item_indices:
        if idx < 0 or idx >= len(items):
            continue
        
        items[idx]['frete_envio'] = frete_por_item
        itens_atualizados.append({
            "indice": idx,
            "codigo_item": items[idx].get('codigo_item', ''),
            "frete_envio": frete_por_item
        })
        
        # Recalcular lucro líquido usando a função centralizada
        calcular_lucro_item(items[idx])
    
    # Salvar alterações
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": items}}
    )
    
    return {
        "success": True,
        "frete_total": frete_total,
        "frete_por_item": frete_por_item,
        "itens_atualizados": itens_atualizados,
        "quantidade_itens": len(itens_atualizados)
    }


@api_router.post("/purchase-orders/{po_id}/rastreio-multiplo")
async def aplicar_rastreio_multiplo(
    po_id: str,
    data: dict,
    current_user: dict = Depends(require_admin)
):
    """
    Aplicar código de rastreio para múltiplos itens de uma OC.
    
    Body:
    {
        "item_indices": [0, 1, 2],           // Índices dos itens
        "codigo_rastreio": "AB123456789BR"   // Código de rastreio dos Correios
    }
    """
    
    item_indices = data.get("item_indices", [])
    codigo_rastreio = data.get("codigo_rastreio", "").strip()
    
    if not item_indices:
        raise HTTPException(status_code=400, detail="Nenhum item selecionado")
    
    if not codigo_rastreio:
        raise HTTPException(status_code=400, detail="Código de rastreio não informado")
    
    # Buscar OC
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    
    # Atualizar cada item com o código de rastreio
    items = po.get('items', [])
    itens_atualizados = []
    
    for idx in item_indices:
        if idx < 0 or idx >= len(items):
            continue
        
        items[idx]['codigo_rastreio'] = codigo_rastreio
        itens_atualizados.append({
            "indice": idx,
            "codigo_item": items[idx].get('codigo_item', ''),
            "codigo_rastreio": codigo_rastreio
        })
    
    # Salvar alterações
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": items}}
    )
    
    return {
        "success": True,
        "codigo_rastreio": codigo_rastreio,
        "itens_atualizados": itens_atualizados,
        "quantidade_itens": len(itens_atualizados)
    }


@api_router.post("/purchase-orders/{po_id}/rastreio-frete-multiplo")
async def aplicar_rastreio_frete_multiplo(
    po_id: str,
    data: dict,
    current_user: dict = Depends(require_admin)
):
    """
    Aplicar código de rastreio E/OU frete para múltiplos itens de uma OC.
    Usado para atualizar/corrigir dados de itens em trânsito.
    
    Body:
    {
        "item_indices": [0, 1, 2],           // Índices dos itens
        "codigo_rastreio": "AB123456789BR", // Código de rastreio (opcional)
        "frete_por_item": 15.50              // Frete por item já calculado (opcional)
    }
    """
    
    item_indices = data.get("item_indices", [])
    codigo_rastreio = data.get("codigo_rastreio", "").strip() if data.get("codigo_rastreio") else None
    frete_por_item = data.get("frete_por_item")
    
    if not item_indices:
        raise HTTPException(status_code=400, detail="Nenhum item selecionado")
    
    if not codigo_rastreio and frete_por_item is None:
        raise HTTPException(status_code=400, detail="Informe código de rastreio e/ou valor do frete")
    
    # Buscar OC
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    
    # Atualizar cada item
    items = po.get('items', [])
    itens_atualizados = []
    
    for idx in item_indices:
        if idx < 0 or idx >= len(items):
            continue
        
        updates = {}
        if codigo_rastreio:
            items[idx]['codigo_rastreio'] = codigo_rastreio
            updates['codigo_rastreio'] = codigo_rastreio
        
        if frete_por_item is not None:
            items[idx]['frete_envio'] = float(frete_por_item)
            updates['frete_envio'] = float(frete_por_item)
        
        itens_atualizados.append({
            "indice": idx,
            "codigo_item": items[idx].get('codigo_item', ''),
            **updates
        })
    
    # Salvar alterações
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": items}}
    )
    
    return {
        "success": True,
        "codigo_rastreio": codigo_rastreio,
        "frete_por_item": frete_por_item,
        "itens_atualizados": itens_atualizados,
        "quantidade_itens": len(itens_atualizados)
    }


@api_router.post("/purchase-orders/{po_id}/status-multiplo")
async def aplicar_status_multiplo(
    po_id: str,
    data: dict,
    current_user: dict = Depends(require_admin)
):
    """
    Aplicar mudança de status para múltiplos itens de uma OC.
    
    Body:
    {
        "item_indices": [0, 1, 2],     // Índices dos itens
        "novo_status": "em_transito"   // Novo status
    }
    """
    from datetime import datetime, timezone
    
    item_indices = data.get("item_indices", [])
    novo_status = data.get("novo_status", "").strip()
    
    if not item_indices:
        raise HTTPException(status_code=400, detail="Nenhum item selecionado")
    
    status_validos = ['pendente', 'cotado', 'comprado', 'em_separacao', 'em_transito', 'entregue']
    if novo_status not in status_validos:
        raise HTTPException(status_code=400, detail=f"Status inválido. Use: {', '.join(status_validos)}")
    
    # Buscar OC
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    
    # Atualizar cada item com o novo status
    items = po.get('items', [])
    itens_atualizados = []
    
    for idx in item_indices:
        if idx < 0 or idx >= len(items):
            continue
        
        items[idx]['status'] = novo_status
        
        # Se mudou para 'comprado', registrar data_compra
        if novo_status == 'comprado' and not items[idx].get('data_compra'):
            items[idx]['data_compra'] = datetime.now(timezone.utc).isoformat()
        
        itens_atualizados.append({
            "indice": idx,
            "codigo_item": items[idx].get('codigo_item', ''),
            "novo_status": novo_status
        })
    
    # Salvar alterações
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": items}}
    )
    
    return {
        "success": True,
        "novo_status": novo_status,
        "itens_atualizados": itens_atualizados,
        "quantidade_itens": len(itens_atualizados)
    }


@api_router.post("/buscar-cep")
async def buscar_cep_endpoint(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Busca CEP pelo endereço usando a API ViaCEP.
    
    Body:
    {
        "endereco": "AVENIDA COMENDADOR FRANCO, 1341, JARDIM BOTANICO, CURITIBA"
    }
    
    Retorna:
    {
        "success": true,
        "cep": "81520-000",
        "cep_numerico": "81520000"
    }
    """
    endereco = data.get("endereco", "")
    
    if not endereco:
        raise HTTPException(status_code=400, detail="Endereço não informado")
    
    cep = buscar_cep_por_endereco(endereco)
    
    if cep:
        return {
            "success": True,
            "cep": cep,
            "cep_numerico": cep.replace("-", "")
        }
    else:
        return {
            "success": False,
            "message": "CEP não encontrado para este endereço"
        }


@api_router.post("/purchase-orders/{po_id}/atualizar-pdf")
async def atualizar_oc_com_pdf(
    po_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(require_admin)
):
    """
    Atualizar uma OC existente com dados de um novo PDF.
    
    Esta função PRESERVA todos os dados importantes dos itens:
    - Status (pendente, cotado, comprado, etc.)
    - Responsável
    - Fontes de compra (fornecedor, preço, link)
    - Notas fiscais anexadas
    - Observações
    - Frete, valor unitário de venda, etc.
    
    Atualiza apenas:
    - Endereço de entrega (se estava vazio)
    - Data de entrega (se estava vazia)
    - Outros campos do cabeçalho da OC
    """
    
    # Buscar OC existente
    existing_po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not existing_po:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    
    # Processar o PDF
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Arquivo deve ser PDF")
    
    content = await file.read()
    
    try:
        pdf_doc = fitz.open(stream=content, filetype="pdf")
        full_text = ""
        for page in pdf_doc:
            full_text += page.get_text()
        pdf_doc.close()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler PDF: {str(e)}")
    
    # Extrair dados do PDF
    # Endereço de Entrega
    endereco_patterns = [
        r'Endere[çc]o de Entrega[:\s]*(.*?)(?:\n\n|Linha|Item)',
        r'Local de Entrega[:\s]*(.*?)(?:\n\n|Linha|Item)',
        r'Entregar em[:\s]*(.*?)(?:\n\n|Linha|Item)'
    ]
    
    novo_endereco = ""
    for pattern in endereco_patterns:
        endereco_match = re.search(pattern, full_text, re.IGNORECASE | re.DOTALL)
        if endereco_match:
            novo_endereco = endereco_match.group(1).strip()
            novo_endereco = ' '.join(novo_endereco.split()).upper()
            break
    
    # Data de Entrega
    nova_data_entrega = None
    data_patterns = [
        r'Data de Entrega[:\s]*(\d{2}/\d{2}/\d{4})',
        r'Data Entrega[:\s]*(\d{2}/\d{2}/\d{4})',
        r'Entrega[:\s]*(\d{2}/\d{2}/\d{4})',
        r'Prazo de Entrega[:\s]*(\d{2}/\d{2}/\d{4})',
        r'Dt\.\s*Entrega[:\s]*(\d{2}/\d{2}/\d{4})'
    ]
    
    for pattern in data_patterns:
        data_match = re.search(pattern, full_text, re.IGNORECASE)
        if data_match:
            try:
                data_str = data_match.group(1)
                dia, mes, ano = data_str.split('/')
                nova_data_entrega = f"{ano}-{mes}-{dia}"
                break
            except:
                pass
    
    # Se não encontrou, buscar na tabela de itens
    if not nova_data_entrega:
        date_after_req = re.findall(r'\d{1,2}\.\d{2}\.\d{6,}\s*(\d{2}/\d{2}/\d{4})', full_text)
        if date_after_req:
            try:
                data_str = date_after_req[0]
                dia, mes, ano = data_str.split('/')
                nova_data_entrega = f"{ano}-{mes}-{dia}"
            except:
                pass
    
    # Preparar atualizações - SÓ atualiza campos vazios ou inexistentes
    updates = {}
    campos_atualizados = []
    
    # Atualizar endereço - SEMPRE atualiza se o PDF tiver o endereço
    endereco_atual = existing_po.get('endereco_entrega', '').strip()
    if novo_endereco:
        # Adicionar CEP ao endereço se não tiver
        if not re.search(r'CEP[:\s]*\d{5}-?\d{3}', novo_endereco, re.IGNORECASE):
            cep = buscar_cep_por_endereco(novo_endereco)
            if cep:
                novo_endereco = f"{novo_endereco}, CEP: {cep}"
        
        if novo_endereco != endereco_atual:
            updates['endereco_entrega'] = novo_endereco
            campos_atualizados.append(f"endereco_entrega: {novo_endereco}")
    
    # Atualizar data de entrega - SEMPRE atualiza se o PDF tiver a data
    data_atual = existing_po.get('data_entrega', '').strip() if existing_po.get('data_entrega') else ''
    if nova_data_entrega and nova_data_entrega != data_atual:
        updates['data_entrega'] = nova_data_entrega
        campos_atualizados.append(f"data_entrega: {nova_data_entrega}")
    
    # Retornar informações de debug também
    debug_info = {
        "endereco_extraido_pdf": novo_endereco or "Não encontrado no PDF",
        "data_extraida_pdf": nova_data_entrega or "Não encontrada no PDF",
        "endereco_atual_oc": endereco_atual or "Vazio",
        "data_atual_oc": data_atual or "Vazia"
    }
    
    if not updates:
        return {
            "success": True,
            "message": "Nenhum campo precisou ser atualizado (todos já estão preenchidos ou não foram encontrados no PDF)",
            "numero_oc": existing_po['numero_oc'],
            "campos_atualizados": [],
            "debug": debug_info
        }
    
    # Aplicar atualizações
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": updates}
    )
    
    return {
        "success": True,
        "message": f"OC {existing_po['numero_oc']} atualizada com sucesso!",
        "numero_oc": existing_po['numero_oc'],
        "campos_atualizados": campos_atualizados,
        "debug": debug_info,
        "dados_preservados": [
            "Status de todos os itens",
            "Responsáveis dos itens",
            "Fontes de compra",
            "Notas fiscais",
            "Observações",
            "Valores de frete e venda"
        ]
    }


@api_router.post("/admin/atualizar-ceps-enderecos")
async def atualizar_ceps_todos_enderecos(
    current_user: dict = Depends(require_admin)
):
    """
    Atualiza todos os endereços de entrega das OCs adicionando o CEP automaticamente
    onde ainda não existe.
    """
    
    # Buscar todas as OCs que têm endereço mas não têm CEP
    ocs_cursor = db.purchase_orders.find(
        {"endereco_entrega": {"$exists": True, "$ne": None, "$ne": ""}},
        {"_id": 0, "id": 1, "numero_oc": 1, "endereco_entrega": 1}
    )
    
    atualizados = 0
    erros = 0
    detalhes = []
    
    async for oc in ocs_cursor:
        endereco = oc.get('endereco_entrega', '')
        
        # Verificar se já tem CEP
        if re.search(r'CEP[:\s]*\d{5}-?\d{3}', endereco, re.IGNORECASE):
            continue
        
        # Buscar CEP
        cep = buscar_cep_por_endereco(endereco)
        if cep:
            novo_endereco = f"{endereco}, CEP: {cep}"
            await db.purchase_orders.update_one(
                {"id": oc['id']},
                {"$set": {"endereco_entrega": novo_endereco}}
            )
            atualizados += 1
            detalhes.append({
                "numero_oc": oc['numero_oc'],
                "cep_adicionado": cep
            })
        else:
            erros += 1
            detalhes.append({
                "numero_oc": oc['numero_oc'],
                "erro": "CEP não encontrado"
            })
    
    return {
        "success": True,
        "atualizados": atualizados,
        "erros": erros,
        "detalhes": detalhes[:20]  # Limitar a 20 para não poluir resposta
    }


@api_router.post("/admin/atualizar-todas-ocs-pdf")
async def atualizar_todas_ocs_com_pdfs(
    files: List[UploadFile] = File(...),
    current_user: dict = Depends(require_admin)
):
    """
    Atualizar múltiplas OCs de uma vez com seus PDFs.
    Cada PDF deve corresponder a uma OC existente (pelo número da OC no PDF).
    
    Preserva todos os dados dos itens, só atualiza campos vazios do cabeçalho.
    """
    
    resultados = []
    
    for file in files:
        if not file.filename.lower().endswith('.pdf'):
            resultados.append({
                "arquivo": file.filename,
                "success": False,
                "erro": "Arquivo não é PDF"
            })
            continue
        
        content = await file.read()
        
        try:
            pdf_doc = fitz.open(stream=content, filetype="pdf")
            full_text = ""
            for page in pdf_doc:
                full_text += page.get_text()
            pdf_doc.close()
        except Exception as e:
            resultados.append({
                "arquivo": file.filename,
                "success": False,
                "erro": f"Erro ao ler PDF: {str(e)}"
            })
            continue
        
        # Extrair número da OC
        oc_patterns = [
            r'(?:OC|Ordem de Compra)[:\s\-]*(\d+[\.\-]?\d+)',
            r'(?:N[úu]mero|N[°º])[:\s]*(\d+[\.\-]?\d+)',
            r'(\d{1,2}\.\d{6,})'
        ]
        
        numero_oc = None
        for pattern in oc_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                numero_oc = f"OC-{match.group(1)}"
                break
        
        if not numero_oc:
            resultados.append({
                "arquivo": file.filename,
                "success": False,
                "erro": "Não foi possível identificar o número da OC no PDF"
            })
            continue
        
        # Buscar OC existente
        existing_po = await db.purchase_orders.find_one({"numero_oc": numero_oc}, {"_id": 0})
        if not existing_po:
            resultados.append({
                "arquivo": file.filename,
                "numero_oc": numero_oc,
                "success": False,
                "erro": f"OC {numero_oc} não encontrada no sistema"
            })
            continue
        
        # Extrair dados do PDF
        # Endereço
        endereco_patterns = [
            r'Endere[çc]o de Entrega[:\s]*(.*?)(?:\n\n|Linha|Item)',
            r'Local de Entrega[:\s]*(.*?)(?:\n\n|Linha|Item)',
            r'Entregar em[:\s]*(.*?)(?:\n\n|Linha|Item)'
        ]
        
        novo_endereco = ""
        for pattern in endereco_patterns:
            endereco_match = re.search(pattern, full_text, re.IGNORECASE | re.DOTALL)
            if endereco_match:
                novo_endereco = endereco_match.group(1).strip()
                novo_endereco = ' '.join(novo_endereco.split()).upper()
                break
        
        # Data de Entrega
        nova_data_entrega = None
        data_patterns = [
            r'Data de Entrega[:\s]*(\d{2}/\d{2}/\d{4})',
            r'Data Entrega[:\s]*(\d{2}/\d{2}/\d{4})',
            r'Entrega[:\s]*(\d{2}/\d{2}/\d{4})',
            r'Prazo de Entrega[:\s]*(\d{2}/\d{2}/\d{4})',
            r'Dt\.\s*Entrega[:\s]*(\d{2}/\d{2}/\d{4})'
        ]
        
        for pattern in data_patterns:
            data_match = re.search(pattern, full_text, re.IGNORECASE)
            if data_match:
                try:
                    data_str = data_match.group(1)
                    dia, mes, ano = data_str.split('/')
                    nova_data_entrega = f"{ano}-{mes}-{dia}"
                    break
                except:
                    pass
        
        if not nova_data_entrega:
            date_after_req = re.findall(r'\d{1,2}\.\d{2}\.\d{6,}\s*(\d{2}/\d{2}/\d{4})', full_text)
            if date_after_req:
                try:
                    data_str = date_after_req[0]
                    dia, mes, ano = data_str.split('/')
                    nova_data_entrega = f"{ano}-{mes}-{dia}"
                except:
                    pass
        
        # ========== EXTRAIR NCM POR ITEM ==========
        ncm_por_item = {}
        lines = full_text.split('\n')
        for i, line in enumerate(lines):
            line_stripped = line.strip()
            
            # Procurar código de 6 dígitos que começa com 0 ou 1
            if re.match(r'^([01]\d{5})$', line_stripped):
                codigo = line_stripped
                
                # Procurar NCM nas próximas linhas
                for j in range(i+1, min(i+40, len(lines))):
                    check_line = lines[j].strip()
                    
                    # Se encontrar outro código de produto, parar
                    if re.match(r'^([01]\d{5})$', check_line):
                        break
                    
                    # Opção 1: NCM completo de 8 dígitos (pode começar com qualquer dígito 1-9)
                    if re.match(r'^[1-9]\d{7}$', check_line):
                        ncm_por_item[codigo] = check_line
                        break
                    
                    # Opção 2: NCM dividido em duas linhas (6 dígitos + 2 dígitos)
                    # NCM pode começar com qualquer dígito 1-9
                    if re.match(r'^[1-9]\d{5}$', check_line):
                        if j+1 < len(lines):
                            next_line = lines[j+1].strip()
                            if re.match(r'^\d{2}$', next_line):
                                ncm_completo = check_line + next_line
                                ncm_por_item[codigo] = ncm_completo
                                break
        
        # Preparar atualizações
        updates = {}
        campos_atualizados = []
        
        # Atualizar endereço - SEMPRE atualiza se o PDF tiver o endereço
        endereco_atual = existing_po.get('endereco_entrega', '').strip()
        if novo_endereco:
            # Adicionar CEP ao endereço se não tiver
            if not re.search(r'CEP[:\s]*\d{5}-?\d{3}', novo_endereco, re.IGNORECASE):
                cep = buscar_cep_por_endereco(novo_endereco)
                if cep:
                    novo_endereco = f"{novo_endereco}, CEP: {cep}"
            
            if novo_endereco != endereco_atual:
                updates['endereco_entrega'] = novo_endereco
                campos_atualizados.append(f"endereco: {novo_endereco[:50]}...")
        
        # Atualizar data de entrega - SEMPRE atualiza se o PDF tiver a data
        data_atual = existing_po.get('data_entrega', '').strip() if existing_po.get('data_entrega') else ''
        if nova_data_entrega and nova_data_entrega != data_atual:
            updates['data_entrega'] = nova_data_entrega
            campos_atualizados.append(f"data_entrega: {nova_data_entrega}")
        
        # Atualizar NCM dos itens
        ncm_atualizados = 0
        if ncm_por_item:
            items_modified = False
            for item in existing_po.get('items', []):
                codigo_item = item.get('codigo_item', '')
                if codigo_item in ncm_por_item:
                    ncm_atual = item.get('ncm', '')
                    ncm_novo = ncm_por_item[codigo_item]
                    if ncm_atual != ncm_novo:
                        item['ncm'] = ncm_novo
                        ncm_atualizados += 1
                        items_modified = True
            
            if items_modified:
                updates['items'] = existing_po['items']
                campos_atualizados.append(f"NCM: {ncm_atualizados} itens")
        
        if updates:
            await db.purchase_orders.update_one(
                {"id": existing_po['id']},
                {"$set": updates}
            )
        
        resultados.append({
            "arquivo": file.filename,
            "numero_oc": numero_oc,
            "success": True,
            "campos_atualizados": campos_atualizados,
            "ncm_encontrados": len(ncm_por_item),
            "ncm_atualizados": ncm_atualizados,
            "message": f"Atualizado: {', '.join(campos_atualizados)}" if campos_atualizados else "Sem alterações"
        })
    
    atualizados = sum(1 for r in resultados if r.get('success') and r.get('campos_atualizados'))
    sem_alteracao = sum(1 for r in resultados if r.get('success') and not r.get('campos_atualizados'))
    erros = sum(1 for r in resultados if not r.get('success'))
    total_ncm = sum(r.get('ncm_atualizados', 0) for r in resultados if r.get('success'))
    
    return {
        "success": True,
        "total_arquivos": len(files),
        "atualizados": atualizados,
        "sem_alteracao": sem_alteracao,
        "erros": erros,
        "total_ncm_atualizados": total_ncm,
        "resultados": resultados
    }


@api_router.post("/admin/migrar-enderecos")
async def migrar_enderecos_itens_para_oc(current_user: dict = Depends(require_admin)):
    """
    Migrar endereços de entrega dos itens para o nível da OC.
    Para cada OC sem endereco_entrega, busca o endereço do primeiro item que tenha.
    """
    
    # Buscar todas as OCs que não têm endereco_entrega
    pos = await db.purchase_orders.find(
        {"$or": [{"endereco_entrega": ""}, {"endereco_entrega": None}, {"endereco_entrega": {"$exists": False}}]},
        {"_id": 0, "id": 1, "numero_oc": 1, "items.endereco_entrega": 1}
    ).to_list(1000)
    
    migrados = 0
    erros = []
    
    for po in pos:
        # Buscar o primeiro item que tenha endereço de entrega
        endereco = None
        for item in po.get('items', []):
            item_endereco = item.get('endereco_entrega', '').strip()
            if item_endereco:
                endereco = item_endereco
                break
        
        if endereco:
            try:
                await db.purchase_orders.update_one(
                    {"id": po['id']},
                    {"$set": {"endereco_entrega": endereco}}
                )
                migrados += 1
            except Exception as e:
                erros.append(f"{po['numero_oc']}: {str(e)}")
    
    return {
        "success": True,
        "total_ocs_processadas": len(pos),
        "enderecos_migrados": migrados,
        "erros": erros
    }


@api_router.post("/admin/recalcular-lucros")
async def recalcular_lucros_todos_itens(current_user: dict = Depends(require_admin)):
    """
    Endpoint de migração para recalcular o lucro de todos os itens.
    Usa a nova fórmula que considera apenas a quantidade necessária, não a quantidade comprada.
    """
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(5000)
    
    itens_recalculados = 0
    ocs_atualizadas = 0
    
    for po in pos:
        items_modified = False
        for item in po.get('items', []):
            old_lucro = item.get('lucro_liquido')
            calcular_lucro_item(item)
            new_lucro = item.get('lucro_liquido')
            
            if old_lucro != new_lucro:
                itens_recalculados += 1
                items_modified = True
        
        if items_modified:
            await db.purchase_orders.update_one(
                {"id": po['id']},
                {"$set": {"items": po['items']}}
            )
            ocs_atualizadas += 1
    
    return {
        "success": True,
        "itens_recalculados": itens_recalculados,
        "ocs_atualizadas": ocs_atualizadas
    }


@api_router.post("/admin/atualizar-ncm-em-massa")
async def atualizar_ncm_em_massa(
    files: List[UploadFile] = File(...),
    current_user: dict = Depends(require_admin)
):
    """
    Atualizar o NCM de todos os itens a partir dos PDFs das OCs.
    Cada PDF será parseado para extrair os NCMs dos itens.
    Os itens são atualizados pelo código_item.
    """
    
    resultados = []
    total_ncm_atualizados = 0
    
    for file in files:
        if not file.filename.lower().endswith('.pdf'):
            resultados.append({
                "arquivo": file.filename,
                "success": False,
                "erro": "Arquivo não é PDF"
            })
            continue
        
        content = await file.read()
        
        try:
            pdf_doc = fitz.open(stream=content, filetype="pdf")
            full_text = ""
            for page in pdf_doc:
                full_text += page.get_text()
            pdf_doc.close()
        except Exception as e:
            resultados.append({
                "arquivo": file.filename,
                "success": False,
                "erro": f"Erro ao ler PDF: {str(e)}"
            })
            continue
        
        # Extrair número da OC
        oc_patterns = [
            r'(?:OC|Ordem de Compra)[:\s\-]*(\d+[\.\-]?\d+)',
            r'(?:N[úu]mero|N[°º])[:\s]*(\d+[\.\-]?\d+)',
            r'(\d{1,2}\.\d{6,})'
        ]
        
        numero_oc = None
        for pattern in oc_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                numero_oc = f"OC-{match.group(1)}"
                break
        
        if not numero_oc:
            resultados.append({
                "arquivo": file.filename,
                "success": False,
                "erro": "Não foi possível identificar o número da OC no PDF"
            })
            continue
        
        # Extrair NCM por item do PDF
        # Criar mapeamento codigo_item -> ncm
        ncm_por_item = {}
        lines = full_text.split('\n')
        
        for i, line in enumerate(lines):
            line_stripped = line.strip()
            
            # Procurar código de 6 dígitos que começa com 0 ou 1
            if re.match(r'^([01]\d{5})$', line_stripped):
                codigo = line_stripped
                
                # Procurar NCM nas próximas linhas
                for j in range(i+1, min(i+40, len(lines))):
                    check_line = lines[j].strip()
                    
                    # Se encontrar outro código de produto, parar
                    if re.match(r'^([01]\d{5})$', check_line):
                        break
                    
                    # Opção 1: NCM completo de 8 dígitos (começando com 8 ou 9)
                    if re.match(r'^[89]\d{7}$', check_line):
                        ncm_por_item[codigo] = check_line
                        break
                    
                    # Opção 2: NCM dividido em duas linhas (6 dígitos + 2 dígitos)
                    # Ex: linha "903033" seguida de "29" = "90303329"
                    if re.match(r'^[89]\d{5}$', check_line):
                        # Verificar se próxima linha tem 2 dígitos
                        if j+1 < len(lines):
                            next_line = lines[j+1].strip()
                            if re.match(r'^\d{2}$', next_line):
                                ncm_completo = check_line + next_line
                                ncm_por_item[codigo] = ncm_completo
                                break
        
        if not ncm_por_item:
            resultados.append({
                "arquivo": file.filename,
                "numero_oc": numero_oc,
                "success": False,
                "erro": "Nenhum NCM encontrado no PDF"
            })
            continue
        
        # Buscar OC existente
        existing_po = await db.purchase_orders.find_one({"numero_oc": numero_oc}, {"_id": 0})
        if not existing_po:
            resultados.append({
                "arquivo": file.filename,
                "numero_oc": numero_oc,
                "success": False,
                "erro": f"OC {numero_oc} não encontrada no sistema"
            })
            continue
        
        # Atualizar NCMs nos itens
        items_atualizados = 0
        for item in existing_po.get('items', []):
            codigo_item = item.get('codigo_item', '')
            if codigo_item in ncm_por_item:
                ncm_atual = item.get('ncm', '')
                ncm_novo = ncm_por_item[codigo_item]
                
                if ncm_atual != ncm_novo:
                    item['ncm'] = ncm_novo
                    items_atualizados += 1
        
        if items_atualizados > 0:
            await db.purchase_orders.update_one(
                {"id": existing_po['id']},
                {"$set": {"items": existing_po['items']}}
            )
            total_ncm_atualizados += items_atualizados
        
        resultados.append({
            "arquivo": file.filename,
            "numero_oc": numero_oc,
            "success": True,
            "ncm_encontrados": len(ncm_por_item),
            "items_atualizados": items_atualizados,
            "ncm_por_item": ncm_por_item
        })
    
    return {
        "success": True,
        "total_arquivos": len(files),
        "total_ncm_atualizados": total_ncm_atualizados,
        "resultados": resultados
    }


# ============== ESTOQUE ==============
@api_router.get("/estoque")
async def listar_estoque(current_user: dict = Depends(get_current_user)):
    """
    Lista todos os itens em estoque (quantidade comprada > quantidade necessária).
    Agrupa por código do item e mostra o total disponível em estoque.
    
    A quantidade comprada pode vir de:
    1. Campo quantidade_comprada do item (se preenchido)
    2. Soma das quantidades das fontes de compra
    """
    
    # Buscar todos os itens de todas as OCs
    pos = await db.purchase_orders.find(
        {},
        {"_id": 0, "id": 1, "numero_oc": 1, "items": 1}
    ).to_list(5000)
    
    # Mapa de estoque: codigo_item -> {info do estoque}
    estoque_map = {}
    
    for po in pos:
        for idx, item in enumerate(po.get('items', [])):
            quantidade_necessaria = item.get('quantidade', 0)
            status = item.get('status', 'pendente')
            
            # Só considerar itens já comprados (comprado ou posterior)
            if status not in ['comprado', 'em_separacao', 'em_transito', 'entregue']:
                continue
            
            # IMPORTANTE: Itens que foram atendidos pelo estoque não geram novo excedente
            # Eles apenas consumiram do estoque de outro item
            if item.get('atendido_por_estoque'):
                continue
            
            # Calcular quantidade comprada:
            # Prioridade: somar quantidades das fontes de compra (mais preciso)
            # Se não tem fontes, usar o campo quantidade_comprada
            fontes = item.get('fontes_compra', [])
            if fontes:
                quantidade_comprada = sum(f.get('quantidade', 0) for f in fontes)
            else:
                quantidade_comprada = item.get('quantidade_comprada', 0)
            
            # Quantidade já usada do estoque
            quantidade_usada_estoque = item.get('quantidade_usada_estoque', 0)
            
            # Se a quantidade comprada é maior que a necessária + já usada, tem excedente
            if quantidade_comprada and quantidade_comprada > (quantidade_necessaria + quantidade_usada_estoque):
                excedente = quantidade_comprada - quantidade_necessaria - quantidade_usada_estoque
                codigo_item = item.get('codigo_item', '')
                
                # Pegar informações da fonte de compra (link, fornecedor)
                link_compra = fontes[0].get('link', '') if fontes else ''
                fornecedor = fontes[0].get('fornecedor', '') if fontes else ''
                preco_unitario = fontes[0].get('preco_unitario', 0) if fontes else item.get('preco_compra', 0)
                
                # Info de onde o estoque foi usado
                estoque_usado_em = item.get('estoque_usado_em', [])
                
                if codigo_item not in estoque_map:
                    estoque_map[codigo_item] = {
                        'codigo_item': codigo_item,
                        'descricao': item.get('descricao', ''),
                        'marca_modelo': item.get('marca_modelo', ''),
                        'unidade': item.get('unidade', 'UN'),
                        'quantidade_estoque': excedente,
                        'link_compra': link_compra,
                        'fornecedor': fornecedor,
                        'preco_unitario': preco_unitario,
                        'imagem_url': item.get('imagem_url'),  # URL da imagem do item
                        'ocs_origem': [{
                            'numero_oc': po.get('numero_oc'),
                            'po_id': po.get('id'),
                            'item_index': idx,  # Índice do item na OC
                            'quantidade_comprada': quantidade_comprada,
                            'quantidade_necessaria': quantidade_necessaria,
                            'quantidade_usada_estoque': quantidade_usada_estoque,
                            'excedente': excedente,
                            'data_compra': item.get('data_compra'),
                            'usado_em': estoque_usado_em  # Para quais OCs foi usado
                        }]
                    }
                else:
                    # Soma ao estoque existente
                    estoque_map[codigo_item]['quantidade_estoque'] += excedente
                    estoque_map[codigo_item]['ocs_origem'].append({
                        'numero_oc': po.get('numero_oc'),
                        'po_id': po.get('id'),
                        'item_index': idx,  # Índice do item na OC
                        'quantidade_comprada': quantidade_comprada,
                        'quantidade_necessaria': quantidade_necessaria,
                        'quantidade_usada_estoque': quantidade_usada_estoque,
                        'excedente': excedente,
                        'data_compra': item.get('data_compra'),
                        'usado_em': estoque_usado_em  # Para quais OCs foi usado
                    })
                    # Atualiza link/fornecedor/imagem se não tinha
                    if not estoque_map[codigo_item]['link_compra'] and link_compra:
                        estoque_map[codigo_item]['link_compra'] = link_compra
                        estoque_map[codigo_item]['fornecedor'] = fornecedor
                    if not estoque_map[codigo_item].get('imagem_url') and item.get('imagem_url'):
                        estoque_map[codigo_item]['imagem_url'] = item.get('imagem_url')
    
    # Converter para lista
    estoque_list = list(estoque_map.values())
    
    # Ordenar por quantidade em estoque (maior primeiro)
    estoque_list.sort(key=lambda x: x['quantidade_estoque'], reverse=True)
    
    return {
        "total_itens_diferentes": len(estoque_list),
        "estoque": estoque_list
    }


@api_router.patch("/estoque/ajustar")
async def ajustar_estoque(
    data: dict,
    current_user: dict = Depends(require_admin)
):
    """
    Ajusta manualmente a quantidade de estoque de um item em uma OC específica.
    Usado para corrigir erros de contagem.
    
    Body:
    {
        "po_id": "...",
        "item_index": 0,
        "nova_quantidade_comprada": 15  // Nova quantidade total comprada
    }
    """
    po_id = data.get('po_id')
    item_index = data.get('item_index')
    nova_qtd = data.get('nova_quantidade_comprada')
    
    if not po_id or item_index is None or nova_qtd is None:
        raise HTTPException(status_code=400, detail="Dados inválidos")
    
    # Buscar OC
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    
    if item_index < 0 or item_index >= len(po.get('items', [])):
        raise HTTPException(status_code=400, detail="Índice de item inválido")
    
    item = po['items'][item_index]
    
    # Atualizar quantidade comprada
    item['quantidade_comprada'] = nova_qtd
    
    # Se tem fontes de compra, ajustar a primeira fonte
    fontes = item.get('fontes_compra', [])
    if fontes and len(fontes) > 0:
        fontes[0]['quantidade'] = nova_qtd
    
    # Salvar
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": po['items']}}
    )
    
    return {
        "success": True,
        "mensagem": f"Quantidade de estoque ajustada para {nova_qtd}"
    }


@api_router.delete("/estoque/limpar/{po_id}/{item_index}")
async def limpar_estoque(
    po_id: str,
    item_index: int,
    current_user: dict = Depends(require_admin)
):
    """
    Remove o excedente de estoque de um item específico.
    Define a quantidade comprada igual à quantidade necessária.
    """
    # Buscar OC
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    
    if item_index < 0 or item_index >= len(po.get('items', [])):
        raise HTTPException(status_code=400, detail="Índice de item inválido")
    
    item = po['items'][item_index]
    quantidade_necessaria = item.get('quantidade', 0)
    
    # Definir quantidade comprada igual à necessária (remove excedente)
    item['quantidade_comprada'] = quantidade_necessaria
    item['quantidade_usada_estoque'] = 0  # Resetar uso
    item['estoque_usado_em'] = []  # Limpar histórico de uso
    
    # Se tem fontes de compra, ajustar
    fontes = item.get('fontes_compra', [])
    if fontes and len(fontes) > 0:
        fontes[0]['quantidade'] = quantidade_necessaria
        # Remover fontes extras se houver
        item['fontes_compra'] = [fontes[0]]
    
    # Salvar
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": po['items']}}
    )
    
    return {
        "success": True,
        "mensagem": f"Estoque limpo. Quantidade definida para {quantidade_necessaria}"
    }


@api_router.post("/estoque/resetar-uso/{po_id}/{item_index}")
async def resetar_uso_estoque(
    po_id: str,
    item_index: int,
    current_user: dict = Depends(require_admin)
):
    """
    Reseta o histórico de uso do estoque de um item.
    Limpa quantidade_usada_estoque e estoque_usado_em.
    """
    # Buscar OC
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    
    if item_index < 0 or item_index >= len(po.get('items', [])):
        raise HTTPException(status_code=400, detail="Índice de item inválido")
    
    item = po['items'][item_index]
    
    # Resetar campos de uso
    item['quantidade_usada_estoque'] = 0
    item['estoque_usado_em'] = []
    
    # Salvar
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": po['items']}}
    )
    
    return {
        "success": True,
        "mensagem": "Histórico de uso do estoque resetado"
    }


@api_router.post("/admin/limpar-dados-estoque-inconsistentes")
async def limpar_dados_estoque_inconsistentes(
    current_user: dict = Depends(require_admin)
):
    """
    Limpa dados de estoque inconsistentes em itens que estão em status pendente/cotado
    mas ainda têm dados de uso de estoque (atendido_por_estoque, estoque_origem, etc).
    
    Esta é uma operação de migração/correção para dados legados.
    """
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(5000)
    
    itens_corrigidos = []
    status_antes_compra = ['pendente', 'cotado']
    
    for po in pos:
        po_modificado = False
        
        for idx, item in enumerate(po.get('items', [])):
            status = item.get('status', 'pendente')
            
            # Se o item está pendente/cotado mas tem dados de estoque, limpar
            if status in status_antes_compra:
                teve_correcao = False
                
                # Verificar se tem dados de estoque a limpar
                if item.get('atendido_por_estoque'):
                    item['atendido_por_estoque'] = False
                    teve_correcao = True
                
                if item.get('quantidade_do_estoque'):
                    item['quantidade_do_estoque'] = 0
                    teve_correcao = True
                
                if item.get('estoque_origem'):
                    # Reverter o uso nas OCs de origem
                    for fonte in item.get('estoque_origem', []):
                        numero_oc_origem = fonte.get('numero_oc')
                        quantidade_usada = fonte.get('quantidade', 0)
                        
                        if not numero_oc_origem or quantidade_usada <= 0:
                            continue
                        
                        # Encontrar a OC de origem
                        po_origem = await db.purchase_orders.find_one(
                            {"numero_oc": numero_oc_origem},
                            {"_id": 0}
                        )
                        
                        if not po_origem:
                            continue
                        
                        codigo_item = item.get('codigo_item')
                        
                        for item_origem in po_origem.get('items', []):
                            if item_origem.get('codigo_item') != codigo_item:
                                continue
                            
                            # Decrementar quantidade_usada_estoque
                            qtd_usada_atual = item_origem.get('quantidade_usada_estoque', 0)
                            nova_qtd_usada = max(0, qtd_usada_atual - quantidade_usada)
                            item_origem['quantidade_usada_estoque'] = nova_qtd_usada
                            
                            # Remover entrada em estoque_usado_em
                            estoque_usado_em = item_origem.get('estoque_usado_em', [])
                            item_origem['estoque_usado_em'] = [
                                uso for uso in estoque_usado_em 
                                if uso.get('po_id') != po.get('id')
                            ]
                            break
                        
                        # Salvar OC de origem
                        await db.purchase_orders.update_one(
                            {"id": po_origem['id']},
                            {"$set": {"items": po_origem['items']}}
                        )
                    
                    item['estoque_origem'] = []
                    teve_correcao = True
                
                if item.get('parcialmente_atendido_estoque'):
                    item['parcialmente_atendido_estoque'] = False
                    teve_correcao = True
                
                if item.get('preco_estoque_unitario'):
                    item['preco_estoque_unitario'] = None
                    teve_correcao = True
                
                # Remover fonte de compra "ESTOQUE INTERNO" se existir
                if item.get('fontes_compra'):
                    fontes_filtradas = [
                        fc for fc in item['fontes_compra'] 
                        if fc.get('fornecedor') != 'ESTOQUE INTERNO'
                    ]
                    if len(fontes_filtradas) != len(item['fontes_compra']):
                        item['fontes_compra'] = fontes_filtradas
                        teve_correcao = True
                
                if teve_correcao:
                    itens_corrigidos.append({
                        'numero_oc': po.get('numero_oc'),
                        'item_index': idx,
                        'codigo_item': item.get('codigo_item'),
                        'status': status
                    })
                    po_modificado = True
        
        if po_modificado:
            await db.purchase_orders.update_one(
                {"id": po['id']},
                {"$set": {"items": po['items']}}
            )
    
    return {
        "success": True,
        "itens_corrigidos": len(itens_corrigidos),
        "detalhes": itens_corrigidos
    }


@api_router.get("/estoque/mapa")
async def get_estoque_mapa(current_user: dict = Depends(get_current_user)):
    """
    Retorna um mapa simplificado de código do item -> quantidade em estoque.
    Usado pelo frontend para mostrar indicadores de estoque disponível em itens pendentes/cotados.
    """
    # Buscar todos os itens de todas as OCs
    pos = await db.purchase_orders.find(
        {},
        {"_id": 0, "items": 1}
    ).to_list(5000)
    
    # Mapa de estoque: codigo_item -> quantidade_estoque
    estoque_mapa = {}
    
    for po in pos:
        for item in po.get('items', []):
            quantidade_necessaria = item.get('quantidade', 0)
            status = item.get('status', 'pendente')
            
            # Só considerar itens já comprados (comprado ou posterior)
            if status not in ['comprado', 'em_separacao', 'em_transito', 'entregue']:
                continue
            
            # IMPORTANTE: Itens que foram atendidos pelo estoque não geram novo excedente
            if item.get('atendido_por_estoque'):
                continue
            
            # Calcular quantidade comprada: priorizar fontes de compra
            fontes = item.get('fontes_compra', [])
            if fontes:
                quantidade_comprada = sum(f.get('quantidade', 0) for f in fontes)
            else:
                quantidade_comprada = item.get('quantidade_comprada', 0)
            
            # Quantidade já usada do estoque
            quantidade_usada_estoque = item.get('quantidade_usada_estoque', 0)
            
            # Se a quantidade comprada é maior que a necessária + já usada, tem excedente
            if quantidade_comprada and quantidade_comprada > (quantidade_necessaria + quantidade_usada_estoque):
                excedente = quantidade_comprada - quantidade_necessaria - quantidade_usada_estoque
                codigo_item = item.get('codigo_item', '')
                
                if codigo_item:
                    if codigo_item not in estoque_mapa:
                        estoque_mapa[codigo_item] = 0
                    estoque_mapa[codigo_item] += excedente
    
    return estoque_mapa


@api_router.get("/estoque/verificar/{codigo_item}")
async def verificar_estoque_item(codigo_item: str, current_user: dict = Depends(get_current_user)):
    """
    Verifica se um item específico tem quantidade em estoque.
    Usado para mostrar no item pendente/cotado se há estoque disponível.
    """
    
    # Buscar todos os itens com esse código
    pos = await db.purchase_orders.find(
        {"items.codigo_item": codigo_item},
        {"_id": 0, "numero_oc": 1, "items": 1}
    ).to_list(5000)
    
    quantidade_total_estoque = 0
    detalhes = []
    
    for po in pos:
        for item in po.get('items', []):
            if item.get('codigo_item') == codigo_item:
                status = item.get('status', 'pendente')
                
                # Só considerar itens já comprados
                if status not in ['comprado', 'em_separacao', 'em_transito', 'entregue']:
                    continue
                
                # IMPORTANTE: Itens que foram atendidos pelo estoque não geram excedente
                if item.get('atendido_por_estoque'):
                    continue
                
                quantidade_necessaria = item.get('quantidade', 0)
                
                # Calcular quantidade comprada: priorizar fontes de compra
                fontes = item.get('fontes_compra', [])
                if fontes:
                    quantidade_comprada = sum(f.get('quantidade', 0) for f in fontes)
                else:
                    quantidade_comprada = item.get('quantidade_comprada', 0)
                
                # Quantidade já usada do estoque
                quantidade_usada_estoque = item.get('quantidade_usada_estoque', 0)
                
                if quantidade_comprada and quantidade_comprada > (quantidade_necessaria + quantidade_usada_estoque):
                    excedente = quantidade_comprada - quantidade_necessaria - quantidade_usada_estoque
                    quantidade_total_estoque += excedente
                    detalhes.append({
                        'numero_oc': po.get('numero_oc'),
                        'excedente': excedente
                    })
    
    return {
        "codigo_item": codigo_item,
        "em_estoque": quantidade_total_estoque > 0,
        "quantidade_disponivel": quantidade_total_estoque,
        "detalhes": detalhes
    }


# ============== PLANILHA DE ITENS CONSOLIDADA ==============
@api_router.get("/planilha-itens")
async def listar_planilha_itens(current_user: dict = Depends(get_current_user)):
    """
    Retorna uma visão consolidada de todos os itens por código.
    Agrupa por código_item e mostra:
    - Total necessário (soma de todas as OCs)
    - Total já comprado (itens com status comprado ou posterior)
    - Diferença (quanto ainda falta comprar)
    - Detalhes de cada ocorrência (lote, responsável, valor, OC)
    """
    
    # Buscar todos os itens de todas as OCs
    pos = await db.purchase_orders.find(
        {},
        {"_id": 0, "id": 1, "numero_oc": 1, "items": 1}
    ).to_list(5000)
    
    # Mapa: codigo_item -> {info consolidada}
    itens_map = {}
    status_comprado_ou_adiante = ['comprado', 'em_separacao', 'em_transito', 'entregue']
    
    for po in pos:
        for idx, item in enumerate(po.get('items', [])):
            codigo_item = item.get('codigo_item', '')
            if not codigo_item:
                continue
            
            quantidade = item.get('quantidade', 0)
            status = item.get('status', 'pendente')
            ja_comprado = status in status_comprado_ou_adiante
            
            # Pegar info das fontes de compra
            fontes = item.get('fontes_compra', [])
            preco_unitario = fontes[0].get('preco_unitario', 0) if fontes else item.get('preco_compra', 0)
            
            ocorrencia = {
                'numero_oc': po.get('numero_oc'),
                'po_id': po.get('id'),
                'lote': item.get('lote', ''),
                'responsavel': item.get('responsavel', ''),
                'marca_modelo': item.get('marca_modelo', ''),
                'preco_unitario': preco_unitario,
                'quantidade': quantidade,
                'status': status,
                'ja_comprado': ja_comprado,
                'quantidade_comprada': item.get('quantidade_comprada'),
                'data_compra': item.get('data_compra')
            }
            
            if codigo_item not in itens_map:
                itens_map[codigo_item] = {
                    'codigo_item': codigo_item,
                    'descricao': item.get('descricao', ''),
                    'unidade': item.get('unidade', 'UN'),
                    'imagem_url': item.get('imagem_url'),  # URL da imagem
                    'quantidade_total_necessaria': quantidade,
                    'quantidade_total_comprada': quantidade if ja_comprado else 0,
                    'quantidade_faltante': 0 if ja_comprado else quantidade,
                    'ocorrencias': [ocorrencia],
                    'lotes_unicos': set([item.get('lote', '')]),
                    'responsaveis_unicos': set([item.get('responsavel', '')]),
                    'marcas_unicas': set([item.get('marca_modelo', '')])
                }
            else:
                itens_map[codigo_item]['quantidade_total_necessaria'] += quantidade
                if ja_comprado:
                    itens_map[codigo_item]['quantidade_total_comprada'] += quantidade
                else:
                    itens_map[codigo_item]['quantidade_faltante'] += quantidade
                itens_map[codigo_item]['ocorrencias'].append(ocorrencia)
                itens_map[codigo_item]['lotes_unicos'].add(item.get('lote', ''))
                itens_map[codigo_item]['responsaveis_unicos'].add(item.get('responsavel', ''))
                itens_map[codigo_item]['marcas_unicas'].add(item.get('marca_modelo', ''))
                # Atualizar imagem_url se não tiver
                if not itens_map[codigo_item].get('imagem_url') and item.get('imagem_url'):
                    itens_map[codigo_item]['imagem_url'] = item.get('imagem_url')
    
    # Converter sets para listas e calcular estatísticas
    resultado = []
    for codigo, dados in itens_map.items():
        dados['lotes'] = sorted([l for l in dados['lotes_unicos'] if l])
        dados['responsaveis'] = sorted([r for r in dados['responsaveis_unicos'] if r])
        dados['marcas'] = sorted([m for m in dados['marcas_unicas'] if m])
        del dados['lotes_unicos']
        del dados['responsaveis_unicos']
        del dados['marcas_unicas']
        
        # Ordenar ocorrências por lote
        dados['ocorrencias'].sort(key=lambda x: x.get('lote', ''))
        
        resultado.append(dados)
    
    # Ordenar por quantidade faltante (maior primeiro)
    resultado.sort(key=lambda x: x['quantidade_faltante'], reverse=True)
    
    # Estatísticas gerais
    total_itens_diferentes = len(resultado)
    total_quantidade_necessaria = sum(r['quantidade_total_necessaria'] for r in resultado)
    total_quantidade_comprada = sum(r['quantidade_total_comprada'] for r in resultado)
    total_quantidade_faltante = sum(r['quantidade_faltante'] for r in resultado)
    
    return {
        "estatisticas": {
            "total_itens_diferentes": total_itens_diferentes,
            "total_quantidade_necessaria": total_quantidade_necessaria,
            "total_quantidade_comprada": total_quantidade_comprada,
            "total_quantidade_faltante": total_quantidade_faltante,
            "percentual_comprado": round((total_quantidade_comprada / total_quantidade_necessaria * 100) if total_quantidade_necessaria > 0 else 0, 1)
        },
        "itens": resultado
    }


# ============== PLANILHA DO CONTRATO (com limites importados) ==============
@api_router.get("/planilha-contrato")
async def listar_planilha_contrato(current_user: dict = Depends(get_current_user)):
    """
    Retorna uma visão consolidada de TODOS os itens do contrato FIEP.
    Usa os limites importados da planilha externa como fonte de verdade.
    
    Para cada item do contrato, cruza com os dados das OCs para mostrar:
    - Quantidade máxima do contrato (da planilha importada)
    - Quantidade já comprada (itens com status comprado ou posterior)
    - Quantidade faltante (limite do contrato - quantidade comprada)
    - Detalhes de cada ocorrência em OCs
    """
    
    # Buscar limites do contrato importados
    limites_cursor = db.limites_contrato.find({}, {"_id": 0})
    limites_list = await limites_cursor.to_list(10000)
    
    if not limites_list:
        # Se não há limites importados, retorna a planilha normal
        return await listar_planilha_itens(current_user)
    
    # Criar mapa de limites: codigo_item -> quantidade_maxima_contrato
    limites_map = {item['codigo_item']: item['quantidade_maxima_contrato'] for item in limites_list}
    
    # Buscar todos os itens de todas as OCs
    pos = await db.purchase_orders.find(
        {},
        {"_id": 0, "id": 1, "numero_oc": 1, "items": 1}
    ).to_list(5000)
    
    # Mapa: codigo_item -> {info consolidada das OCs}
    ocs_map = {}
    status_comprado_ou_adiante = ['comprado', 'em_separacao', 'em_transito', 'entregue']
    
    for po in pos:
        for idx, item in enumerate(po.get('items', [])):
            codigo_item = item.get('codigo_item', '')
            if not codigo_item:
                continue
            
            quantidade = item.get('quantidade', 0)
            status = item.get('status', 'pendente')
            ja_comprado = status in status_comprado_ou_adiante
            
            # Considerar quantidade_comprada se disponível (pode ser maior que quantidade necessária)
            quantidade_efetivamente_comprada = item.get('quantidade_comprada', quantidade) if ja_comprado else 0
            
            # Pegar info das fontes de compra
            fontes = item.get('fontes_compra', [])
            preco_unitario = fontes[0].get('preco_unitario', 0) if fontes else item.get('preco_compra', 0)
            
            ocorrencia = {
                'numero_oc': po.get('numero_oc'),
                'po_id': po.get('id'),
                'lote': item.get('lote', ''),
                'responsavel': item.get('responsavel', ''),
                'marca_modelo': item.get('marca_modelo', ''),
                'preco_unitario': preco_unitario,
                'quantidade': quantidade,
                'status': status,
                'ja_comprado': ja_comprado,
                'quantidade_comprada': item.get('quantidade_comprada'),
                'data_compra': item.get('data_compra')
            }
            
            if codigo_item not in ocs_map:
                ocs_map[codigo_item] = {
                    'descricao': item.get('descricao', ''),
                    'unidade': item.get('unidade', 'UN'),
                    'imagem_url': item.get('imagem_url'),
                    'quantidade_nas_ocs': quantidade,
                    'quantidade_comprada': quantidade_efetivamente_comprada,
                    'ocorrencias': [ocorrencia],
                    'lotes_unicos': set([item.get('lote', '')]),
                    'responsaveis_unicos': set([item.get('responsavel', '')]),
                    'marcas_unicas': set([item.get('marca_modelo', '')])
                }
            else:
                ocs_map[codigo_item]['quantidade_nas_ocs'] += quantidade
                ocs_map[codigo_item]['quantidade_comprada'] += quantidade_efetivamente_comprada
                ocs_map[codigo_item]['ocorrencias'].append(ocorrencia)
                ocs_map[codigo_item]['lotes_unicos'].add(item.get('lote', ''))
                ocs_map[codigo_item]['responsaveis_unicos'].add(item.get('responsavel', ''))
                ocs_map[codigo_item]['marcas_unicas'].add(item.get('marca_modelo', ''))
                # Atualizar imagem_url e descrição se não tiver
                if not ocs_map[codigo_item].get('imagem_url') and item.get('imagem_url'):
                    ocs_map[codigo_item]['imagem_url'] = item.get('imagem_url')
                if not ocs_map[codigo_item].get('descricao') and item.get('descricao'):
                    ocs_map[codigo_item]['descricao'] = item.get('descricao')
    
    # Construir resultado final baseado nos limites do contrato
    resultado = []
    
    for codigo_item, qtd_maxima in limites_map.items():
        oc_data = ocs_map.get(codigo_item, {})
        
        quantidade_comprada = oc_data.get('quantidade_comprada', 0)
        quantidade_faltante = max(0, qtd_maxima - quantidade_comprada)
        
        # Converter sets para listas
        lotes = sorted([l for l in oc_data.get('lotes_unicos', set()) if l]) if oc_data else []
        responsaveis = sorted([r for r in oc_data.get('responsaveis_unicos', set()) if r]) if oc_data else []
        marcas = sorted([m for m in oc_data.get('marcas_unicas', set()) if m]) if oc_data else []
        
        # Ordenar ocorrências por lote
        ocorrencias = oc_data.get('ocorrencias', [])
        ocorrencias.sort(key=lambda x: x.get('lote', ''))
        
        item_resultado = {
            'codigo_item': codigo_item,
            'descricao': oc_data.get('descricao', ''),
            'unidade': oc_data.get('unidade', 'UN'),
            'imagem_url': oc_data.get('imagem_url'),
            'quantidade_contrato': qtd_maxima,  # Do Excel importado
            'quantidade_nas_ocs': oc_data.get('quantidade_nas_ocs', 0),  # Total nas OCs
            'quantidade_comprada': quantidade_comprada,  # Já comprado
            'quantidade_faltante': quantidade_faltante,  # Falta para completar o contrato
            'tem_oc': len(ocorrencias) > 0,
            'lotes': lotes,
            'responsaveis': responsaveis,
            'marcas': marcas,
            'ocorrencias': ocorrencias
        }
        
        resultado.append(item_resultado)
    
    # Ordenar: primeiro os com OC e faltantes, depois os sem OC
    resultado.sort(key=lambda x: (
        -1 if x['tem_oc'] and x['quantidade_faltante'] > 0 else 0,  # OCs com faltante primeiro
        -x['quantidade_faltante'],  # Maior faltante primeiro
        0 if x['tem_oc'] else 1,  # Com OC antes de sem OC
        x['codigo_item']
    ))
    
    # Estatísticas gerais
    total_itens_diferentes = len(resultado)
    total_quantidade_contrato = sum(r['quantidade_contrato'] for r in resultado)
    total_quantidade_comprada = sum(r['quantidade_comprada'] for r in resultado)
    total_quantidade_faltante = sum(r['quantidade_faltante'] for r in resultado)
    itens_com_oc = sum(1 for r in resultado if r['tem_oc'])
    itens_completos = sum(1 for r in resultado if r['quantidade_faltante'] == 0)
    
    return {
        "estatisticas": {
            "total_itens_diferentes": total_itens_diferentes,
            "total_quantidade_contrato": total_quantidade_contrato,
            "total_quantidade_comprada": total_quantidade_comprada,
            "total_quantidade_faltante": total_quantidade_faltante,
            "percentual_comprado": round((total_quantidade_comprada / total_quantidade_contrato * 100) if total_quantidade_contrato > 0 else 0, 1),
            "itens_com_oc": itens_com_oc,
            "itens_completos": itens_completos
        },
        "itens": resultado,
        "fonte": "contrato_fiep"
    }


# ============== USAR ESTOQUE ==============
@api_router.post("/estoque/usar")
async def usar_estoque(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Usa itens do estoque para atender um item pendente/cotado.
    
    - Se a quantidade do estoque atende 100%: muda para "Comprado"
    - Se atende parcialmente: mantém status + marca como "parcialmente atendido"
    - Registra o preço original do estoque
    - Registra de qual OC veio o estoque
    
    Body:
    {
        "po_id": "...",
        "item_index": 0,
        "quantidade_usar": 10
    }
    """
    
    po_id = data.get('po_id')
    item_index = data.get('item_index')
    quantidade_usar = data.get('quantidade_usar', 0)
    
    if not po_id or item_index is None or quantidade_usar <= 0:
        raise HTTPException(status_code=400, detail="Dados inválidos")
    
    # Buscar OC destino
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    
    if item_index < 0 or item_index >= len(po.get('items', [])):
        raise HTTPException(status_code=400, detail="Índice de item inválido")
    
    item = po['items'][item_index]
    codigo_item = item.get('codigo_item')
    quantidade_necessaria = item.get('quantidade', 0)
    
    # Verificar estoque disponível e buscar detalhes
    estoque_response = await verificar_estoque_item(codigo_item, current_user)
    quantidade_disponivel = estoque_response.get('quantidade_disponivel', 0)
    
    if quantidade_disponivel <= 0:
        raise HTTPException(status_code=400, detail="Não há estoque disponível para este item")
    
    # Para itens pendentes/cotados, ignorar quantidade_do_estoque anterior
    # (o usuário pode ter voltado o status para testar ou corrigir)
    status_item = item.get('status', 'pendente')
    if status_item in ['pendente', 'cotado']:
        # Resetar campos de estoque anterior se o item voltou para pendente/cotado
        item['quantidade_do_estoque'] = 0
        item['estoque_origem'] = []
        item['parcialmente_atendido_estoque'] = False
        item['atendido_por_estoque'] = False
    
    # Limitar a quantidade a usar ao necessário e ao disponível
    quantidade_faltante = quantidade_necessaria - item.get('quantidade_do_estoque', 0)
    quantidade_efetiva = min(quantidade_usar, quantidade_disponivel, quantidade_faltante)
    
    if quantidade_efetiva <= 0:
        raise HTTPException(status_code=400, detail="Este item já foi totalmente atendido ou não há quantidade faltante")
    
    # Buscar informações do estoque (preço, OC de origem)
    pos_com_estoque = await db.purchase_orders.find(
        {"items.codigo_item": codigo_item},
        {"_id": 0, "id": 1, "numero_oc": 1, "items": 1}
    ).to_list(1000)
    
    # Encontrar as OCs que têm excedente deste item
    fontes_estoque = []
    for po_origem in pos_com_estoque:
        for idx, item_origem in enumerate(po_origem.get('items', [])):
            if item_origem.get('codigo_item') != codigo_item:
                continue
            
            status_origem = item_origem.get('status', 'pendente')
            if status_origem not in ['comprado', 'em_separacao', 'em_transito', 'entregue']:
                continue
            
            # IMPORTANTE: Itens que foram atendidos pelo estoque não geram excedente
            if item_origem.get('atendido_por_estoque'):
                continue
            
            # Calcular quantidade comprada: PRIORIZAR fontes_compra (mais preciso)
            fontes = item_origem.get('fontes_compra', [])
            if fontes:
                qtd_comprada = sum(f.get('quantidade', 0) for f in fontes)
            else:
                qtd_comprada = item_origem.get('quantidade_comprada', 0)
            
            qtd_necessaria_origem = item_origem.get('quantidade', 0)
            
            # Já usado anteriormente
            qtd_ja_usada = item_origem.get('quantidade_usada_estoque', 0)
            
            # Excedente disponível
            excedente = qtd_comprada - qtd_necessaria_origem - qtd_ja_usada
            
            if excedente > 0:
                # Pegar preço unitário
                if fontes:
                    preco_unitario = fontes[0].get('preco_unitario', 0)
                    frete_unitario = fontes[0].get('frete', 0)
                    link_compra = fontes[0].get('link', '')
                    fornecedor = fontes[0].get('fornecedor', '')
                else:
                    preco_unitario = item_origem.get('preco_compra', 0)
                    frete_unitario = item_origem.get('frete_compra', 0)
                    link_compra = item_origem.get('link_compra', '')
                    fornecedor = item_origem.get('fornecedor', '')
                
                # Capturar todos os dados do item original para copiar
                fontes_estoque.append({
                    'po_id': po_origem.get('id'),
                    'numero_oc': po_origem.get('numero_oc'),
                    'item_index': idx,
                    'excedente_disponivel': excedente,
                    'preco_unitario': preco_unitario,
                    # Dados adicionais do item original
                    'frete': frete_unitario,
                    'link': link_compra,
                    'fornecedor': fornecedor,
                    'observacao': item_origem.get('observacao', ''),
                    'imagem_url': item_origem.get('imagem_url', ''),
                    'imagem_filename': item_origem.get('imagem_filename', ''),
                    'marca_modelo': item_origem.get('marca_modelo', '')
                })
    
    if not fontes_estoque:
        raise HTTPException(status_code=400, detail="Estoque não encontrado")
    
    # Usar o estoque das fontes disponíveis
    quantidade_restante = quantidade_efetiva
    ocs_utilizadas = []
    preco_medio_ponderado = 0
    total_usado = 0
    
    for fonte in fontes_estoque:
        if quantidade_restante <= 0:
            break
        
        qtd_usar_desta_fonte = min(quantidade_restante, fonte['excedente_disponivel'])
        
        # Registrar uso na OC de origem
        po_origem = await db.purchase_orders.find_one({"id": fonte['po_id']}, {"_id": 0})
        if po_origem:
            item_origem = po_origem['items'][fonte['item_index']]
            qtd_ja_usada = item_origem.get('quantidade_usada_estoque', 0)
            item_origem['quantidade_usada_estoque'] = qtd_ja_usada + qtd_usar_desta_fonte
            
            # Registrar para quais OCs foi usado
            if 'estoque_usado_em' not in item_origem:
                item_origem['estoque_usado_em'] = []
            item_origem['estoque_usado_em'].append({
                'po_id': po_id,
                'numero_oc': po.get('numero_oc'),
                'quantidade': qtd_usar_desta_fonte,
                'data': datetime.now(timezone.utc).strftime('%Y-%m-%d')
            })
            
            await db.purchase_orders.update_one(
                {"id": fonte['po_id']},
                {"$set": {"items": po_origem['items']}}
            )
        
        ocs_utilizadas.append({
            'numero_oc': fonte['numero_oc'],
            'quantidade': qtd_usar_desta_fonte,
            'preco_unitario': fonte['preco_unitario']
        })
        
        preco_medio_ponderado += fonte['preco_unitario'] * qtd_usar_desta_fonte
        total_usado += qtd_usar_desta_fonte
        quantidade_restante -= qtd_usar_desta_fonte
    
    # Calcular preço médio
    preco_medio = preco_medio_ponderado / total_usado if total_usado > 0 else 0
    
    # Pegar dados do primeiro item de estoque usado (para copiar dados)
    dados_item_origem = fontes_estoque[0] if fontes_estoque else {}
    
    # Atualizar item destino
    quantidade_anterior_estoque = item.get('quantidade_do_estoque', 0)
    item['quantidade_do_estoque'] = quantidade_anterior_estoque + total_usado
    item['preco_estoque_unitario'] = preco_medio
    
    # ========= COPIAR DADOS DO ITEM ORIGINAL DO ESTOQUE =========
    # Copiar observação se não tiver ou estiver vazia
    if dados_item_origem.get('observacao') and not item.get('observacao'):
        item['observacao'] = dados_item_origem['observacao']
    
    # Copiar imagem se não tiver
    if dados_item_origem.get('imagem_url') and not item.get('imagem_url'):
        item['imagem_url'] = dados_item_origem['imagem_url']
        if dados_item_origem.get('imagem_filename'):
            item['imagem_filename'] = dados_item_origem['imagem_filename']
    
    # Copiar marca/modelo se não tiver
    if dados_item_origem.get('marca_modelo') and not item.get('marca_modelo'):
        item['marca_modelo'] = dados_item_origem['marca_modelo']
    
    # Registrar origem do estoque
    if 'estoque_origem' not in item:
        item['estoque_origem'] = []
    item['estoque_origem'].extend(ocs_utilizadas)
    
    # Verificar se atende toda a necessidade
    quantidade_total_atendida = item.get('quantidade_do_estoque', 0)
    
    if quantidade_total_atendida >= quantidade_necessaria:
        # Atende 100% - mudar para Comprado
        item['status'] = 'comprado'
        item['atendido_por_estoque'] = True
        atualizar_data_compra(item, 'comprado')
        
        # Adicionar fonte de compra com dados do item original do estoque
        if 'fontes_compra' not in item or not item['fontes_compra']:
            item['fontes_compra'] = []
        
        # Usar dados do item original (frete, link, fornecedor)
        item['fontes_compra'].append({
            'id': str(uuid.uuid4()),
            'quantidade': quantidade_necessaria,
            'preco_unitario': preco_medio,
            'frete': dados_item_origem.get('frete', 0),
            'link': dados_item_origem.get('link', ''),
            'fornecedor': dados_item_origem.get('fornecedor', '') or 'ESTOQUE INTERNO'
        })
        
        # Recalcular lucro
        calcular_lucro_item(item)
        mensagem = f"Item totalmente atendido pelo estoque ({total_usado} UN)"
    else:
        # Atende parcialmente
        item['parcialmente_atendido_estoque'] = True
        item['quantidade_faltante'] = quantidade_necessaria - quantidade_total_atendida
        
        # Adicionar fonte de compra parcial do estoque
        if 'fontes_compra' not in item:
            item['fontes_compra'] = []
        
        # Verificar se já existe fonte do estoque interno
        fonte_estoque_existente = next((f for f in item['fontes_compra'] if f.get('fornecedor') == 'ESTOQUE INTERNO'), None)
        if fonte_estoque_existente:
            # Atualizar quantidade da fonte existente
            fonte_estoque_existente['quantidade'] = quantidade_total_atendida
        else:
            # Adicionar nova fonte do estoque
            item['fontes_compra'].append({
                'id': str(uuid.uuid4()),
                'quantidade': total_usado,
                'preco_unitario': preco_medio,
                'frete': dados_item_origem.get('frete', 0),
                'link': dados_item_origem.get('link', ''),
                'fornecedor': 'ESTOQUE INTERNO (parcial)'
            })
        
        mensagem = f"Item parcialmente atendido pelo estoque ({total_usado} UN). Faltam {item['quantidade_faltante']} UN"
    
    # Salvar item destino
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"items": po['items']}}
    )
    
    return {
        "success": True,
        "codigo_item": codigo_item,
        "quantidade_usada": total_usado,
        "quantidade_total_atendida": quantidade_total_atendida,
        "quantidade_necessaria": quantidade_necessaria,
        "atendido_totalmente": quantidade_total_atendida >= quantidade_necessaria,
        "preco_unitario_estoque": round(preco_medio, 2),
        "ocs_origem": ocs_utilizadas,
        "mensagem": mensagem
    }


@api_router.get("/estoque/detalhes/{codigo_item}")
async def get_estoque_detalhes(codigo_item: str, current_user: dict = Depends(get_current_user)):
    """
    Retorna detalhes do estoque disponível para um item específico.
    Usado no modal de "Usar do Estoque" para mostrar de onde vem e o preço.
    """
    # Buscar todas as OCs com este item
    pos = await db.purchase_orders.find(
        {"items.codigo_item": codigo_item},
        {"_id": 0, "id": 1, "numero_oc": 1, "items": 1}
    ).to_list(1000)
    
    fontes = []
    total_disponivel = 0
    
    for po in pos:
        for idx, item in enumerate(po.get('items', [])):
            if item.get('codigo_item') != codigo_item:
                continue
            
            status = item.get('status', 'pendente')
            if status not in ['comprado', 'em_separacao', 'em_transito', 'entregue']:
                continue
            
            # IMPORTANTE: Itens que foram atendidos pelo estoque não geram excedente
            if item.get('atendido_por_estoque'):
                continue
            
            # Calcular quantidade comprada: PRIORIZAR fontes_compra (mais preciso)
            fontes_compra = item.get('fontes_compra', [])
            if fontes_compra:
                qtd_comprada = sum(f.get('quantidade', 0) for f in fontes_compra)
            else:
                qtd_comprada = item.get('quantidade_comprada', 0)
            
            qtd_necessaria = item.get('quantidade', 0)
            qtd_ja_usada = item.get('quantidade_usada_estoque', 0)
            
            excedente = qtd_comprada - qtd_necessaria - qtd_ja_usada
            
            if excedente > 0:
                # Pegar preço unitário
                if fontes_compra:
                    preco = fontes_compra[0].get('preco_unitario', 0)
                    fornecedor = fontes_compra[0].get('fornecedor', '')
                else:
                    preco = item.get('preco_compra', 0)
                    fornecedor = ''
                
                fontes.append({
                    'numero_oc': po.get('numero_oc'),
                    'quantidade_disponivel': excedente,
                    'preco_unitario': preco,
                    'fornecedor': fornecedor,
                    'data_compra': item.get('data_compra')
                })
                total_disponivel += excedente
    
    return {
        "codigo_item": codigo_item,
        "total_disponivel": total_disponivel,
        "fontes": fontes
    }


@app.on_event("startup")
async def startup_event():
    """Iniciar job de verificação de rastreios e criar índices no MongoDB"""
    global rastreio_task
    
    # Criar índices para otimizar queries
    try:
        # Índices para purchase_orders
        await db.purchase_orders.create_index("id", unique=True)
        await db.purchase_orders.create_index("numero_oc")
        await db.purchase_orders.create_index("items.status")
        await db.purchase_orders.create_index("items.responsavel")
        await db.purchase_orders.create_index("created_at")
        
        # Índices para users
        await db.users.create_index("email", unique=True)
        await db.users.create_index("owner_name")
        
        # Índices para reference_items
        await db.reference_items.create_index("codigo_item")
        await db.reference_items.create_index("descricao")
        
        # Índices para notifications
        await db.notifications.create_index("user_email")
        await db.notifications.create_index("created_at")
        
        logging.getLogger(__name__).info("Índices do MongoDB criados/verificados com sucesso")
    except Exception as e:
        logging.getLogger(__name__).warning(f"Erro ao criar índices: {e}")
    
    # Iniciar job de verificação de rastreios
    rastreio_task = asyncio.create_task(verificar_rastreios_em_transito())
    logging.getLogger(__name__).info("Job de verificação de rastreios iniciado")

# Inicializar routers modulares com database e constantes
init_admin_routes(db, LOT_ASSIGNMENTS, LOT_TO_OWNER, EXCLUDED_OCS_FROM_COMMISSION)
init_backup_routes(db)
init_fornecedores_routes(db)
init_dashboard_routes(db)
init_estoque_routes(db)
init_limites_routes(db)

# Incluir routers modulares no api_router
api_router.include_router(auth_router)
api_router.include_router(rastreio_router)
api_router.include_router(notificacao_router)
api_router.include_router(admin_router)
api_router.include_router(backup_router)
api_router.include_router(fornecedores_router)
api_router.include_router(dashboard_router)
api_router.include_router(estoque_router)
api_router.include_router(limites_router)

# Include the router in the main app
app.include_router(api_router)

# CORS Configuration - usando variável de ambiente
cors_origins_env = os.environ.get('CORS_ORIGINS', '*')
if cors_origins_env == '*':
    # Wildcard não pode ser usado com credentials
    app.add_middleware(
        CORSMiddleware,
        allow_credentials=False,
        allow_origins=["*"],
        allow_methods=["*"],
        allow_headers=["*"],
    )
else:
    # Lista específica de origens permite credentials
    cors_origins_list = [origin.strip() for origin in cors_origins_env.split(',')]
    app.add_middleware(
        CORSMiddleware,
        allow_credentials=True,
        allow_origins=cors_origins_list,
        allow_methods=["*"],
        allow_headers=["*"],
    )

# Middleware para headers anti-cache (forçar atualização)
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request

class NoCacheMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        return response

app.add_middleware(NoCacheMiddleware)

# Logger já configurado no início do arquivo

@app.on_event("shutdown")
async def shutdown_db_client():
    global rastreio_task
    if rastreio_task:
        rastreio_task.cancel()
    client.close()
