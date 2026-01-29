"""
Limites Contrato Routes - Rotas para limites contratuais
"""
from fastapi import APIRouter, HTTPException, Depends, File, UploadFile
from datetime import datetime, timezone
import logging

from auth import get_current_user, require_admin

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/limites-contrato", tags=["Limites Contrato"])

# Database reference
db = None

def init_limites_routes(database):
    """Initialize limites routes with database"""
    global db
    db = database


@router.get("")
async def listar_limites_contrato(current_user: dict = Depends(get_current_user)):
    """Lista todos os limites de contrato importados"""
    limites = await db.limites_contrato.find({}, {"_id": 0}).to_list(10000)
    return {
        "total": len(limites),
        "limites": limites
    }


@router.get("/mapa")
async def get_limites_contrato_mapa(current_user: dict = Depends(get_current_user)):
    """
    Retorna um mapa de código_item -> quantidade_maxima_contrato
    Usado pelo frontend para exibir os limites rapidamente
    """
    limites = await db.limites_contrato.find({}, {"_id": 0}).to_list(10000)
    mapa = {item['codigo_item']: item['quantidade_maxima_contrato'] for item in limites}
    return mapa


@router.post("/importar")
async def importar_limites_contrato(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_admin)
):
    """
    Importa a planilha de limites do contrato FIEP.
    A planilha deve ter:
    - Coluna J (10): Código do item
    - Coluna H (8): Quantidade máxima permitida no contrato
    """
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser Excel (.xlsx ou .xls)")
    
    try:
        import openpyxl
        from io import BytesIO
        
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        sheet = wb.active
        
        limites_por_codigo = {}
        linhas_processadas = 0
        
        for row in range(2, sheet.max_row + 1):
            codigo = sheet.cell(row=row, column=10).value  # Coluna J
            qtd_maxima = sheet.cell(row=row, column=8).value  # Coluna H
            
            if codigo and qtd_maxima:
                codigo_str = str(codigo).strip()
                try:
                    qtd = int(qtd_maxima)
                    if codigo_str not in limites_por_codigo:
                        limites_por_codigo[codigo_str] = 0
                    limites_por_codigo[codigo_str] += qtd
                    linhas_processadas += 1
                except (ValueError, TypeError):
                    continue
        
        # Limpar coleção existente e inserir novos limites
        await db.limites_contrato.delete_many({})
        
        documentos = [
            {
                'codigo_item': codigo,
                'quantidade_maxima_contrato': quantidade,
                'data_importacao': datetime.now(timezone.utc).isoformat()
            }
            for codigo, quantidade in limites_por_codigo.items()
        ]
        
        if documentos:
            await db.limites_contrato.insert_many(documentos)
        
        logger.info(f"Importados limites para {len(limites_por_codigo)} códigos de itens")
        
        return {
            "success": True,
            "itens_importados": len(limites_por_codigo),
            "linhas_processadas": linhas_processadas,
            "mensagem": f"Limites do contrato importados para {len(limites_por_codigo)} códigos de itens"
        }
        
    except Exception as e:
        logger.error(f"Erro ao importar limites: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao processar planilha: {str(e)}")
